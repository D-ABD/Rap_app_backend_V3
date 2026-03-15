# rap_app/admin/declic_admin.py
import logging
from django.contrib import admin, messages
from django.http import HttpResponse
from django.utils.html import format_html
from django.utils.timezone import localtime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.contrib.admin import SimpleListFilter

from ..models.declic import Declic, ObjectifDeclic
from ..models.centres import Centre

logger = logging.getLogger("rap_app.admin.declic")


# -------------------------------------------------------------------
# 🔎 FILTRE CP
# -------------------------------------------------------------------
class CodePostalFilter(SimpleListFilter):
    title = "Code postal du centre"
    parameter_name = "code_postal"

    def lookups(self, request, model_admin):
        codes = (
            Centre.objects.exclude(code_postal__isnull=True)
            .exclude(code_postal__exact="")
            .values_list("code_postal", flat=True)
            .distinct()
        )
        return [(c, c) for c in sorted(codes)]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(centre__code_postal=self.value())
        return queryset


# -------------------------------------------------------------------
# 📤 EXPORT ATELIERS UNIQUEMENT
# -------------------------------------------------------------------
@admin.action(description="📤 Exporter les ateliers en Excel")
def export_declic_xlsx(modeladmin, request, queryset):

    wb = Workbook()
    ws = wb.active
    ws.title = "Déclic – Ateliers"

    ws.merge_cells("A1:H1")
    ws["A1"] = "Export Déclic – Ateliers (IC supprimée)"
    ws["A1"].font = Font(bold=True, size=14, color="004C99")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])

    headers = [
        "ID", "Type atelier", "Date", "Centre",
        "Inscrits", "Présents", "Absents",
        "Taux présence (%)",
    ]
    ws.append(headers)

    fill_header = PatternFill("solid", fgColor="DCE6F1")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for cell in ws[2]:
        cell.font = Font(bold=True, color="002060")
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for obj in queryset:
        ws.append([
            obj.id,
            obj.get_type_declic_display(),
            obj.date_declic.strftime("%d/%m/%Y") if obj.date_declic else "",
            getattr(obj.centre, "nom", "—"),
            obj.nb_inscrits_declic,
            obj.nb_presents_declic,
            obj.nb_absents_declic,
            obj.taux_presence_declic,
        ])
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.font = Font(size=10)

    # Largeurs
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="declic_ateliers.xlsx"'
    return response


@admin.register(ObjectifDeclic)
class ObjectifDeclicAdmin(admin.ModelAdmin):
    """🎯 Objectifs annuels Déclic (ateliers uniquement)."""

    # 👉 Méthodes d’affichage (déclarées AVANT)
    def taux_atteinte_display(self, obj):
        return f"{obj.taux_atteinte:.1f} %" if obj.taux_atteinte else "—"
    taux_atteinte_display.short_description = "Taux atteinte"

    def reste_a_faire_display(self, obj):
        return obj.reste_a_faire or 0
    reste_a_faire_display.short_description = "Reste à faire"

    def updated_at_display(self, obj):
        return localtime(obj.updated_at).strftime("%d/%m/%Y %H:%M") if obj.updated_at else "—"
    updated_at_display.short_description = "Modifié le"

    # 👉 Maintenant on peut les utiliser :
    list_display = (
        "centre",
        "annee",
        "valeur_objectif",
        "taux_atteinte_display",
        "reste_a_faire_display",
        "updated_at_display",
    )

    list_filter = ("annee", CodePostalFilter)
    search_fields = ("centre__nom",)
    ordering = ("-annee", "centre__nom")

    readonly_fields = (
        "taux_atteinte_display",
        "reste_a_faire_display",
        "created_at",
        "updated_at",
        "created_by",
        "updated_by",
    )

    fieldsets = (
        (None, {
            "fields": ("centre", "annee", "valeur_objectif", "commentaire")
        }),
        ("📊 Données calculées", {
            "fields": ("taux_atteinte_display", "reste_a_faire_display"),
        }),
        ("🕒 Métadonnées", {
            "fields": ("created_at", "updated_at", "created_by", "updated_by"),
        }),
    )

# -------------------------------------------------------------------
# 📊 ADMIN DÉCLIC (ATELIERS UNIQUEMENT)
# -------------------------------------------------------------------
@admin.register(Declic)
class DeclicAdmin(admin.ModelAdmin):

    list_display = (
        "date_display",
        "type_badge",
        "centre",
        "nb_presents_declic",
        "nb_absents_declic",
        "taux_presence_declic_display",
        "updated_at_display",
    )

    list_filter = ("type_declic", CodePostalFilter, ("date_declic", admin.DateFieldListFilter))
    search_fields = ("centre__nom", "commentaire")
    ordering = ("-date_declic",)
    actions = [export_declic_xlsx]

    fieldsets = (
        ("📅 Informations générales", {
            "fields": ("type_declic", "date_declic", "centre", "commentaire")
        }),
        ("🧩 Données Ateliers Déclic", {
            "fields": ("nb_inscrits_declic", "nb_presents_declic", "nb_absents_declic")
        }),
        ("🕒 Métadonnées", {
            "fields": ("created_at", "updated_at", "created_by", "updated_by")
        }),
    )

    # Affichage
    def date_display(self, obj):
        return obj.date_declic.strftime("%d/%m/%Y") if obj.date_declic else "—"

    def taux_presence_declic_display(self, obj):
        return f"{obj.taux_presence_declic:.1f} %"

    def updated_at_display(self, obj):
        return localtime(obj.updated_at).strftime("%d/%m/%Y %H:%M") if obj.updated_at else "—"

    def type_badge(self, obj):
        return format_html(
            f'<span style="color:white; background:#6a1b9a; padding:2px 8px; border-radius:5px;">'
            f'{obj.get_type_declic_display()}</span>'
        )
