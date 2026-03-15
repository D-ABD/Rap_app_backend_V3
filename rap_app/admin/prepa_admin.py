# rap_app/admin/prepa_admin.py
import logging
from django.contrib import admin, messages
from django.http import HttpResponse
from django.utils.html import format_html
from django.utils.timezone import localtime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.contrib.admin import SimpleListFilter

from ..models.prepa import Prepa, ObjectifPrepa

logger = logging.getLogger("rap_app.admin.prepa")


# -------------------------------------------------------------------
# ðŸ”Ž FILTRE PERSONNALISÃ‰ : Code postal du centre (fallback)
# -------------------------------------------------------------------
class CodePostalFilter(SimpleListFilter):
    title = "Code postal du centre"
    parameter_name = "code_postal"

    def lookups(self, request, model_admin):
        from ..models.centres import Centre
        codes = (
            Centre.objects.exclude(code_postal__isnull=True)
            .exclude(code_postal__exact="")
            .values_list("code_postal", flat=True)
            .distinct()
        )
        return [(c, c) for c in sorted(codes)]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(centre__code_postal=self.value())
        return queryset


# -------------------------------------------------------------------
# ðŸ“¤ ACTION Dâ€™EXPORT EXCEL
# -------------------------------------------------------------------
@admin.action(description="ðŸ“¤ Exporter la sÃ©lection en Excel")
def export_prepa_xlsx(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "PrÃ©pa â€“ SÃ©ances"

    ws.merge_cells("A1:O1")
    ws["A1"] = "Export PrÃ©pa â€“ RAP_APP"
    ws["A1"].font = Font(bold=True, size=14, color="004C99")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])

    headers = [
        "ID", "Type activitÃ©", "Date", "Centre",
        "Places ouvertes", "Prescriptions", "PrÃ©sents IC", "Absents IC", "AdhÃ©sions",
        "Inscrits atelier", "PrÃ©sents atelier", "Absents atelier",
        "Taux prÃ©sence IC", "Taux adhÃ©sion", "Taux prÃ©sence atelier",
    ]
    ws.append(headers)

    fill_header = PatternFill("solid", fgColor="DCE6F1")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for cell in ws[2]:
        cell.font = Font(bold=True, color="002060")
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for i, obj in enumerate(queryset, start=1):
        ws.append([
            obj.id,
            obj.get_type_prepa_display(),
            obj.date_prepa.strftime("%d/%m/%Y") if obj.date_prepa else "",
            getattr(obj.centre, "nom", "â€”"),
            obj.nombre_places_ouvertes,
            obj.nombre_prescriptions,
            obj.nb_presents_info,
            obj.nb_absents_info,
            obj.nb_adhesions,
            obj.nb_inscrits_prepa,
            obj.nb_presents_prepa,
            obj.nb_absents_prepa,
            obj.taux_presence_info,
            obj.taux_adhesion,
            obj.taux_presence_prepa,
        ])
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.font = Font(size=10)

    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="prepa_export.xlsx"'
    return response


# -------------------------------------------------------------------
# ðŸŽ¯ OBJECTIF PRÃ‰PA ADMIN
# -------------------------------------------------------------------
@admin.register(ObjectifPrepa)
class ObjectifPrepaAdmin(admin.ModelAdmin):
    """ðŸŽ¯ Objectifs annuels PrÃ©pa."""
    list_display = (
        "centre",
        "annee",
        "valeur_objectif",
        "taux_atteinte_display",
        "reste_a_faire_display",
        "updated_at_display",
    )
    list_filter = ("annee", CodePostalFilter)
    search_fields = ("centre__nom",)
    ordering = ("-annee", "centre__nom")

    readonly_fields = (
        "taux_atteinte_display", "reste_a_faire_display",
        "created_at", "updated_at", "created_by", "updated_by"
    )

    fieldsets = (
        (None, {
            "fields": ("centre", "annee", "valeur_objectif", "commentaire")
        }),
        ("ðŸ“Š DonnÃ©es calculÃ©es", {
            "fields": ("taux_atteinte_display", "reste_a_faire_display"),
            "description": "Champs calculÃ©s automatiquement selon les sÃ©ances PrÃ©pa."
        }),
        ("ðŸ•’ MÃ©tadonnÃ©es", {
            "fields": ("created_at", "updated_at", "created_by", "updated_by"),
        }),
    )

    def taux_atteinte_display(self, obj):
        return f"{obj.taux_atteinte:.1f} %" if obj.taux_atteinte else "â€”"
    taux_atteinte_display.short_description = "Taux atteinte"

    def reste_a_faire_display(self, obj):
        return obj.reste_a_faire or 0
    reste_a_faire_display.short_description = "Reste Ã  faire"

    def updated_at_display(self, obj):
        return localtime(obj.updated_at).strftime("%d/%m/%Y %H:%M") if obj.updated_at else "â€”"
    updated_at_display.short_description = "ModifiÃ© le"


# -------------------------------------------------------------------
# ðŸ“Š PRÃ‰PA ADMIN
# -------------------------------------------------------------------
@admin.register(Prepa)
class PrepaAdmin(admin.ModelAdmin):
    """ðŸ“Š SÃ©ances PrÃ©pa."""

    list_display = (
        "date_display",
        "type_badge",
        "centre",
        "presents_display",
        "absents_display",
        "adhesions_display",
        "taux_presence_info_display",
        "taux_presence_prepa_display",
        "updated_at_display",
    )
    list_filter = ("type_prepa", CodePostalFilter, ("date_prepa", admin.DateFieldListFilter))
    search_fields = ("centre__nom", "commentaire")
    ordering = ("-date_prepa",)
    readonly_fields = ("created_at", "updated_at", "created_by", "updated_by")
    actions = [export_prepa_xlsx]

    fieldsets = (
        ("ðŸ“… Informations gÃ©nÃ©rales", {
            "fields": ("type_prepa", "date_prepa", "centre", "commentaire")
        }),
        ("ðŸ“Š DonnÃ©es Information collective", {
            "fields": (
                "nombre_places_ouvertes", "nombre_prescriptions",
                "nb_presents_info", "nb_absents_info", "nb_adhesions"
            )
        }),
        ("ðŸ§© DonnÃ©es Ateliers PrÃ©pa", {
            "fields": (
                "nb_inscrits_prepa", "nb_presents_prepa", "nb_absents_prepa"
            )
        }),
        ("ðŸ•’ MÃ©tadonnÃ©es", {
            "fields": ("created_at", "updated_at", "created_by", "updated_by")
        }),
    )

    # ðŸ§© Colonnes dynamiques
    def presents_display(self, obj):
        return obj.nb_presents_info if obj.type_prepa == "info_collective" else obj.nb_presents_prepa
    presents_display.short_description = "PrÃ©sents"

    def absents_display(self, obj):
        return obj.nb_absents_info if obj.type_prepa == "info_collective" else obj.nb_absents_prepa
    absents_display.short_description = "Absents"

    def adhesions_display(self, obj):
        return obj.nb_adhesions if obj.type_prepa == "info_collective" else "â€”"
    adhesions_display.short_description = "AdhÃ©sions"

    def date_display(self, obj):
        return obj.date_prepa.strftime("%d/%m/%Y") if obj.date_prepa else "â€”"
    date_display.short_description = "Date"

    def type_badge(self, obj):
        colors = {
            "info_collective": "#0277bd",
            "session_1": "#00695c",
            "session_2": "#2e7d32",
            "session_3": "#558b2f",
            "session_4": "#9ccc65",
            "session_5": "#7cb342",
            "session_6": "#33691e",
        }
        color = colors.get(obj.type_prepa, "#555")
        return format_html(
            f'<span style="color:white; background:{color}; padding:2px 8px; border-radius:5px;">{obj.get_type_prepa_display()}</span>'
        )
    type_badge.short_description = "Type"

    def taux_presence_info_display(self, obj):
        return f"{obj.taux_presence_info:.1f} %" if obj.taux_presence_info else "â€”"
    taux_presence_info_display.short_description = "PrÃ©sence IC"

    def taux_presence_prepa_display(self, obj):
        return f"{obj.taux_presence_prepa:.1f} %" if obj.taux_presence_prepa else "â€”"
    taux_presence_prepa_display.short_description = "PrÃ©sence PrÃ©pa"

    def updated_at_display(self, obj):
        return localtime(obj.updated_at).strftime("%d/%m/%Y %H:%M") if obj.updated_at else "â€”"
    updated_at_display.short_description = "ModifiÃ© le"
