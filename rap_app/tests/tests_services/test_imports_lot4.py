"""Tests import/export Excel Lot 4 — Candidat et CVThèque (§2.10.1)."""

import io

from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework import status

from rap_app.models.candidat import Candidat
from rap_app.models.centres import Centre
from rap_app.models.custom_user import CustomUser
from rap_app.models.cvtheque import CVTheque
from rap_app.models.formations import Formation
from rap_app.models.statut import Statut
from rap_app.models.types_offre import TypeOffre
from rap_app.services.imports.handlers_lot4 import CVThequeExcelHandler, CandidatExcelHandler
from rap_app.tests.factories import UserFactory
from rap_app.tests.tests_services.test_imports_lot1 import _http_response_bytes
from rap_app.tests.test_utils import AuthenticatedTestCase


def _cand_export_ie():
    return reverse("import-export-lot1-export-xlsx", kwargs={"resource": "candidat"})


def _cand_import_ie():
    return reverse("import-export-lot1-import-xlsx", kwargs={"resource": "candidat"})


def _cv_import_ie():
    return reverse("import-export-lot1-import-xlsx", kwargs={"resource": "cvtheque"})


def _cv_export_ie():
    return reverse("import-export-lot1-export-xlsx", kwargs={"resource": "cvtheque"})


class CandidatCVThequeExcelTests(AuthenticatedTestCase):
    def setUp(self):
        super().setUp()
        self.admin = UserFactory(role=CustomUser.ROLE_ADMIN)
        self.client.force_authenticate(user=self.admin)

    def test_meta_resource_candidat_on_template(self):
        h = CandidatExcelHandler()
        buf = io.BytesIO(h.build_template_bytes())
        wb = load_workbook(buf, read_only=True, data_only=True)
        try:
            meta = {
                str(r[0].value).strip(): r[1].value
                for r in wb["Meta"].iter_rows(min_col=1, max_col=2)
                if r[0].value
            }
            self.assertEqual(meta.get("resource"), "candidat")
        finally:
            wb.close()

    def test_candidat_export_import_round_trip_dry_run(self):
        centre = Centre.objects.create(nom="C Cand IE", code_postal="75001", created_by=self.admin)
        to = TypeOffre.objects.create(nom="crif", created_by=self.admin)
        st = Statut.objects.create(nom="non_defini", couleur="#000000", created_by=self.admin)
        formation = Formation.objects.create(
            nom="F Cand IE",
            centre=centre,
            type_offre=to,
            statut=st,
            created_by=self.admin,
        )
        cand = Candidat.objects.create(
            nom="Dupont",
            prenom="Jean",
            email="jean.dupont@example.com",
            formation=formation,
            created_by=self.admin,
        )
        export = self.client.get(_cand_export_ie())
        self.assertEqual(export.status_code, status.HTTP_200_OK)
        upload = SimpleUploadedFile(
            "c.xlsx",
            _http_response_bytes(export),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        r = self.client.post(f"{_cand_import_ie()}?dry_run=true", {"file": upload}, format="multipart")
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertTrue(r.data.get("dry_run"))
        self.assertEqual(r.data.get("resource"), "candidat")
        self.assertEqual(r.data["summary"]["failed"], 0)

    def test_staff_read_cannot_post_candidat_import(self):
        Candidat.objects.create(nom="X", prenom="Y", email="x@example.com", created_by=self.admin)
        staff_read = UserFactory(role=CustomUser.ROLE_STAFF_READ)
        self.client.force_authenticate(user=staff_read)
        h = CandidatExcelHandler()
        upload = SimpleUploadedFile(
            "t.xlsx",
            h.build_template_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        r = self.client.post(_cand_import_ie(), {"file": upload}, format="multipart")
        self.assertEqual(r.status_code, status.HTTP_403_FORBIDDEN)

    def test_cvtheque_import_returns_not_supported(self):
        h = CVThequeExcelHandler()
        upload = SimpleUploadedFile(
            "cv.xlsx",
            h.build_template_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        r = self.client.post(_cv_import_ie(), {"file": upload}, format="multipart")
        self.assertEqual(r.status_code, status.HTTP_400_BAD_REQUEST)
        code = r.data.get("code")
        if isinstance(code, list):
            code = code[0]
        self.assertEqual(code, "not_supported")

    def test_cvtheque_export_respects_candidat_query_param(self):
        """Parité liste REST / export import-export : filtre ``candidat`` (§2.9.1)."""
        centre = Centre.objects.create(nom="C CV IE", code_postal="75001", created_by=self.admin)
        to = TypeOffre.objects.create(nom="crif", created_by=self.admin)
        st = Statut.objects.create(nom="non_defini", couleur="#000000", created_by=self.admin)
        formation = Formation.objects.create(
            nom="F CV IE", centre=centre, type_offre=to, statut=st, created_by=self.admin
        )
        c1 = Candidat.objects.create(
            nom="E1",
            prenom="C",
            email="e1.cv.ie@example.com",
            formation=formation,
            created_by=self.admin,
        )
        c2 = Candidat.objects.create(
            nom="E2",
            prenom="C",
            email="e2.cv.ie@example.com",
            formation=formation,
            created_by=self.admin,
        )
        pdf1 = SimpleUploadedFile("v1.pdf", b"%PDF-1.4 v1", content_type="application/pdf")
        pdf2 = SimpleUploadedFile("v2.pdf", b"%PDF-1.4 v2", content_type="application/pdf")
        cv1 = CVTheque.objects.create(
            candidat=c1,
            document_type="CV",
            fichier=pdf1,
            titre="CV export filtre 1",
            created_by=self.admin,
        )
        cv2 = CVTheque.objects.create(
            candidat=c2,
            document_type="CV",
            fichier=pdf2,
            titre="CV export filtre 2",
            created_by=self.admin,
        )
        self.addCleanup(cv1.fichier.close)
        self.addCleanup(cv2.fichier.close)

        def _ids_from_export(resp):
            self.assertEqual(resp.status_code, status.HTTP_200_OK)
            buf = io.BytesIO(_http_response_bytes(resp))
            wb = load_workbook(buf, read_only=True, data_only=True)
            try:
                ws = wb["Données"]
                out = []
                for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                    v = row[0]
                    if v is not None:
                        out.append(int(v))
                return out
            finally:
                wb.close()

        all_ids = set(_ids_from_export(self.client.get(_cv_export_ie())))
        self.assertEqual(all_ids, {cv1.id, cv2.id})

        c1_ids = set(_ids_from_export(self.client.get(_cv_export_ie(), {"candidat": c1.id})))
        self.assertEqual(c1_ids, {cv1.id})
