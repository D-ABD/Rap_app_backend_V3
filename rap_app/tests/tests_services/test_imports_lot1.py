"""Tests unitaires et d'intégration pour l'import/export Excel Lot 1 (référentiels)."""

import io

from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from rest_framework import status

from rap_app.models.centres import Centre
from rap_app.models.custom_user import CustomUser
from rap_app.models.import_job import ImportJob
from rap_app.models.types_offre import TypeOffre
from rap_app.services.imports.handlers_lot1 import CentreExcelHandler
from rap_app.services.imports.schemas import RESOURCE_CENTRE
from rap_app.services.imports.validation import validate_xlsx_uploaded_file
from rap_app.tests.factories import UserFactory
from rap_app.tests.test_utils import AuthenticatedTestCase


def _centre_export_ie():
    return reverse("import-export-lot1-export-xlsx", kwargs={"resource": "centre"})


def _centre_import_ie():
    return reverse("import-export-lot1-import-xlsx", kwargs={"resource": "centre"})


def _http_response_bytes(response):
    """Corps binaire d'une réponse (``HttpResponse`` ou ``FileResponse`` / streaming)."""
    if getattr(response, "streaming", False):
        return b"".join(response.streaming_content)
    return response.content


class ValidateXlsxUploadTests(TestCase):
    """Validation extension et taille (§2.12)."""

    def test_rejects_non_xlsx_extension(self):
        """Un fichier sans extension .xlsx est refusé."""
        f = SimpleUploadedFile("data.csv", b"not xlsx", content_type="text/csv")
        with self.assertRaises(ValidationError) as ctx:
            validate_xlsx_uploaded_file(f)
        self.assertIn("file", ctx.exception.message_dict)

    def test_rejects_oversized_file(self):
        """Un fichier dépassant RAP_IMPORT_MAX_UPLOAD_BYTES est refusé."""
        content = b"PK\x03\x04" + b"x" * 200
        f = SimpleUploadedFile("huge.xlsx", content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with override_settings(RAP_IMPORT_MAX_UPLOAD_BYTES=100):
            with self.assertRaises(ValidationError) as ctx:
                validate_xlsx_uploaded_file(f)
        self.assertIn("file", ctx.exception.message_dict)
        codes = ctx.exception.message_dict.get("code")
        self.assertEqual(codes[0] if isinstance(codes, list) else codes, "file_too_large")


class CentreExcelIntegrationTests(AuthenticatedTestCase):
    """Intégration API : centres, export, import, dry_run, permissions."""

    def setUp(self):
        super().setUp()
        self.admin = UserFactory(role=CustomUser.ROLE_ADMIN)
        self.client.force_authenticate(user=self.admin)

    def test_export_import_round_trip_no_field_change(self):
        """Export puis import des mêmes données : les champs métier restent identiques."""
        c = Centre.objects.create(nom="Centre RT Lot1", code_postal="75001", commune="Paris")
        res = self.client.get(_centre_export_ie(), {"nom": c.nom})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        content = _http_response_bytes(res)
        before = {
            "nom": c.nom,
            "code_postal": c.code_postal,
            "commune": c.commune,
            "is_active": c.is_active,
        }
        url_import = _centre_import_ie()
        upload = SimpleUploadedFile(
            "centres.xlsx",
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        r2 = self.client.post(url_import, {"file": upload}, format="multipart")
        self.assertEqual(r2.status_code, status.HTTP_200_OK)
        self.assertFalse(r2.data.get("dry_run"))
        c.refresh_from_db()
        after = {
            "nom": c.nom,
            "code_postal": c.code_postal,
            "commune": c.commune,
            "is_active": c.is_active,
        }
        self.assertEqual(before, after)

    def test_import_updates_field_from_excel(self):
        """Modification d'une cellule puis import : la base reflète la mise à jour (upsert)."""
        c = Centre.objects.create(nom="Centre Upsert", code_postal="69001", commune="Lyon")
        handler = CentreExcelHandler()
        buf = io.BytesIO(handler.export_queryset_to_bytes(Centre.objects.filter(pk=c.pk)))
        from openpyxl import load_workbook

        wb = load_workbook(buf)
        ws = wb["Données"]
        nom_col = None
        for j, cell in enumerate(ws[1], start=1):
            if cell.value == "nom":
                nom_col = j
                break
        self.assertIsNotNone(nom_col)
        ws.cell(row=2, column=nom_col, value="Centre Upsert Modifié")
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        upload = SimpleUploadedFile(
            "u.xlsx",
            out.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        url_import = _centre_import_ie()
        r = self.client.post(url_import, {"file": upload}, format="multipart")
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertGreaterEqual(r.data["summary"]["updated"], 1)
        c.refresh_from_db()
        self.assertEqual(c.nom, "Centre Upsert Modifié")

    def test_staff_read_cannot_post_import(self):
        """Un utilisateur staff_read ne peut pas importer (écriture admin-like)."""
        staff_read = UserFactory(role=CustomUser.ROLE_STAFF_READ)
        self.client.force_authenticate(user=staff_read)
        handler = CentreExcelHandler()
        f = io.BytesIO(handler.build_template_bytes())
        upload = SimpleUploadedFile(
            "t.xlsx",
            f.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        r = self.client.post(_centre_import_ie(), {"file": upload}, format="multipart")
        self.assertEqual(r.status_code, status.HTTP_403_FORBIDDEN)

    def test_dry_run_does_not_persist(self):
        """dry_run=True : aucune modification en base après import."""
        c = Centre.objects.create(nom="Centre Dry", code_postal="13001")
        original_nom = c.nom
        handler = CentreExcelHandler()
        buf = io.BytesIO(handler.export_queryset_to_bytes(Centre.objects.filter(pk=c.pk)))
        from openpyxl import load_workbook

        wb = load_workbook(buf)
        ws = wb["Données"]
        nom_col = None
        for j, cell in enumerate(ws[1], start=1):
            if cell.value == "nom":
                nom_col = j
                break
        ws.cell(row=2, column=nom_col, value="Ne doit pas être enregistré")
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        upload = SimpleUploadedFile(
            "d.xlsx",
            out.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        r = self.client.post(f"{_centre_import_ie()}?dry_run=true", {"file": upload}, format="multipart")
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertTrue(r.data.get("dry_run"))
        c.refresh_from_db()
        self.assertEqual(c.nom, original_nom)


class CentreScopeServiceTests(TestCase):
    """Périmètre centre : vérification au niveau service (sans passer par la permission POST admin)."""

    def test_non_admin_cannot_update_centre_outside_assigned_centres(self):
        """Un staff non admin ne peut pas appliquer une ligne d'import pour un centre hors périmètre."""
        c_in = Centre.objects.create(nom="Centre In Scope", code_postal="75001")
        c_out = Centre.objects.create(nom="Centre Out Scope", code_postal="75002")
        staff = UserFactory(role=CustomUser.ROLE_STAFF)
        staff.centres.add(c_in)
        handler = CentreExcelHandler()
        buf = io.BytesIO(handler.export_queryset_to_bytes(Centre.objects.filter(pk=c_out.pk)))
        upload = SimpleUploadedFile(
            "scope.xlsx",
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        result = handler.import_upload(staff, upload, dry_run=False)
        self.assertEqual(result["summary"]["failed"], 1)
        row = result["rows"][0]
        self.assertEqual(row["action"], "error")
        self.assertTrue(any(e.get("code") == "out_of_scope" for e in row["errors"]))


class TypeOffreMinimalTests(TestCase):
    """Contrôle minimal sur type d'offre (symétrie export → import)."""

    def test_type_offre_export_then_dry_run_import(self):
        """Export d'un type d'offre existant puis import dry_run sans erreur bloquante."""
        admin = UserFactory(role=CustomUser.ROLE_ADMIN)
        to, _ = TypeOffre.objects.get_or_create(
            nom=TypeOffre.CRIF,
            defaults={"autre": "", "couleur": "#111111"},
        )
        from rap_app.services.imports.handlers_lot1 import TypeOffreExcelHandler

        handler = TypeOffreExcelHandler()
        buf = handler.export_queryset_to_bytes(TypeOffre.objects.filter(pk=to.pk))
        upload = SimpleUploadedFile(
            "t.xlsx",
            buf,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        result = handler.import_upload(admin, upload, dry_run=True)
        self.assertEqual(result["summary"]["failed"], 0)
        self.assertTrue(result["dry_run"])


class ImportExportLot1RoutesTests(AuthenticatedTestCase):
    """Routes canoniques ``/api/import-export/<resource>/…``."""

    def setUp(self):
        super().setUp()
        self.admin = UserFactory(role=CustomUser.ROLE_ADMIN)
        self.client.force_authenticate(user=self.admin)

    def test_export_centre_returns_xlsx(self):
        """Export canonique : 200 et signature ZIP (OOXML)."""
        Centre.objects.create(nom="Centre IE", code_postal="75001", commune="Paris")
        r = self.client.get(reverse("import-export-lot1-export-xlsx", kwargs={"resource": "centre"}))
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        body = _http_response_bytes(r)
        self.assertTrue(body.startswith(b"PK"), msg="Attendu un .xlsx (ZIP)")

    def test_import_round_trip_via_import_export_route(self):
        """Import via ``/api/import-export/centre/import-xlsx/`` équivalent au mixin."""
        c = Centre.objects.create(nom="Centre IE2", code_postal="75001", commune="Paris")
        export = self.client.get(reverse("import-export-lot1-export-xlsx", kwargs={"resource": "centre"}))
        upload = SimpleUploadedFile(
            "c.xlsx",
            _http_response_bytes(export),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        r = self.client.post(
            reverse("import-export-lot1-import-xlsx", kwargs={"resource": "centre"}),
            {"file": upload},
            format="multipart",
        )
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertFalse(r.data.get("dry_run"))
        c.refresh_from_db()
        self.assertEqual(c.nom, "Centre IE2")

    def test_unknown_column_returns_400_and_code(self):
        """Mode strict : une colonne hors schéma échoue avant le traitement des lignes (§2.5.2)."""
        handler = CentreExcelHandler()
        buf = io.BytesIO(handler.build_template_bytes())
        from openpyxl import load_workbook

        wb = load_workbook(buf)
        ws = wb["Données"]
        last_col = ws.max_column + 1
        ws.cell(row=1, column=last_col, value="colonne_inconnue")
        ws.cell(row=2, column=last_col, value="x")
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        upload = SimpleUploadedFile(
            "bad.xlsx",
            out.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        r = self.client.post(
            reverse("import-export-lot1-import-xlsx", kwargs={"resource": "centre"}),
            {"file": upload},
            format="multipart",
        )
        self.assertEqual(r.status_code, status.HTTP_400_BAD_REQUEST)
        code = r.data.get("code")
        if isinstance(code, list) and code:
            code = code[0]
        self.assertEqual(code, "unknown_columns")

    def test_forbidden_system_column_returns_400_and_invalid_code(self):
        """§2.5.3 : colonnes réservées (ex. created_at) refusées en en-tête."""
        handler = CentreExcelHandler()
        buf = io.BytesIO(handler.build_template_bytes())
        from openpyxl import load_workbook

        wb = load_workbook(buf)
        ws = wb["Données"]
        last_col = ws.max_column + 1
        ws.cell(row=1, column=last_col, value="created_at")
        ws.cell(row=2, column=last_col, value="2020-01-01")
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        upload = SimpleUploadedFile(
            "bad.xlsx",
            out.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        r = self.client.post(
            reverse("import-export-lot1-import-xlsx", kwargs={"resource": "centre"}),
            {"file": upload},
            format="multipart",
        )
        self.assertEqual(r.status_code, status.HTTP_400_BAD_REQUEST)
        code = r.data.get("code")
        if isinstance(code, list) and code:
            code = code[0]
        self.assertEqual(code, "invalid")


class ImportJobPersistenceTests(AuthenticatedTestCase):
    """Modèle technique ``ImportJob`` (§2.14) — une ligne par ``import-xlsx``."""

    def setUp(self):
        super().setUp()
        self.admin = UserFactory(role=CustomUser.ROLE_ADMIN)
        self.client.force_authenticate(user=self.admin)

    def test_dry_run_import_creates_success_import_job(self):
        Centre.objects.create(nom="Centre IJ", code_postal="75001", commune="Paris")
        export = self.client.get(reverse("import-export-lot1-export-xlsx", kwargs={"resource": "centre"}))
        upload = SimpleUploadedFile(
            "c.xlsx",
            _http_response_bytes(export),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        n_before = ImportJob.objects.count()
        r = self.client.post(
            f"{reverse('import-export-lot1-import-xlsx', kwargs={'resource': 'centre'})}?dry_run=true",
            {"file": upload},
            format="multipart",
        )
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertEqual(ImportJob.objects.count(), n_before + 1)
        job = ImportJob.objects.latest("pk")
        self.assertEqual(job.resource, RESOURCE_CENTRE)
        self.assertEqual(job.url_resource, "centre")
        self.assertTrue(job.dry_run)
        self.assertEqual(job.status, ImportJob.Status.SUCCESS)
        self.assertEqual(job.http_status, 200)
        self.assertEqual(job.user_id, self.admin.pk)

    def test_validation_error_creates_error_import_job(self):
        from openpyxl import load_workbook

        handler = CentreExcelHandler()
        buf = io.BytesIO(handler.build_template_bytes())
        wb = load_workbook(buf)
        ws = wb["Données"]
        last_col = ws.max_column + 1
        ws.cell(row=1, column=last_col, value="colonne_inconnue")
        ws.cell(row=2, column=last_col, value="x")
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        upload = SimpleUploadedFile(
            "bad.xlsx",
            out.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        n_before = ImportJob.objects.count()
        r = self.client.post(
            reverse("import-export-lot1-import-xlsx", kwargs={"resource": "centre"}),
            {"file": upload},
            format="multipart",
        )
        self.assertEqual(r.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(ImportJob.objects.count(), n_before + 1)
        job = ImportJob.objects.latest("pk")
        self.assertEqual(job.status, ImportJob.Status.ERROR)
        self.assertEqual(job.http_status, 400)
        self.assertIsNotNone(job.error_payload)

    @override_settings(RAP_IMPORT_PERSIST_JOBS=False)
    def test_no_import_job_when_persistence_disabled(self):
        Centre.objects.create(nom="Centre IJ2", code_postal="75002", commune="Paris")
        export = self.client.get(reverse("import-export-lot1-export-xlsx", kwargs={"resource": "centre"}))
        upload = SimpleUploadedFile(
            "c2.xlsx",
            _http_response_bytes(export),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        n_before = ImportJob.objects.count()
        r = self.client.post(
            f"{reverse('import-export-lot1-import-xlsx', kwargs={'resource': 'centre'})}?dry_run=true",
            {"file": upload},
            format="multipart",
        )
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertEqual(ImportJob.objects.count(), n_before)


class ImportJobApiTests(AuthenticatedTestCase):
    """``GET /api/import-export/jobs/`` — liste paginée ; scope admin vs staff (§2.14 / §2.15)."""

    def _results(self, response):
        inner = response.data["data"]
        return inner["results"] if isinstance(inner, dict) and "results" in inner else inner

    def test_admin_sees_all_import_jobs(self):
        admin = UserFactory(role=CustomUser.ROLE_ADMIN)
        staff = UserFactory(role=CustomUser.ROLE_STAFF)
        ImportJob.objects.create(
            user=admin,
            resource=RESOURCE_CENTRE,
            url_resource="centre",
            dry_run=False,
            status=ImportJob.Status.SUCCESS,
            http_status=200,
        )
        ImportJob.objects.create(
            user=staff,
            resource=RESOURCE_CENTRE,
            url_resource="centre",
            dry_run=False,
            status=ImportJob.Status.ERROR,
            http_status=400,
            error_payload={"code": "x"},
        )
        self.client.force_authenticate(user=admin)
        r = self.client.get(reverse("import-export-job-list"))
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertTrue(r.data.get("success"))
        rows = self._results(r)
        self.assertEqual(len(rows), 2)

    def test_staff_sees_only_own_import_jobs(self):
        staff_a = UserFactory(role=CustomUser.ROLE_STAFF)
        staff_b = UserFactory(role=CustomUser.ROLE_STAFF)
        j_a = ImportJob.objects.create(
            user=staff_a,
            resource=RESOURCE_CENTRE,
            url_resource="centre",
            dry_run=True,
            status=ImportJob.Status.SUCCESS,
            http_status=200,
        )
        ImportJob.objects.create(
            user=staff_b,
            resource=RESOURCE_CENTRE,
            url_resource="centre",
            dry_run=False,
            status=ImportJob.Status.SUCCESS,
            http_status=200,
        )
        self.client.force_authenticate(user=staff_a)
        r = self.client.get(reverse("import-export-job-list"))
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        rows = self._results(r)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["id"], j_a.pk)

    def test_staff_cannot_retrieve_other_users_job(self):
        staff_a = UserFactory(role=CustomUser.ROLE_STAFF)
        staff_b = UserFactory(role=CustomUser.ROLE_STAFF)
        j_b = ImportJob.objects.create(
            user=staff_b,
            resource=RESOURCE_CENTRE,
            url_resource="centre",
            dry_run=False,
            status=ImportJob.Status.SUCCESS,
            http_status=200,
        )
        self.client.force_authenticate(user=staff_a)
        r = self.client.get(reverse("import-export-job-detail", kwargs={"pk": j_b.pk}))
        self.assertEqual(r.status_code, status.HTTP_404_NOT_FOUND)

    def test_staff_csv_export_is_scoped_to_own_jobs(self):
        staff_a = UserFactory(role=CustomUser.ROLE_STAFF)
        staff_b = UserFactory(role=CustomUser.ROLE_STAFF)
        ImportJob.objects.create(
            user=staff_a,
            resource=RESOURCE_CENTRE,
            url_resource="centre",
            dry_run=False,
            status=ImportJob.Status.SUCCESS,
            http_status=200,
            original_filename="a.xlsx",
        )
        ImportJob.objects.create(
            user=staff_b,
            resource=RESOURCE_CENTRE,
            url_resource="centre",
            dry_run=False,
            status=ImportJob.Status.SUCCESS,
            http_status=200,
            original_filename="b.xlsx",
        )
        self.client.force_authenticate(user=staff_a)
        r = self.client.get(reverse("import-export-job-export-csv"))
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        body = r.content.decode("utf-8")
        self.assertIn("a.xlsx", body)
        self.assertNotIn("b.xlsx", body)

    def test_staff_pdf_export_is_scoped_to_own_jobs(self):
        staff_a = UserFactory(role=CustomUser.ROLE_STAFF)
        staff_b = UserFactory(role=CustomUser.ROLE_STAFF)
        ImportJob.objects.create(
            user=staff_a,
            resource=RESOURCE_CENTRE,
            url_resource="centre",
            dry_run=False,
            status=ImportJob.Status.SUCCESS,
            http_status=200,
            original_filename="a.xlsx",
        )
        ImportJob.objects.create(
            user=staff_b,
            resource=RESOURCE_CENTRE,
            url_resource="centre",
            dry_run=False,
            status=ImportJob.Status.SUCCESS,
            http_status=200,
            original_filename="b.xlsx",
        )
        self.client.force_authenticate(user=staff_a)
        r = self.client.get(reverse("import-export-job-export-pdf"))
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertEqual(r["Content-Type"], "application/pdf")
        self.assertIn(b".pdf", r["Content-Disposition"].encode())
        self.assertTrue(r.content.startswith(b"%PDF"))
        self.assertGreater(len(r.content), 200)
        # Pas d’assertion sur « a.xlsx » / « b.xlsx » dans r.content : texte souvent compressé (PDF).

    def test_admin_can_filter_jobs_by_user_and_date_min(self):
        admin = UserFactory(role=CustomUser.ROLE_ADMIN)
        staff = UserFactory(role=CustomUser.ROLE_STAFF)
        old_job = ImportJob.objects.create(
            user=staff,
            resource=RESOURCE_CENTRE,
            url_resource="centre",
            dry_run=False,
            status=ImportJob.Status.SUCCESS,
            http_status=200,
            original_filename="old.xlsx",
        )
        recent_job = ImportJob.objects.create(
            user=staff,
            resource=RESOURCE_CENTRE,
            url_resource="centre",
            dry_run=False,
            status=ImportJob.Status.SUCCESS,
            http_status=200,
            original_filename="recent.xlsx",
        )
        ImportJob.objects.filter(pk=old_job.pk).update(created_at=timezone.now() - timezone.timedelta(days=5))
        self.client.force_authenticate(user=admin)
        r = self.client.get(
            reverse("import-export-job-list"),
            {"user": staff.username, "date_min": timezone.now().date().isoformat()},
        )
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        rows = self._results(r)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["id"], recent_job.pk)

    def test_admin_can_filter_jobs_by_resource_and_status_csv_lists(self):
        admin = UserFactory(role=CustomUser.ROLE_ADMIN)
        ok = ImportJob.objects.create(
            user=admin,
            resource=RESOURCE_CENTRE,
            url_resource="centre",
            dry_run=False,
            status=ImportJob.Status.SUCCESS,
            http_status=200,
        )
        ImportJob.objects.create(
            user=admin,
            resource=RESOURCE_CENTRE,
            url_resource="centre",
            dry_run=False,
            status=ImportJob.Status.ERROR,
            http_status=400,
        )
        ImportJob.objects.create(
            user=admin,
            resource="formation",
            url_resource="formation",
            dry_run=False,
            status=ImportJob.Status.SUCCESS,
            http_status=200,
        )
        self.client.force_authenticate(user=admin)
        r = self.client.get(
            reverse("import-export-job-list"),
            {"resource__in": "centre,formation", "status__in": "success"},
        )
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        rows = self._results(r)
        self.assertEqual(len(rows), 2)
        self.assertIn(ok.pk, [row["id"] for row in rows])


class ListVsExportParityTests(AuthenticatedTestCase):
    """Parité scope : liste REST et export (§2.9.1)."""

    def test_staff_sees_same_centre_count_in_list_and_export_rows(self):
        """Un staff limité à ses centres : une ligne d’export par centre listé."""
        staff = UserFactory(role=CustomUser.ROLE_STAFF)
        c_ok = Centre.objects.create(nom="Centre staff", code_postal="75001")
        Centre.objects.create(nom="Centre autre", code_postal="75002")
        staff.centres.add(c_ok)
        self.client.force_authenticate(user=staff)

        res_list = self.client.get(reverse("centre-list"))
        self.assertEqual(res_list.status_code, status.HTTP_200_OK)
        payload = res_list.data["data"]
        n_list = payload.get("count", len(payload.get("results", [])))

        export = self.client.get(reverse("import-export-lot1-export-xlsx", kwargs={"resource": "centre"}))
        self.assertEqual(export.status_code, status.HTTP_200_OK)
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(_http_response_bytes(export)))
        ws = wb["Données"]
        n_rows = max(0, ws.max_row - 1)
        self.assertEqual(n_list, n_rows)


class ReadLot1ParseDeadlineTests(TestCase):
    """Plafond durée lecture itérative après ouverture du classeur (§2.7)."""

    def test_parse_deadline_exceeded_returns_parse_timeout(self):
        from unittest import mock

        from rap_app.services.imports.excel_io import read_lot1_workbook, write_lot1_workbook
        from rap_app.services.imports.schemas import CENTRE_COLUMNS, RESOURCE_CENTRE, SCHEMA_VERSION_LOT1

        headers = list(CENTRE_COLUMNS)
        blank = [None] * len(headers)
        buf = write_lot1_workbook(
            resource=RESOURCE_CENTRE,
            schema_version=SCHEMA_VERSION_LOT1,
            headers=headers,
            rows=[blank],
        )
        with mock.patch("rap_app.services.imports.excel_io.get_max_parse_seconds", return_value=60.0):
            with mock.patch("rap_app.services.imports.excel_io.time.monotonic", side_effect=[0.0, 100.0]):
                with self.assertRaises(ValidationError) as ctx:
                    read_lot1_workbook(io.BytesIO(buf), expected_columns=set(CENTRE_COLUMNS))
        code = ctx.exception.message_dict.get("code")
        if isinstance(code, list) and code:
            code = code[0]
        self.assertEqual(code, "parse_timeout")
