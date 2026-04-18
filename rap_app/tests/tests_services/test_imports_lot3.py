"""Tests import/export Excel Lot 3 — Formation et Document (§2.10.1)."""

import io

from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from openpyxl import Workbook, load_workbook
from rest_framework import status
from rest_framework.request import Request
from rest_framework.test import APIRequestFactory, force_authenticate

from rap_app.models.centres import Centre
from rap_app.models.custom_user import CustomUser
from rap_app.models.documents import Document
from rap_app.models.formations import Formation
from rap_app.models.statut import Statut
from rap_app.models.types_offre import TypeOffre
from rap_app.services.imports.handlers_lot3 import DocumentExcelHandler, FormationExcelHandler
from rap_app.tests.factories import UserFactory
from rap_app.tests.tests_services.test_imports_lot1 import _http_response_bytes
from rap_app.tests.test_utils import AuthenticatedTestCase


def _form_export_ie():
    return reverse("import-export-lot1-export-xlsx", kwargs={"resource": "formation"})


def _form_import_ie():
    return reverse("import-export-lot1-import-xlsx", kwargs={"resource": "formation"})


def _doc_import_ie():
    return reverse("import-export-lot1-import-xlsx", kwargs={"resource": "document"})


def _doc_export_ie():
    return reverse("import-export-lot1-export-xlsx", kwargs={"resource": "document"})


class FormationDocumentExcelTests(AuthenticatedTestCase):
    def setUp(self):
        super().setUp()
        self.admin = UserFactory(role=CustomUser.ROLE_ADMIN)
        self.client.force_authenticate(user=self.admin)

    def test_meta_resource_formation_on_template(self):
        h = FormationExcelHandler()
        buf = io.BytesIO(h.build_template_bytes())
        wb = load_workbook(buf, read_only=True, data_only=True)
        try:
            meta = {
                str(r[0].value).strip(): r[1].value
                for r in wb["Meta"].iter_rows(min_col=1, max_col=2)
                if r[0].value
            }
            self.assertEqual(meta.get("resource"), "formation")
        finally:
            wb.close()

    def test_formation_export_import_round_trip_dry_run(self):
        centre = Centre.objects.create(nom="C Form IE", code_postal="75001", created_by=self.admin)
        to = TypeOffre.objects.create(nom="crif", created_by=self.admin)
        st = Statut.objects.create(nom="non_defini", couleur="#000000", created_by=self.admin)
        f = Formation.objects.create(
            nom="Formation Excel Lot3",
            centre=centre,
            type_offre=to,
            statut=st,
            created_by=self.admin,
        )
        export = self.client.get(_form_export_ie())
        self.assertEqual(export.status_code, status.HTTP_200_OK)
        upload = SimpleUploadedFile(
            "f.xlsx",
            _http_response_bytes(export),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        r = self.client.post(f"{_form_import_ie()}?dry_run=true", {"file": upload}, format="multipart")
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertTrue(r.data.get("dry_run"))
        self.assertEqual(r.data.get("resource"), "formation")
        self.assertEqual(r.data["summary"]["failed"], 0)

    def test_commercial_cannot_post_formation_import(self):
        commercial = UserFactory(role=CustomUser.ROLE_COMMERCIAL)
        self.client.force_authenticate(user=commercial)
        h = FormationExcelHandler()
        upload = SimpleUploadedFile(
            "t.xlsx",
            h.build_template_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        r = self.client.post(_form_import_ie(), {"file": upload}, format="multipart")
        self.assertEqual(r.status_code, status.HTTP_403_FORBIDDEN)

    def test_document_import_returns_not_supported(self):
        h = DocumentExcelHandler()
        upload = SimpleUploadedFile(
            "d.xlsx",
            h.build_template_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        r = self.client.post(_doc_import_ie(), {"file": upload}, format="multipart")
        self.assertEqual(r.status_code, status.HTTP_400_BAD_REQUEST)
        code = r.data.get("code")
        if isinstance(code, list):
            code = code[0]
        self.assertEqual(code, "not_supported")

    def test_document_export_respects_formation_query_param(self):
        """Parité liste REST / export import-export : filtre ``formation`` (§2.9.1)."""
        centre = Centre.objects.create(nom="C Doc IE", code_postal="75001", created_by=self.admin)
        to = TypeOffre.objects.create(nom="crif", created_by=self.admin)
        st = Statut.objects.create(nom="non_defini", couleur="#000000", created_by=self.admin)
        f1 = Formation.objects.create(
            nom="F1 Doc IE", centre=centre, type_offre=to, statut=st, created_by=self.admin
        )
        f2 = Formation.objects.create(
            nom="F2 Doc IE", centre=centre, type_offre=to, statut=st, created_by=self.admin
        )
        pdf = SimpleUploadedFile("a.pdf", b"%PDF-1.4 minimal", content_type="application/pdf")
        pdf2 = SimpleUploadedFile("b.pdf", b"%PDF-1.4 deux", content_type="application/pdf")
        d1 = Document.objects.create(
            formation=f1,
            nom_fichier="a.pdf",
            fichier=pdf,
            type_document=Document.PDF,
            created_by=self.admin,
        )
        d2 = Document.objects.create(
            formation=f2,
            nom_fichier="b.pdf",
            fichier=pdf2,
            type_document=Document.PDF,
            created_by=self.admin,
        )
        self.addCleanup(d1.fichier.close)
        self.addCleanup(d2.fichier.close)

        def _ids_from_export(resp):
            self.assertEqual(resp.status_code, status.HTTP_200_OK)
            buf = io.BytesIO(_http_response_bytes(resp))
            wb = load_workbook(buf, read_only=True, data_only=True)
            try:
                ws = wb["Données"]
                out = []
                for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                    v = row[0]
                    if v is not None:
                        out.append(int(v))
                return out
            finally:
                wb.close()

        all_ids = set(_ids_from_export(self.client.get(_doc_export_ie())))
        self.assertEqual(all_ids, {d1.id, d2.id})

        f1_ids = set(_ids_from_export(self.client.get(_doc_export_ie(), {"formation": f1.id})))
        self.assertEqual(f1_ids, {d1.id})

    def test_formation_import_french_headers_export_sheet_without_meta_dry_run(self):
        """Feuille « Formations », libellés FR, centre par nom + slugs type/statut sans feuille Meta."""
        centre = Centre.objects.create(nom="CRIF Import FR", code_postal="75001", created_by=self.admin)
        to = TypeOffre.objects.create(nom="crif", created_by=self.admin)
        st = Statut.objects.create(nom="non_defini", couleur="#000000", created_by=self.admin)
        self.assertIsNotNone(to.pk)
        self.assertIsNotNone(st.pk)

        wb = Workbook()
        ws = wb.active
        ws.title = "Formations"
        headers = ["Centre", "Formation", "Type d'offre", "Statut", "Date début", "Date fin"]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=4, column=col, value=h)
        ws.cell(row=5, column=1, value=centre.nom)
        ws.cell(row=5, column=2, value="Formation import FR pytest")
        ws.cell(row=5, column=3, value="crif")
        ws.cell(row=5, column=4, value="non_defini")
        ws.cell(row=5, column=5, value="01/01/2026")
        ws.cell(row=5, column=6, value="31/12/2026")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        upload = SimpleUploadedFile(
            "formations_export.xlsx",
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        factory = APIRequestFactory()
        django_request = factory.get("/api/import-export/formation/import-xlsx/")
        force_authenticate(django_request, user=self.admin)
        request = Request(django_request)

        h = FormationExcelHandler()
        out = h.import_upload(self.admin, upload, dry_run=True, request=request)
        self.assertEqual(out.get("summary", {}).get("failed"), 0, out)
        self.assertEqual(out.get("summary", {}).get("created"), 1)
