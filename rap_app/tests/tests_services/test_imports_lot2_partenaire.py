"""Tests import/export Excel Lot 2 — Partenaire."""

import io

from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework import status

from rap_app.models.centres import Centre
from rap_app.models.custom_user import CustomUser
from rap_app.models.partenaires import Partenaire
from rap_app.services.imports.handlers_lot2 import PartenaireExcelHandler
from rap_app.tests.factories import UserFactory
from rap_app.tests.tests_services.test_imports_lot1 import _http_response_bytes
from rap_app.tests.test_utils import AuthenticatedTestCase


def _part_export_ie():
    return reverse("import-export-lot1-export-xlsx", kwargs={"resource": "partenaire"})


def _part_import_ie():
    return reverse("import-export-lot1-import-xlsx", kwargs={"resource": "partenaire"})


class PartenaireExcelHandlerTests(AuthenticatedTestCase):
    def setUp(self):
        super().setUp()
        self.admin = UserFactory(role=CustomUser.ROLE_ADMIN)
        self.client.force_authenticate(user=self.admin)

    def test_meta_resource_partenaire_on_template(self):
        h = PartenaireExcelHandler()
        buf = io.BytesIO(h.build_template_bytes())
        wb = load_workbook(buf, read_only=True, data_only=True)
        try:
            meta = {str(r[0].value).strip(): r[1].value for r in wb["Meta"].iter_rows(min_col=1, max_col=2) if r[0].value}
            self.assertEqual(meta.get("resource"), "partenaire")
        finally:
            wb.close()

    def test_export_import_round_trip_dry_run(self):
        p = Partenaire.objects.create(nom="Partenaire Excel Lot2 Unique", type=Partenaire.TYPE_ENTREPRISE)
        export = self.client.get(_part_export_ie())
        self.assertEqual(export.status_code, status.HTTP_200_OK)
        upload = SimpleUploadedFile(
            "p.xlsx",
            _http_response_bytes(export),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        r = self.client.post(f"{_part_import_ie()}?dry_run=true", {"file": upload}, format="multipart")
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertTrue(r.data.get("dry_run"))
        self.assertEqual(r.data.get("resource"), "partenaire")
        self.assertEqual(r.data["summary"]["failed"], 0)

    def test_staff_read_cannot_post_import(self):
        Partenaire.objects.create(nom="Partenaire SR", type=Partenaire.TYPE_ENTREPRISE)
        staff_read = UserFactory(role=CustomUser.ROLE_STAFF_READ)
        self.client.force_authenticate(user=staff_read)
        h = PartenaireExcelHandler()
        upload = SimpleUploadedFile(
            "t.xlsx",
            h.build_template_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        r = self.client.post(_part_import_ie(), {"file": upload}, format="multipart")
        self.assertEqual(r.status_code, status.HTTP_403_FORBIDDEN)

    def test_default_centre_scope_staff(self):
        """Staff limité : default_centre_id hors périmètre → erreur de ligne."""
        c_ok = Centre.objects.create(nom="Centre P Scope", code_postal="75001")
        c_out = Centre.objects.create(nom="Centre P Hors", code_postal="75002")
        staff = UserFactory(role=CustomUser.ROLE_STAFF)
        staff.centres.add(c_ok)
        handler = PartenaireExcelHandler()
        buf = io.BytesIO(handler.build_template_bytes())
        wb = load_workbook(buf)
        ws = wb["Données"]
        headers = [c.value for c in ws[1]]
        col_n = headers.index("nom") + 1
        col_dc = headers.index("default_centre_id") + 1
        ws.cell(row=2, column=col_n, value="Partenaire Hors Centre Default")
        ws.cell(row=2, column=col_dc, value=c_out.id)
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        upload = SimpleUploadedFile(
            "p.xlsx",
            out.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        from rest_framework.test import APIRequestFactory

        factory = APIRequestFactory()
        django_request = factory.post("/api/import-export/partenaire/import-xlsx/")
        django_request.user = staff
        result = handler.import_upload(staff, upload, dry_run=True, request=django_request)
        self.assertEqual(result["summary"]["failed"], 1)
        err = result["rows"][0]["errors"][0]
        self.assertEqual(err.get("code"), "out_of_scope")
