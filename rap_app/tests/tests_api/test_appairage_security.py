from io import BytesIO

from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from django.utils import timezone

from rap_app.models.appairage import Appairage, AppairageStatut
from rap_app.models.commentaires_appairage import CommentaireAppairage
from rap_app.models.candidat import Candidat
from rap_app.models.centres import Centre
from rap_app.models.formations import Formation
from rap_app.models.partenaires import Partenaire
from rap_app.models.statut import Statut
from rap_app.models.types_offre import TypeOffre
from rap_app.services.placement_services import AppairagePlacementService, defer_appairage_snapshot_sync
from rap_app.tests.factories import UserFactory


class AppairageSecurityTests(APITestCase):
    def setUp(self):
        self.type_offre = TypeOffre.objects.create(nom=TypeOffre.CRIF)
        self.statut = Statut.objects.create(nom=Statut.NON_DEFINI)

    def _create_appairage(self, centre, formation_nom="Formation X", partenaire_nom="Entreprise Test"):
        formation = Formation.objects.create(
            nom=formation_nom,
            centre=centre,
            type_offre=self.type_offre,
            statut=self.statut,
        )

        candidat = Candidat.objects.create(
            nom="Dupont",
            prenom="Jean",
            formation=formation,
        )

        partenaire = Partenaire.objects.create(
            nom=partenaire_nom,
        )

        return Appairage.objects.create(
            candidat=candidat,
            partenaire=partenaire,
            formation=formation,
        )

    def test_appairage_retrieve_outside_centre_forbidden(self):
        centre_a = Centre.objects.create(nom="Centre A")
        centre_b = Centre.objects.create(nom="Centre B")

        user = UserFactory(role="staff")
        user.centres.add(centre_a)

        appairage = self._create_appairage(centre_b)

        self.client.force_authenticate(user=user)

        url = f"/api/appairages/{appairage.id}/"
        response = self.client.get(url)

        self.assertIn(
            response.status_code,
            [status.HTTP_403_FORBIDDEN, status.HTTP_404_NOT_FOUND],
        )

    def test_appairage_retrieve_same_centre_allowed(self):
        centre = Centre.objects.create(nom="Centre A")

        user = UserFactory(role="staff")
        user.centres.add(centre)

        appairage = self._create_appairage(centre)

        self.client.force_authenticate(user=user)

        url = f"/api/appairages/{appairage.id}/"
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_commercial_can_create_appairage_in_scope(self):
        centre = Centre.objects.create(nom="Centre Commercial")
        user = UserFactory(role="commercial")
        user.centres.add(centre)

        formation = Formation.objects.create(
            nom="Formation Commerciale",
            centre=centre,
            type_offre=self.type_offre,
            statut=self.statut,
        )
        candidat = Candidat.objects.create(
            nom="Martin",
            prenom="Luc",
            formation=formation,
        )
        partenaire = Partenaire.objects.create(nom="Entreprise Commerciale")

        self.client.force_authenticate(user=user)
        response = self.client.post(
            "/api/appairages/",
            {
                "candidat": candidat.id,
                "partenaire": partenaire.id,
                "formation": formation.id,
                "statut": AppairageStatut.TRANSMIS,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        payload = response.data.get("data", response.data)
        self.assertEqual(payload["formation"], formation.id)

    def test_charge_recrutement_cannot_create_appairage_outside_scope(self):
        centre_autorise = Centre.objects.create(nom="Centre Autorise")
        centre_hors_scope = Centre.objects.create(nom="Centre Hors Scope")

        user = UserFactory(role="charge_recrutement")
        user.centres.add(centre_autorise)

        formation = Formation.objects.create(
            nom="Formation Hors Scope",
            centre=centre_hors_scope,
            type_offre=self.type_offre,
            statut=self.statut,
        )
        candidat = Candidat.objects.create(
            nom="Durand",
            prenom="Lea",
            formation=formation,
        )
        partenaire = Partenaire.objects.create(nom="Entreprise Hors Scope")

        self.client.force_authenticate(user=user)
        response = self.client.post(
            "/api/appairages/",
            {
                "candidat": candidat.id,
                "partenaire": partenaire.id,
                "formation": formation.id,
                "statut": AppairageStatut.TRANSMIS,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_staff_read_cannot_archive_appairage(self):
        centre = Centre.objects.create(nom="Centre A")

        user = UserFactory(role="staff_read")
        user.centres.add(centre)

        appairage = self._create_appairage(centre)

        self.client.force_authenticate(user=user)

        url = f"/api/appairages/{appairage.id}/archiver/"
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_appairage_export_xlsx_respects_centre_scope(self):
        centre_a = Centre.objects.create(nom="Centre A")
        centre_b = Centre.objects.create(nom="Centre B")

        user = UserFactory(role="staff")
        user.centres.add(centre_a)

        self._create_appairage(
            centre=centre_a,
            formation_nom="Formation Visible",
            partenaire_nom="Entreprise Visible",
        )
        visible = Appairage.objects.filter(
            formation__centre=centre_a,
            partenaire__nom="Entreprise Visible",
        ).first()
        CommentaireAppairage.objects.create(
            appairage=visible,
            body="Commentaire visible export",
            created_by=user,
        )

        self._create_appairage(
            centre=centre_b,
            formation_nom="Formation Cachee",
            partenaire_nom="Entreprise Cachee",
        )

        self.client.force_authenticate(user=user)

        response = self.client.get("/api/appairages/export-xlsx/")

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
            msg=getattr(response, "content", b"")[:300],
        )
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        wb = load_workbook(filename=BytesIO(response.content))
        ws = wb.active

        exported_text = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    exported_text.append(str(cell))

        exported_blob = "\n".join(exported_text)

        self.assertIn("Formation Visible", exported_blob)
        self.assertIn("Entreprise Visible", exported_blob)
        self.assertIn("Commentaire visible export", exported_blob)

        self.assertNotIn("Formation Cachee", exported_blob)
        self.assertNotIn("Entreprise Cachee", exported_blob)

    def test_archiving_appairage_refreshes_candidate_snapshot(self):
        centre = Centre.objects.create(nom="Centre A")
        user = UserFactory(role="staff")
        user.centres.add(centre)

        appairage = self._create_appairage(centre)
        AppairagePlacementService.sync_after_save(appairage, actor=user)
        candidat = appairage.candidat
        candidat.refresh_from_db()
        self.assertEqual(candidat.placement_appairage_id, appairage.id)

        self.client.force_authenticate(user=user)
        response = self.client.post(f"/api/appairages/{appairage.id}/archiver/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        candidat.refresh_from_db()
        self.assertIsNone(candidat.placement_appairage_id)
        self.assertIsNone(candidat.entreprise_placement_id)

    def test_delete_archives_last_appairage_and_clears_candidate_snapshot(self):
        centre = Centre.objects.create(nom="Centre A")
        user = UserFactory(role="staff")
        user.centres.add(centre)

        appairage = self._create_appairage(centre)
        AppairagePlacementService.sync_after_save(appairage, actor=user)
        candidat = appairage.candidat
        candidat.refresh_from_db()
        self.assertEqual(candidat.placement_appairage_id, appairage.id)

        self.client.force_authenticate(user=user)
        response = self.client.delete(f"/api/appairages/{appairage.id}/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        appairage.refresh_from_db()
        self.assertEqual(appairage.activite, "archive")
        candidat.refresh_from_db()
        self.assertIsNone(candidat.placement_appairage_id)
        self.assertIsNone(candidat.entreprise_placement_id)

    def test_unarchiving_appairage_restores_candidate_snapshot(self):
        centre = Centre.objects.create(nom="Centre A")
        user = UserFactory(role="staff")
        user.centres.add(centre)

        appairage = self._create_appairage(centre)
        AppairagePlacementService.sync_after_save(appairage, actor=user)

        self.client.force_authenticate(user=user)
        archive_response = self.client.post(f"/api/appairages/{appairage.id}/archiver/")
        self.assertEqual(archive_response.status_code, status.HTTP_200_OK)

        unarchive_response = self.client.post(f"/api/appairages/{appairage.id}/desarchiver/")
        self.assertEqual(unarchive_response.status_code, status.HTTP_200_OK)

        candidat = appairage.candidat
        candidat.refresh_from_db()
        self.assertEqual(candidat.placement_appairage_id, appairage.id)
        self.assertEqual(candidat.entreprise_placement_id, appairage.partenaire_id)

    def test_list_appairages_can_include_archived_items(self):
        centre = Centre.objects.create(nom="Centre A")
        user = UserFactory(role="staff")
        user.centres.add(centre)

        appairage = self._create_appairage(centre)
        appairage.archiver(user=user)

        self.client.force_authenticate(user=user)
        response = self.client.get("/api/appairages/", {"avec_archivees": "true"})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.data.get("data", response.data)
        returned_ids = [item["id"] for item in payload["results"]]
        self.assertIn(appairage.id, returned_ids)

    def test_list_appairages_can_show_archived_only(self):
        centre = Centre.objects.create(nom="Centre A")
        user = UserFactory(role="staff")
        user.centres.add(centre)

        archived = self._create_appairage(centre, formation_nom="Formation Archivee")
        archived.archiver(user=user)
        active = self._create_appairage(centre, formation_nom="Formation Active", partenaire_nom="Entreprise Active")

        self.client.force_authenticate(user=user)
        response = self.client.get("/api/appairages/", {"activite": "archive"})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        payload = response.data.get("data", response.data)
        returned_ids = [item["id"] for item in payload["results"]]
        self.assertIn(archived.id, returned_ids)
        self.assertNotIn(active.id, returned_ids)

    def test_delete_archives_active_appairage_and_falls_back_to_other_active_one(self):
        centre = Centre.objects.create(nom="Centre A")
        user = UserFactory(role="staff")
        user.centres.add(centre)

        appairage = self._create_appairage(centre, formation_nom="Formation A", partenaire_nom="Entreprise A")
        candidat = appairage.candidat
        other_partenaire = Partenaire.objects.create(nom="Entreprise B")
        with defer_appairage_snapshot_sync():
            fallback_appairage = Appairage.objects.create(
                candidat=candidat,
                partenaire=other_partenaire,
                formation=appairage.formation,
                statut=AppairageStatut.ACCEPTE,
                date_appairage=timezone.now(),
            )

        AppairagePlacementService.sync_after_save(fallback_appairage, actor=user)
        candidat.refresh_from_db()
        self.assertEqual(candidat.placement_appairage_id, fallback_appairage.id)

        self.client.force_authenticate(user=user)
        response = self.client.delete(f"/api/appairages/{fallback_appairage.id}/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        fallback_appairage.refresh_from_db()
        self.assertEqual(fallback_appairage.activite, "archive")
        candidat.refresh_from_db()
        self.assertEqual(candidat.placement_appairage_id, appairage.id)
        self.assertEqual(candidat.entreprise_placement_id, appairage.partenaire_id)
