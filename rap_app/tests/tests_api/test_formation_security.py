from datetime import timedelta

from django.utils import timezone
from rest_framework import status
from rest_framework.test import APITestCase

from rap_app.models.centres import Centre
from rap_app.models.commentaires import Commentaire
from rap_app.models.formations import Formation
from rap_app.models.statut import Statut
from rap_app.models.types_offre import TypeOffre
from rap_app.tests.factories import UserFactory


class FormationSecurityTests(APITestCase):
    def _create_formation(self, centre, nom="Formation Test"):
        today = timezone.localdate()

        type_offre = TypeOffre.objects.create(nom=TypeOffre.CRIF)

        statut = Statut.objects.create(nom=Statut.NON_DEFINI)

        return Formation.objects.create(
            nom=nom,
            centre=centre,
            type_offre=type_offre,
            statut=statut,
            start_date=today + timedelta(days=7),
            end_date=today + timedelta(days=30),
        )

    def test_formation_retrieve_outside_centre_forbidden(self):
        centre_a = Centre.objects.create(nom="Centre A")
        centre_b = Centre.objects.create(nom="Centre B")

        user = UserFactory(role="staff")
        user.centres.add(centre_a)

        formation = self._create_formation(centre_b, nom="Formation B")

        self.client.force_authenticate(user=user)

        url = f"/api/formations/{formation.id}/"
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_formation_retrieve_same_centre_allowed(self):
        centre = Centre.objects.create(nom="Centre A")

        user = UserFactory(role="staff")
        user.centres.add(centre)

        formation = self._create_formation(centre, nom="Formation A")

        self.client.force_authenticate(user=user)

        url = f"/api/formations/{formation.id}/"
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_formation_archiver_outside_centre_forbidden(self):
        centre_a = Centre.objects.create(nom="Centre A")
        centre_b = Centre.objects.create(nom="Centre B")

        user = UserFactory(role="staff")
        user.centres.add(centre_a)

        formation = self._create_formation(centre_b, nom="Formation B")

        self.client.force_authenticate(user=user)

        url = f"/api/formations/{formation.id}/archiver/"
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_staff_read_cannot_archive_formation(self):
        centre = Centre.objects.create(nom="Centre A")

        user = UserFactory(role="staff_read")
        user.centres.add(centre)

        formation = self._create_formation(centre, nom="Formation A")

        self.client.force_authenticate(user=user)

        url = f"/api/formations/{formation.id}/archiver/"
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_commercial_can_list_formations_in_scope_but_cannot_create(self):
        centre = Centre.objects.create(nom="Centre Commercial")
        user = UserFactory(role="commercial")
        user.centres.add(centre)

        self._create_formation(centre, nom="Formation Commerciale")

        self.client.force_authenticate(user=user)

        list_response = self.client.get("/api/formations/")
        self.assertEqual(list_response.status_code, status.HTTP_200_OK)
        self.assertEqual(list_response.data["data"]["count"], 1)

        payload = {
            "nom": "Nouvelle formation commerciale",
            "centre_id": centre.id,
            "type_offre_id": TypeOffre.objects.create(nom=TypeOffre.CRIF).id,
            "statut_id": Statut.objects.create(nom=Statut.NON_DEFINI).id,
            "start_date": str(timezone.localdate() + timedelta(days=3)),
            "end_date": str(timezone.localdate() + timedelta(days=10)),
        }
        create_response = self.client.post("/api/formations/", payload, format="json")
        self.assertEqual(create_response.status_code, status.HTTP_403_FORBIDDEN)

    def test_charge_recrutement_can_retrieve_but_cannot_update_formation(self):
        centre = Centre.objects.create(nom="Centre Recrutement")
        user = UserFactory(role="charge_recrutement")
        user.centres.add(centre)

        formation = self._create_formation(centre, nom="Formation Recrutement")

        self.client.force_authenticate(user=user)

        detail_response = self.client.get(f"/api/formations/{formation.id}/")
        self.assertEqual(detail_response.status_code, status.HTTP_200_OK)

        patch_response = self.client.patch(
            f"/api/formations/{formation.id}/",
            data={"nom": "Formation Interdite"},
            format="json",
        )
        self.assertEqual(patch_response.status_code, status.HTTP_403_FORBIDDEN)
        formation.refresh_from_db()
        self.assertEqual(formation.nom, "Formation Recrutement")

    def test_list_dans_filter_stays_scoped_to_user_centres(self):
        centre_a = Centre.objects.create(nom="Centre A")
        centre_b = Centre.objects.create(nom="Centre B")

        user = UserFactory(role="staff")
        user.centres.add(centre_a)

        formation_visible = self._create_formation(centre_a, nom="Formation Visible")
        formation_hidden = self._create_formation(centre_b, nom="Formation Cachée")

        self.client.force_authenticate(user=user)

        response = self.client.get("/api/formations/?dans=4w")

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        results = response.data.get("data", {}).get("results", [])
        returned_ids = [item["id"] for item in results]

        self.assertIn(formation_visible.id, returned_ids)
        self.assertNotIn(formation_hidden.id, returned_ids)


from io import BytesIO
from datetime import timedelta

import pytest
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APIClient

from rap_app.models.centres import Centre
from rap_app.models.formations import Formation
from rap_app.models.statut import Statut
from rap_app.models.types_offre import TypeOffre
from rap_app.tests.factories import UserFactory


@pytest.mark.django_db
def test_export_xlsx_respects_centre_scope():

    client = APIClient()

    centre_a = Centre.objects.create(nom="Centre A")
    centre_b = Centre.objects.create(nom="Centre B")

    user = UserFactory(role="staff")
    user.centres.add(centre_a)

    type_offre = TypeOffre.objects.create(nom=TypeOffre.CRIF)
    statut = Statut.objects.create(nom=Statut.NON_DEFINI)

    today = timezone.localdate()

    formation_visible = Formation.objects.create(
        nom="Formation Visible",
        centre=centre_a,
        type_offre=type_offre,
        statut=statut,
        start_date=today + timedelta(days=5),
        end_date=today + timedelta(days=30),
    )
    Commentaire.objects.create(
        formation=formation_visible,
        contenu="Commentaire export visible",
        created_by=user,
    )

    formation_cachee = Formation.objects.create(
        nom="Formation Cachee",
        centre=centre_b,
        type_offre=type_offre,
        statut=statut,
        start_date=today + timedelta(days=5),
        end_date=today + timedelta(days=30),
    )

    client.force_authenticate(user=user)

    response = client.get("/api/formations/export-xlsx/")

    assert response.status_code == 200

    # Vérifie que c'est bien un Excel
    assert response["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    workbook = load_workbook(filename=BytesIO(response.content))
    worksheet = workbook.active
    exported_values = {
        str(cell)
        for row in worksheet.iter_rows(values_only=True)
        for cell in row
        if cell not in (None, "")
    }

    # La formation du centre visible doit apparaître
    assert "Formation Visible" in exported_values
    assert any("Commentaire export visible" in value for value in exported_values)

    # La formation d'un autre centre ne doit PAS apparaître
    assert "Formation Cachee" not in exported_values
