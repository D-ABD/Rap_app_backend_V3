"""
Lecture et écriture de classeurs Excel pour le Lot 1 (feuilles Meta, Données, Aide V1).

Convention : ligne 1 de la feuille « Données » = en-têtes ``snake_case`` ;
lignes 2+ = données ; ``row_number`` dans les rapports = indice 1-based (ligne Excel).
"""

from __future__ import annotations

import io
import time
from datetime import datetime, timezone
from typing import Any

from django.conf import settings
from django.core.exceptions import ValidationError
from openpyxl import Workbook, load_workbook

from .schemas import (
    FORBIDDEN_IMPORT_COLUMN_NAMES,
    SCHEMA_VERSION_LOT1,
    ERR_INVALID,
    ERR_INVALID_FILE,
    ERR_PARSE_TIMEOUT,
    ERR_SCHEMA_MISMATCH,
    ERR_UNKNOWN_COLUMNS,
)

SHEET_DATA = "Données"
SHEET_META = "Meta"
SHEET_AIDE = "Aide"

# §2.11.2 — Aide V1 (texte court, pas de grosses tables FK).
AIDE_V1_LINES: tuple[str, ...] = (
    "Lot 1 — Modèle symétrique import/export. Ne modifiez pas la feuille « Meta » (schema_version, resource, generated_at).",
    "Feuille « Données » : ligne 1 = en-têtes snake_case ; chaque ligne suivante = un enregistrement.",
    "Import : utilisez l’endpoint POST …/import-xlsx/ avec le champ multipart « file ». Simulation : ajoutez le paramètre de requête dry_run=true (aucune écriture en base ; summary = résultat simulé).",
    "Colonnes interdites : champs système / audit (ex. created_at, updated_at, created_by, updated_by) — ne pas les ajouter ; le fichier sera refusé.",
)


def get_max_lot1_data_rows() -> int:
    """Plafond de lignes de données (hors en-tête) pour limiter CPU/mémoire (§2.7)."""
    return int(getattr(settings, "RAP_IMPORT_MAX_LOT1_DATA_ROWS", 50_000))


def get_max_parse_seconds() -> float:
    """
    Plafond de durée (secondes) pour la lecture itérative après ``load_workbook``.

    ``0`` désactive la vérification. N’interrompt pas l’ouverture ZIP/openpyxl elle-même (best-effort §2.7).
    """
    raw = getattr(settings, "RAP_IMPORT_MAX_PARSE_SECONDS", 120)
    try:
        v = float(raw)
    except (TypeError, ValueError):
        return 0.0
    return max(0.0, v)


def _check_parse_deadline(deadline: float | None) -> None:
    if deadline is None:
        return
    if time.monotonic() > deadline:
        raise ValidationError(
            {
                "file": (
                    "Temps maximum d’analyse du classeur dépassé. Réduisez le fichier ou contactez l’administrateur."
                ),
                "code": ERR_PARSE_TIMEOUT,
            }
        )


def validate_data_sheet_headers_strict(headers: list[str], expected: set[str]) -> None:
    """
    Contrôle strict des en-têtes (§2.5.2) : exactement l'ensemble attendu, ni colonne
    manquante ni colonne inconnue — exécuté juste après la ligne d'en-têtes.

    Args:
        headers: En-têtes lus sur la feuille « Données » (ordre libre).
        expected: Ensemble des noms de colonnes autorisés (snake_case).

    Raises:
        django.core.exceptions.ValidationError: Colonnes manquantes ou inconnues ;
            clé ``code`` = ``unknown_columns`` ou ``invalid`` (colonnes réservées §2.5.3).
    """
    got = {str(h).strip() for h in headers if h is not None and str(h).strip()}
    forbidden = got & FORBIDDEN_IMPORT_COLUMN_NAMES
    if forbidden:
        raise ValidationError(
            {
                "file": f"Colonnes réservées (audit / système), non importables : {', '.join(sorted(forbidden))}.",
                "code": ERR_INVALID,
            }
        )
    unknown = got - expected
    missing = expected - got
    if unknown:
        raise ValidationError(
            {
                "file": f"Colonnes inconnues : {', '.join(sorted(unknown))}.",
                "code": ERR_UNKNOWN_COLUMNS,
            }
        )
    if missing:
        raise ValidationError({"file": f"Colonnes manquantes : {', '.join(sorted(missing))}."})


def python_to_cell(value: Any) -> Any:
    """
    Prépare une valeur Python pour écriture dans une cellule Excel.

    Les ``datetime`` sont conservés ; None devient une cellule vide.

    Args:
        value: Valeur à sérialiser.

    Returns:
        Valeur compatible openpyxl.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    return value


def cell_to_python(value: Any, *, as_bool: bool = False) -> Any:
    """
    Interprète une valeur lue depuis Excel (types hétérogènes selon les clients).

    Args:
        value: Valeur brute ``cell.value``.
        as_bool: Si True, interprète texte/entier comme booléen.

    Returns:
        Valeur normalisée ; chaînes vides -> None pour les champs optionnels hors bool.
    """
    if value is None:
        return None if not as_bool else False
    if as_bool:
        if isinstance(value, bool):
            return value
        s = str(value).strip().lower()
        if s in ("", "0", "false", "non", "no", "f", "n"):
            return False
        return True
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return value
    if isinstance(value, str):
        s = value.strip()
        return s if s else None
    return value


def write_lot1_workbook(
    *,
    resource: str,
    schema_version: int,
    headers: list[str],
    rows: list[list[Any]],
    include_aide_sheet: bool = True,
) -> bytes:
    """
    Construit un classeur .xlsx avec feuilles Données, Meta et (V1) Aide.

    Args:
        resource: Identifiant ``resource`` (ex. ``centre``).
        schema_version: Version du schéma embarquée en Meta.
        headers: Liste ordonnée des noms de colonnes (ligne 1).
        rows: Lignes de valeurs alignées sur ``headers``.
        include_aide_sheet: Si True, ajoute la feuille « Aide » (§2.11.2).

    Returns:
        bytes: Contenu binaire du fichier ``.xlsx``.
    """
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = SHEET_DATA
    for col, h in enumerate(headers, start=1):
        ws_data.cell(row=1, column=col, value=h)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws_data.cell(row=r_idx, column=c_idx, value=python_to_cell(val))

    ws_meta = wb.create_sheet(SHEET_META)
    meta_rows = [
        ("schema_version", schema_version),
        ("resource", resource),
        ("generated_at", datetime.now(timezone.utc).replace(microsecond=0).isoformat()),
    ]
    for i, (k, v) in enumerate(meta_rows, start=1):
        ws_meta.cell(row=i, column=1, value=k)
        ws_meta.cell(row=i, column=2, value=v)

    if include_aide_sheet:
        ws_aide = wb.create_sheet(SHEET_AIDE)
        for i, line in enumerate(AIDE_V1_LINES, start=1):
            ws_aide.cell(row=i, column=1, value=line)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def read_lot1_workbook(
    file_obj,
    *,
    expected_columns: set[str] | None = None,
    max_data_rows: int | None = None,
) -> tuple[dict[str, Any], list[str], list[dict[str, Any]]]:
    """
    Lit un classeur Lot 1 : valide Meta, retourne en-têtes et lignes sous forme de dicts.

    Args:
        file_obj: Fichier binaire (``BytesIO``, ``UploadedFile``, chemin supporté par openpyxl).
        expected_columns: Si fourni, contrôle strict des en-têtes (§2.5.2) **avant** le
            parcours des lignes de données.
        max_data_rows: Plafond de lignes de données (défaut : setting ``RAP_IMPORT_MAX_LOT1_DATA_ROWS``).

    Returns:
        tuple: (meta dict, liste d'en-têtes, liste de dicts par ligne de données).

    Raises:
        django.core.exceptions.ValidationError: Si la feuille, les métadonnées ou les en-têtes sont invalides.
    """
    if max_data_rows is None:
        max_data_rows = get_max_lot1_data_rows()
    if hasattr(file_obj, "seek"):
        try:
            file_obj.seek(0)
        except (OSError, AttributeError):
            pass
    try:
        wb = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError(
            {
                "file": "Fichier corrompu ou non compatible (ouverture Excel impossible).",
                "code": ERR_INVALID_FILE,
            }
        ) from exc

    max_sec = get_max_parse_seconds()
    parse_deadline = (time.monotonic() + max_sec) if max_sec > 0 else None

    try:
        if SHEET_DATA not in wb.sheetnames:
            raise ValidationError({"file": f"Feuille « {SHEET_DATA} » introuvable."})
        ws = wb[SHEET_DATA]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            raise ValidationError({"file": "Feuille de données vide."})
        headers = ["" if c is None else str(c).strip() for c in header_row]
        while headers and headers[-1] == "":
            headers.pop()
        if not headers or not any(h for h in headers):
            raise ValidationError({"file": "Ligne d'en-têtes invalide."})

        if expected_columns is not None:
            validate_data_sheet_headers_strict(headers, expected_columns)

        data_rows: list[dict[str, Any]] = []
        for excel_row_idx, row in enumerate(rows_iter, start=2):
            _check_parse_deadline(parse_deadline)
            if all(v is None or (isinstance(v, str) and not str(v).strip()) for v in row):
                continue
            if len(data_rows) >= max_data_rows:
                raise ValidationError(
                    {
                        "file": f"Trop de lignes de données dans la feuille « {SHEET_DATA} » (max. {max_data_rows})."
                    }
                )
            d: dict[str, Any] = {}
            for i, key in enumerate(headers):
                if i < len(row):
                    d[key] = row[i]
                else:
                    d[key] = None
            d["_excel_row_number"] = excel_row_idx
            data_rows.append(d)

        meta: dict[str, Any] = {}
        if SHEET_META in wb.sheetnames:
            m = wb[SHEET_META]
            for row in m.iter_rows(min_row=1, max_col=2, values_only=True):
                _check_parse_deadline(parse_deadline)
                k, v = row[0], row[1] if len(row) > 1 else None
                if k is None:
                    continue
                meta[str(k).strip()] = v
    finally:
        wb.close()

    return meta, headers, data_rows


def assert_meta_matches(meta: dict[str, Any], *, expected_resource: str) -> None:
    """
    Vérifie ``schema_version`` et ``resource`` issus de la feuille Meta.

    Args:
        meta: Dictionnaire lu depuis la feuille Meta.
        expected_resource: Ressource attendue pour l'endpoint appelé.

    Raises:
        django.core.exceptions.ValidationError: Si la version ou la ressource ne correspond pas.
    """
    ver = meta.get("schema_version")
    try:
        ver_int = int(float(ver)) if ver is not None else None
    except (TypeError, ValueError):
        ver_int = None
    if ver_int != SCHEMA_VERSION_LOT1:
        raise ValidationError(
            {
                "file": "Ce fichier a été produit avec une version obsolète du modèle ; téléchargez le modèle à jour.",
                "code": ERR_SCHEMA_MISMATCH,
            }
        )
    res = meta.get("resource")
    if res is None or str(res).strip() != expected_resource:
        raise ValidationError(
            {
                "file": f"Ce fichier ne correspond pas à la ressource attendue ({expected_resource}).",
                "code": ERR_SCHEMA_MISMATCH,
            }
        )
