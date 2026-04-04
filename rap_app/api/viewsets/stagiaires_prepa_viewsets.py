"""ViewSet des stagiaires Prépa."""

from io import BytesIO

from django.db.models import Q
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.utils.timezone import localdate
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.response import Response

from ...models.centres import Centre
from ...models.prepa import Prepa, StagiairePrepa
from ..mixins import HardDeleteArchivedMixin
from ..permissions import IsPrepaStaffOrAbove
from ..roles import is_admin_like, is_candidate, is_prepa_staff, is_staff_read, is_staff_standard
from ..serializers.prepa_serializers import StagiairePrepaSerializer


class StagiairePrepaViewSet(HardDeleteArchivedMixin, viewsets.ModelViewSet):
    """
    CRUD et exports du suivi nominatif des stagiaires Prépa.

    Accès :
    - admin/superadmin : accès global ;
    - staff + staff_read : accès transverse restreint à leurs centres ;
    - prepa_staff : accès principal restreint à leurs centres ;
    - commercial / charge_recrutement / candidats : aucun accès métier.

    Les créations et mises à jour contrôlent à la fois le `centre` direct
    et la `prepa_origine`, afin d'éviter un rattachement inter-centres.
    """

    serializer_class = StagiairePrepaSerializer
    permission_classes = [IsPrepaStaffOrAbove]
    hard_delete_enabled = True
    queryset = StagiairePrepa.objects.select_related("centre", "prepa_origine", "prepa_origine__centre").all()

    def _admin_like(self, user) -> bool:
        return is_admin_like(user)

    def _accessible_centre_ids(self, user):
        if self._admin_like(user):
            return None
        if is_prepa_staff(user) or is_staff_standard(user) or is_staff_read(user):
            centres = getattr(user, "centres", None)
            if not centres or not centres.exists():
                return []
            return list(centres.values_list("id", flat=True))
        return []

    def _assert_user_can_use_centre(self, centre):
        if not centre:
            return
        centre_ids = self._accessible_centre_ids(self.request.user)
        if centre_ids is None:
            return
        if getattr(centre, "id", None) not in set(centre_ids):
            raise PermissionDenied("Centre hors de votre périmètre d'accès.")

    def _assert_user_can_use_prepa_origine(self, prepa):
        if not prepa:
            return
        self._assert_user_can_use_centre(getattr(prepa, "centre", None))

    def _scope_qs(self, qs):
        user = self.request.user
        if not user.is_authenticated or is_candidate(user):
            return qs.none()

        centre_ids = self._accessible_centre_ids(user)
        if centre_ids is None:
            return qs
        if not centre_ids:
            return qs.none()

        model = getattr(qs, "model", None)
        if model and getattr(model._meta, "model_name", "") == "centre":
            return qs.filter(id__in=centre_ids)

        field_names = [f.name for f in model._meta.get_fields()] if model else []
        if "centre_id" in field_names or "centre" in field_names:
            return qs.filter(centre_id__in=centre_ids)

        return qs.none()

    def get_queryset(self):
        qs = self._scope_qs(
            StagiairePrepa.objects.select_related("centre", "prepa_origine", "prepa_origine__centre").all()
        )
        params = self.request.query_params
        truthy = {"1", "true", "yes", "on"}

        include_archived = str(params.get("avec_archivees", "")).lower() in truthy
        archived_only = str(params.get("archives_seules", "")).lower() in truthy

        if archived_only:
            qs = qs.filter(is_active=False)
        elif not include_archived:
            qs = qs.filter(is_active=True)

        search = params.get("search")
        centre = params.get("centre")
        statut = params.get("statut_parcours")
        prepa_origine = params.get("prepa_origine")
        annee = params.get("annee")
        type_atelier = params.get("type_atelier")
        ordering = params.get("ordering") or "nom"

        if search:
            qs = qs.filter(
                Q(nom__icontains=search)
                | Q(prenom__icontains=search)
                | Q(telephone__icontains=search)
                | Q(email__icontains=search)
                | Q(centre__nom__icontains=search)
            )
        if centre:
            qs = qs.filter(centre_id=centre)
        if statut:
            qs = qs.filter(statut_parcours=statut)
        if prepa_origine:
            qs = qs.filter(prepa_origine_id=prepa_origine)
        if annee:
            qs = qs.filter(Q(date_entree_parcours__year=annee) | Q(prepa_origine__date_prepa__year=annee))
        if type_atelier:
            flag_field = StagiairePrepa.atelier_flag_map().get(type_atelier, (None, None))[0]
            if flag_field:
                qs = qs.filter(**{flag_field: True})

        if ordering in {
            "nom",
            "-nom",
            "prenom",
            "-prenom",
            "date_entree_parcours",
            "-date_entree_parcours",
            "date_sortie_parcours",
            "-date_sortie_parcours",
            "updated_at",
            "-updated_at",
        }:
            qs = qs.order_by(ordering, "prenom", "id")
        else:
            qs = qs.order_by("nom", "prenom", "id")

        return qs

    def get_archived_aware_object(self):
        lookup_value = self.kwargs.get(self.lookup_url_kwarg or self.lookup_field)
        base_qs = self._scope_qs(
            StagiairePrepa.objects.select_related("centre", "prepa_origine", "prepa_origine__centre").all()
        )
        return get_object_or_404(base_qs, **{self.lookup_field: lookup_value})

    def destroy(self, request, *args, **kwargs):
        """
        Conserve `DELETE` pour compatibilité mais archive
        logiquement le stagiaire Prépa.
        """
        instance = self.get_object()
        if not instance.is_active:
            return Response(
                {
                    "success": True,
                    "message": "Stagiaire Prépa déjà archivé.",
                    "data": self.get_serializer(instance).data,
                },
                status=status.HTTP_200_OK,
            )

        instance.is_active = False
        instance.save(user=request.user, update_fields=["is_active"])
        return Response(
            {
                "success": True,
                "message": "Stagiaire Prépa archivé avec succès.",
                "data": self.get_serializer(instance).data,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=["post"], url_path="desarchiver")
    def desarchiver(self, request, pk=None):
        """
        Restaure un stagiaire Prépa archivé et renvoie l'enveloppe API standard.
        """
        instance = self.get_archived_aware_object()
        if instance.is_active:
            return Response(
                {
                    "success": True,
                    "message": "Stagiaire Prépa déjà actif.",
                    "data": self.get_serializer(instance).data,
                },
                status=status.HTTP_200_OK,
            )

        instance.is_active = True
        instance.save(user=request.user, update_fields=["is_active"])
        return Response(
            {
                "success": True,
                "message": "Stagiaire Prépa désarchivé avec succès.",
                "data": self.get_serializer(instance).data,
            },
            status=status.HTTP_200_OK,
        )

    def create(self, request, *args, **kwargs):
        """
        Crée un stagiaire Prépa et renvoie l'enveloppe API standard.
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(
            {
                "success": True,
                "message": "Stagiaire Prépa créé avec succès.",
                "data": self.get_serializer(serializer.instance).data,
            },
            status=status.HTTP_201_CREATED,
            headers=headers,
        )

    def retrieve(self, request, *args, **kwargs):
        """
        Retourne le détail d'un stagiaire Prépa dans l'enveloppe API standard.
        """
        instance = self.get_object()
        return Response(
            {
                "success": True,
                "message": "Stagiaire Prépa récupéré avec succès.",
                "data": self.get_serializer(instance).data,
            },
            status=status.HTTP_200_OK,
        )

    def update(self, request, *args, **kwargs):
        """
        Met à jour un stagiaire Prépa et renvoie l'enveloppe API standard.
        """
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(
            {
                "success": True,
                "message": "Stagiaire Prépa mis à jour avec succès.",
                "data": self.get_serializer(serializer.instance).data,
            },
            status=status.HTTP_200_OK,
        )

    def perform_create(self, serializer):
        centre = serializer.validated_data.get("centre")
        prepa_origine = serializer.validated_data.get("prepa_origine")
        self._assert_user_can_use_centre(centre)
        self._assert_user_can_use_prepa_origine(prepa_origine)
        serializer.save()

    def perform_update(self, serializer):
        instance = serializer.instance
        centre = serializer.validated_data.get("centre", getattr(instance, "centre", None))
        prepa_origine = serializer.validated_data.get("prepa_origine", getattr(instance, "prepa_origine", None))
        self._assert_user_can_use_centre(centre)
        self._assert_user_can_use_prepa_origine(prepa_origine)
        serializer.save()

    @action(detail=False, methods=["get"], url_path="meta")
    def meta(self, request):
        """
        Retourne les métadonnées utiles au frontend dans l'enveloppe
        API standard (centres, statuts, ateliers, séances et années).
        """
        centres = self._scope_qs(Centre.objects.all()).order_by("nom")
        annees = (
            self.get_queryset()
            .order_by()
            .values_list("date_entree_parcours__year", flat=True)
            .distinct()
        )
        annees = sorted([a for a in annees if a], reverse=True)
        prepas_origine = (
            self._scope_qs(Prepa.objects.select_related("centre").all())
            .order_by("-date_prepa", "-id")[:200]
        )

        return Response(
            {
                "success": True,
                "message": "Métadonnées stagiaires Prépa récupérées avec succès.",
                "data": {
                    "centres": [
                        {"id": c.id, "nom": c.nom, "departement": c.departement, "code_postal": c.code_postal}
                        for c in centres
                    ],
                    "statut_parcours": [
                        {"value": value, "label": label} for value, label in StagiairePrepa.StatutParcours.choices
                    ],
                    "type_atelier": [
                        {"value": value, "label": label}
                        for value, label in Prepa.TypePrepa.choices
                        if value.startswith("atelier") or value == Prepa.TypePrepa.AUTRE
                    ],
                    "prepas_origine": [
                        {
                            "id": p.id,
                            "label": f"{p.get_type_prepa_display()} du {p.date_prepa:%d/%m/%Y}"
                            + (f" - {p.centre.nom}" if p.centre else ""),
                        }
                        for p in prepas_origine
                    ],
                    "annees": annees or [localdate().year],
                },
            },
            status=status.HTTP_200_OK,
        )

    def _filtered_export_qs(self, request):
        qs = self.filter_queryset(self.get_queryset())
        if request.method.lower() == "post":
            ids = request.data.get("ids") or []
            if ids:
                qs = qs.filter(id__in=ids)
        return qs

    @action(detail=False, methods=["get", "post"], url_path="export-xlsx")
    def export_xlsx(self, request):
        qs = self._filtered_export_qs(request)

        wb = Workbook()
        ws = wb.active
        ws.title = "Stagiaires Prepa"
        ws.append(
            [
                "Nom",
                "Prénom",
                "Téléphone",
                "Email",
                "Centre",
                "Prépa d'origine",
                "Statut",
                "Ateliers réalisés",
                "Dernier atelier",
                "Date d'entrée",
                "Date de sortie",
                "Motif abandon",
            ]
        )

        for obj in qs:
            ws.append(
                [
                    obj.nom,
                    obj.prenom,
                    obj.telephone or "",
                    obj.email or "",
                    getattr(obj.centre, "nom", ""),
                    str(obj.prepa_origine) if obj.prepa_origine else "",
                    obj.get_statut_parcours_display(),
                    ", ".join(obj.ateliers_realises_labels),
                    obj.dernier_atelier_label or "",
                    obj.date_entree_parcours.strftime("%d/%m/%Y") if obj.date_entree_parcours else "",
                    obj.date_sortie_parcours.strftime("%d/%m/%Y") if obj.date_sortie_parcours else "",
                    obj.motif_abandon or "",
                ]
            )

        self._style_sheet(ws)
        return self._xlsx_response(wb, "stagiaires_prepa.xlsx")

    @action(detail=False, methods=["get", "post"], url_path="export-emargement-xlsx")
    def export_emargement_xlsx(self, request):
        qs = self._filtered_export_qs(request)
        wb = Workbook()
        ws = wb.active
        ws.title = "Emargement Prepa"

        title = "Feuille d'émargement - Stagiaires Prépa"
        type_atelier = request.query_params.get("type_atelier")
        if type_atelier:
            try:
                title = f"{title} - {Prepa.TypePrepa(type_atelier).label}"
            except ValueError:
                pass

        ws.merge_cells("A1:H1")
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.append([])
        ws.append(["Nom", "Prénom", "Centre", "Statut", "Téléphone", "Email", "Présence", "Signature"])

        for obj in qs:
            ws.append(
                [
                    obj.nom,
                    obj.prenom,
                    getattr(obj.centre, "nom", ""),
                    obj.get_statut_parcours_display(),
                    obj.telephone or "",
                    obj.email or "",
                    "",
                    "",
                ]
            )

        self._style_sheet(ws, header_row=3)
        return self._xlsx_response(wb, "stagiaires_prepa_emargement.xlsx")

    @action(detail=False, methods=["get", "post"], url_path="export-presence-xlsx")
    def export_presence_xlsx(self, request):
        qs = self._filtered_export_qs(request)
        wb = Workbook()
        ws = wb.active
        ws.title = "Presence Prepa"

        title = "Feuille de présence - Stagiaires Prépa"
        type_atelier = request.query_params.get("type_atelier")
        if type_atelier:
            try:
                title = f"{title} - {Prepa.TypePrepa(type_atelier).label}"
            except ValueError:
                pass

        ws.merge_cells("A1:G1")
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.append([])
        ws.append(["Nom", "Prénom", "Centre", "Statut", "Téléphone", "Email", "Présent"])

        for obj in qs:
            ws.append(
                [
                    obj.nom,
                    obj.prenom,
                    getattr(obj.centre, "nom", ""),
                    obj.get_statut_parcours_display(),
                    obj.telephone or "",
                    obj.email or "",
                    "",
                ]
            )

        self._style_sheet(ws, header_row=3)
        return self._xlsx_response(wb, "stagiaires_prepa_presence.xlsx")

    def _style_sheet(self, ws, header_row=1):
        fill = PatternFill("solid", fgColor="DCE6F1")
        border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="002060")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top")

        for col in ws.columns:
            if not col:
                continue
            letter = get_column_letter(col[0].column)
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[letter].width = min(max_len + 2, 40)

    def _xlsx_response(self, wb, filename):
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
