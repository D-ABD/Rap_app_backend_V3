# rap_app/api/viewsets/declic_objectifs_viewset.py

from collections import OrderedDict
from io import BytesIO

from django.http import HttpResponse
from drf_spectacular.utils import (
    OpenApiParameter,
    extend_schema,
    extend_schema_view,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from ...api.roles import (
    is_admin_like,
    is_candidate,
    is_declic_staff,
    is_staff_or_staffread,
)
from ...models.centres import Centre
from ...models.declic import Declic, ObjectifDeclic  # 🔁 on importe aussi Declic
from ..mixins import ApiResponseMixin
from ..permissions import IsDeclicStaffOrAbove
from ..serializers.declic_objectifs_serializers import ObjectifDeclicSerializer


@extend_schema_view(
    list=extend_schema(
        summary="Lister tous les objectifs Déclic",
        description="Retourne la liste paginée des objectifs Déclic (format DRF standard).",
        parameters=[
            OpenApiParameter(
                name="annee",
                description="Filtrer par année",
                required=False,
                type=int,
            ),
            OpenApiParameter(
                name="centre_id",
                description="Filtrer par identifiant de centre",
                required=False,
                type=int,
            ),
            OpenApiParameter(
                name="departement",
                description="Filtrer par code département (ex: 59, 75)",
                required=False,
                type=str,
            ),
        ],
        responses={200: ObjectifDeclicSerializer(many=True)},
    ),
)
class ObjectifDeclicViewSet(ApiResponseMixin, viewsets.ModelViewSet):
    """
    ViewSet CRUD des objectifs annuels Déclic.

    Le périmètre est restreint par centre selon le rôle :
    admin-like, profils Déclic, staff/staff_read. Le fichier conserve un
    helper local de scope historique et expose en plus des actions utilitaires
    de filtres, synthèse et export Excel.
    """

    serializer_class = ObjectifDeclicSerializer
    permission_classes = [IsAuthenticated, IsDeclicStaffOrAbove]
    queryset = ObjectifDeclic.objects.select_related("centre").all()
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["centre__nom", "centre__code_postal", "annee"]
    ordering_fields = ["annee", "centre__nom"]

    # -----------------------------------------------------------
    # 🔹 Helpers de périmètre (scope)
    # -----------------------------------------------------------
    def _centre_ids_for_user(self, user):
        # Règle : renvoie None si admin-like (accès total)
        #         sinon, limite aux centres autorisés (centres_acces ou centres selon rôle)
        #         staff non "declic" : accès centré sur user.centres (ManyToMany)
        #         Tous les autres cas : accès interdit (liste vide).
        if is_admin_like(user):
            return None
        if is_declic_staff(user):
            centres = getattr(user, "centres_acces", None) or getattr(user, "centres", None)
            return list(centres.values_list("id", flat=True)) if centres else []
        if is_staff_or_staffread(user):
            try:
                return list(user.centres.values_list("id", flat=True))
            except Exception:
                return []
        return []

    def _scope_qs_to_user_centres(self, qs):
        # Applique le filtrage de visibilité selon le rôle de l'utilisateur.
        user = self.request.user
        if not user.is_authenticated or is_candidate(user):
            return qs.none()
        centre_ids = self._centre_ids_for_user(user)
        if centre_ids is None:
            return qs
        if centre_ids:
            return qs.filter(centre_id__in=centre_ids).distinct()
        return qs.none()

    def _assert_user_can_use_centre(self, centre):
        # Check de sécurité : bloquer toute tentative de création/màj hors périmètre permise.
        if not centre:
            return
        user = self.request.user
        if is_admin_like(user):
            return
        allowed = set(self._centre_ids_for_user(user) or [])
        if getattr(centre, "id", None) not in allowed:
            raise PermissionDenied("Centre hors de votre périmètre d'accès.")

    # -----------------------------------------------------------
    # 🔹 Queryset principal
    # -----------------------------------------------------------
    def get_queryset(self):
        """
        Retourne le queryset d'objectifs Déclic visibles par l'utilisateur actuel.

        - Si admin-like : accès total à tous les ObjectifDeclic.
        - Si declic_staff : accès restreint aux centres autorisés (user.centres_acces ou user.centres).
        - Si staff / staff_read : accès uniquement à user.centres (ManyToMany).
        - Si non authentifié ou candidat : aucun accès.

        Filtres complémentaires appliqués selon la query string :
            - "annee" : filtre exacte sur le champ annee.
            - "centre_id" : filtre exacte sur le centre lié.
            - "departement" : filtre via le préfixe du code_postal du centre.

        Trie le résultat par année décroissante puis nom de centre.

        La structure exacte du résultat dépend du serializer.
        """
        qs = ObjectifDeclic.objects.select_related("centre")
        qs = self._scope_qs_to_user_centres(qs)

        params = self.request.query_params
        annee = params.get("annee")
        centre_id = params.get("centre_id")
        departement = params.get("departement")

        if annee:
            qs = qs.filter(annee=annee)
        if centre_id:
            qs = qs.filter(centre_id=centre_id)
        if departement:
            # on reste sur le préfixe CP (comme avant), cohérent avec le front
            qs = qs.filter(centre__code_postal__startswith=departement)

        return qs.order_by("-annee", "centre__nom")

    # -----------------------------------------------------------
    # 🔹 list() — pagination DRF standard
    # -----------------------------------------------------------
    def list(self, request, *args, **kwargs):
        """
        Renvoie une liste paginée d’ObjectifDeclic accessibles (DRF : count, next, previous, results).
        - Permissions : IS_AUTHENTICATED et IsDeclicStaffOrAbove + filtrage par périmètre utilisateur.
        - Format : pagination DRF standard, items serializés via ObjectifDeclicSerializer.
        """
        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return self.success_response(data=serializer.data, message="Liste des objectifs Déclic récupérée avec succès.")

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return self.success_response(data=serializer.data, message="Objectif Déclic récupéré avec succès.")

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return self.success_response(
            data=self.get_serializer(serializer.instance).data,
            message="Objectif Déclic créé avec succès.",
            status_code=status.HTTP_201_CREATED,
        )

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return self.success_response(
            data=self.get_serializer(serializer.instance).data,
            message="Objectif Déclic mis à jour avec succès.",
        )

    def partial_update(self, request, *args, **kwargs):
        kwargs["partial"] = True
        return self.update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return self.success_response(data=None, message="Objectif Déclic supprimé avec succès.")

    # -----------------------------------------------------------
    # 🔹 create / update sécurisés
    # -----------------------------------------------------------
    def perform_create(self, serializer):
        """
        Sécurise la création : interdit la création sur un centre non autorisé.
        Le format de la réponse dépend du serializer.
        """
        self._assert_user_can_use_centre(serializer.validated_data.get("centre"))
        serializer.save()

    def perform_update(self, serializer):
        """
        Sécurise la mise à jour : bloque si le centre cible échappe au périmètre de l’utilisateur.
        Réponse conforme au serializer.
        """
        current = serializer.instance
        new_centre = serializer.validated_data.get("centre", getattr(current, "centre", None))
        self._assert_user_can_use_centre(new_centre)
        serializer.save()

    # -----------------------------------------------------------
    # 🔹 Filtres
    # -----------------------------------------------------------
    @action(detail=False, methods=["get"], url_path="filters")
    def filters(self, request):
        """
        Retourne les listes de choix (année, centre, département) en fonction du périmètre actuel de l’utilisateur.
        Méthode : GET.
        Permissions : identiques au ViewSet.
        Format JSON : voir exemple ci-dessous.
          {
              "annee": [{"value": ..., "label": ...}, ...],
              "centre": [{"value": ..., "label": ...}, ...],
              "departement": [{"value": ..., "label": ...}, ...]
          }
        """
        qs = self._scope_qs_to_user_centres(ObjectifDeclic.objects.select_related("centre"))

        annees = qs.order_by("-annee").values_list("annee", flat=True).distinct()
        centres = qs.values("centre__id", "centre__nom", "centre__code_postal").distinct()
        departements = sorted({(c["centre__code_postal"] or "")[:2] for c in centres if c.get("centre__code_postal")})

        data = OrderedDict(
            annee=[{"value": a, "label": str(a)} for a in annees],
            centre=[
                {
                    "value": c["centre__id"],
                    "label": f'{c["centre__nom"]} ({c["centre__code_postal"]})',
                }
                for c in centres
            ],
            departement=[{"value": d, "label": f"Département {d}"} for d in departements],
        )
        return self.success_response(data=data, message="Filtres objectifs Déclic récupérés avec succès.")

    # -----------------------------------------------------------
    # 🔹 Synthèse
    # -----------------------------------------------------------
    @action(detail=False, methods=["get"], url_path="synthese")
    def synthese(self, request):
        """
        Retourne la synthèse globale des ObjectifDeclic visibles par l’utilisateur.
        Méthode : GET.
        Permissions : identiques au ViewSet.
        Réponse : liste d’objets définis via obj.synthese_globale().
        Format JSON dépendant du modèle (non documenté ici).
        """
        qs = self.get_queryset()
        data = [obj.synthese_globale() for obj in qs]
        return self.success_response(data=data, message="Synthèse objectifs Déclic récupérée avec succès.")

    # -----------------------------------------------------------
    # 🔹 Export Excel
    # -----------------------------------------------------------
    @action(detail=False, methods=["get"], url_path="export-xlsx")
    def export_xlsx(self, request):
        """
        Export Excel des ObjectifDeclic filtrés selon permissions et filtres en cours.
        - Méthode : GET.
        - Permissions : identiques au ViewSet.
        - Réponse : Fichier xlsx, entête Content-Disposition 'attachment; filename="objectifs_declic.xlsx"'.
          Colonnes :
            * Centre
            * Département
            * Année
            * Objectif
            * Réalisé (tous ateliers cumulés)
            * Taux atteinte (%)
            * Taux rétention (%) (calculé par Declic.taux_retention)
            * Reste à faire
        - Si aucun résultat : JSON 404 {"detail": "Aucun objectif à exporter."}
        - Format/structure des lignes dépend du serializer/model (cf. synthese_globale()).
        """
        qs = self.get_queryset()
        if not qs.exists():
            return Response(
                {"success": False, "message": "Aucun objectif à exporter.", "data": None},
                status=status.HTTP_404_NOT_FOUND,
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Objectifs Déclic"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # === En-têtes ===
        headers = [
            "Centre",
            "Département",
            "Année",
            "Objectif",
            "Réalisé (tous ateliers cumulés)",  # ✅ libellé mis à jour
            "Taux atteinte (%)",
            "Taux rétention (%)",  # 🆕 calculé avec Declic.taux_retention
            "Reste à faire",
        ]

        ws.append(headers)
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # === Données ===
        for obj in qs:
            data = obj.synthese_globale()
            # 🔢 Rétention A1 → A6 calculée côté modèle Déclic
            taux_retention = Declic.taux_retention(obj.centre, obj.annee) if obj.centre and obj.annee else 0

            ws.append(
                [
                    data["centre"],
                    obj.centre.departement,
                    data["annee"],
                    data["objectif"],
                    data["realise"],  # 🔁 maintenant = total ateliers (1→6 + autre)
                    data["taux_presence"],
                    data["taux_adhesion"],
                    data["taux_atteinte"],
                    taux_retention,
                    data["reste_a_faire"],
                ]
            )

        # === Ajustement largeur colonnes ===
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        )
        response["Content-Disposition"] = 'attachment; filename="objectifs_declic.xlsx"'
        return response
