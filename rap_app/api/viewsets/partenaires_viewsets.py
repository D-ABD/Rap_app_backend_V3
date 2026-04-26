"""ViewSet principal des partenaires."""

import datetime
from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.core.exceptions import ValidationError as DjangoValidationError
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone as dj_timezone
from django_filters.rest_framework import DjangoFilterBackend, FilterSet
from django_filters.rest_framework import filters as dj_filters
from drf_spectacular.utils import OpenApiResponse, extend_schema, extend_schema_view
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied, ValidationError as DRFValidationError
from rest_framework.permissions import SAFE_METHODS, BasePermission
from rest_framework.response import Response

from ...api.permissions import (
    IsOwnerOrStaffOrAbove,
    UserVisibilityScopeMixin,
)
from ...api.roles import is_admin_like, is_candidate, is_centre_scoped_staff, is_staff_read
from ...models.logs import LogUtilisateur
from ...models.partenaires import Partenaire
from ..mixins import ApiResponseMixin, HardDeleteArchivedMixin
from ..serializers.partenaires_serializers import (
    PartenaireChoicesResponseSerializer,
    PartenaireSerializer,
)

# -------------------- helpers accès candidat (alignés sur ``get_queryset``) --------------------


def _request_cible_reafficher_dans_ma_liste(request, view) -> bool:
    """
    DRF n’alimente parfois pas ``view.action`` assez tôt : on complète par le path.
    """
    if getattr(view, "action", None) == "reafficher_dans_ma_liste":
        return True
    path = (getattr(request, "path", "") or "").lower()
    return "reafficher-dans-ma-liste" in path


def _candidat_lie_pour_fiche_liste(user, partenaire: Partenaire) -> bool:
    """
    Même idée de rattachement que le filtre candidat (sans exclusion retrait) :
    créateur, prospections, dernière maj, saisi_par.
    Sert retrait de liste, réaffichage, et cohérence des permissions.
    """
    if not (user and getattr(user, "is_authenticated", False) and getattr(user, "pk", None)):
        return False
    if partenaire.created_by_id == user.id:
        return True
    if partenaire.updated_by_id == user.id:
        return True
    if partenaire.saisi_par.filter(id=user.id).exists():
        return True
    if partenaire.prospections.filter(owner_id=user.id).exists():
        return True
    return False


# -------------------- permission locale --------------------


class PartenaireAccessPermission(BasePermission):
    """
    Permission d'accès aux instances de Partenaire en fonction de
    l'authentification, du rôle (admin/staff) et de la relation avec
    le partenaire ou ses prospections liées.
    """

    message = "Accès restreint."

    def has_permission(self, request, view):
        user = request.user
        if not (user and user.is_authenticated):
            return False

        if request.method in SAFE_METHODS:
            return True

        action = getattr(view, "action", None)

        # Import Excel : réservé au staff (même règle qu’avant, y compris §2.15).
        if action == "import_xlsx":
            return is_admin_like(user) or (is_centre_scoped_staff(user) and not is_staff_read(user))

        # Création : staff inchangé ; candidats / stagiaires alignés sur `perform_create`
        # (centre issu de la formation via `user.centre`, obligatoire).
        if action == "create":
            if is_admin_like(user) or (is_centre_scoped_staff(user) and not is_staff_read(user)):
                return True
            if is_candidate(user) and getattr(user, "centre", None) is not None:
                return True
            return False

        return True

    def has_object_permission(self, request, view, obj):
        user = request.user
        if is_admin_like(user):
            return True

        if is_centre_scoped_staff(user):
            if is_staff_read(user):
                return request.method in SAFE_METHODS
            return True

        # Accès complet si créateur de ce partenaire
        if getattr(obj, "created_by_id", None) == user.id:
            return True

        # Candidat / parcours : a « réouvert » une fiche existante (réutilisation doublon) → updated_by
        # ou enregistrement explicite dans saisi_par (plusieurs candidats, même fiche)
        if getattr(obj, "updated_by_id", None) == user.id:
            return True
        if hasattr(obj, "saisi_par") and obj.saisi_par.filter(id=user.id).exists():
            return True

        # Accès en lecture seule si utilisateur propriétaire d'une prospection liée à ce partenaire
        if request.method in SAFE_METHODS and hasattr(obj, "prospections"):
            try:
                return obj.prospections.filter(owner=user).exists()
            except Exception:
                return False

        # Retrait de liste (DELETE) / retour de liste (POST) : toute personne
        # « liée » comme en liste (p. ex. seulement un propriétaire de prospection)
        if not (is_admin_like(user) or is_centre_scoped_staff(user)) and _candidat_lie_pour_fiche_liste(
            user, obj
        ):
            act = getattr(view, "action", None)
            if (request.method == "DELETE" and act == "destroy") or (
                request.method == "POST" and _request_cible_reafficher_dans_ma_liste(request, view)
            ):
                return True

        return False


class InlinePartenaireFilter(FilterSet):
    """
    Jeu de filtres pour les partenaires incluant champs standards
    (type, activité, localisation, créateur, centre) et indicateurs
    booléens sur l'existence de relations (appairages, prospections,
    formations, candidats).
    """

    type = dj_filters.CharFilter(lookup_expr="exact")
    is_active = dj_filters.BooleanFilter()
    city = dj_filters.CharFilter(lookup_expr="icontains")
    secteur_activite = dj_filters.CharFilter(lookup_expr="icontains")
    created_by = dj_filters.NumberFilter()
    # ✅ filtrer par centre par défaut du partenaire
    centre_id = dj_filters.NumberFilter(field_name="default_centre_id")

    has_appairages = dj_filters.BooleanFilter(method="filter_has_appairages")
    has_prospections = dj_filters.BooleanFilter(method="filter_has_prospections")
    has_formations = dj_filters.BooleanFilter(method="filter_has_formations")
    has_candidats = dj_filters.BooleanFilter(method="filter_has_candidats")

    class Meta:
        model = Partenaire
        fields = [
            "type",
            "is_active",
            "city",
            "secteur_activite",
            "created_by",
            "centre_id",
            "has_appairages",
            "has_prospections",
            "has_formations",
            "has_candidats",
        ]

    def _bool(self, value):
        if isinstance(value, bool):
            return value
        if value is None:
            return None
        s = str(value).strip().lower()
        return s in {"1", "true", "yes", "y", "on"}

    def filter_has_appairages(self, queryset, name, value):
        b = self._bool(value)
        if b is None:
            return queryset
        return queryset.filter(appairages_count__gt=0) if b else queryset.filter(appairages_count=0)

    def filter_has_prospections(self, queryset, name, value):
        b = self._bool(value)
        if b is None:
            return queryset
        return queryset.filter(prospections_count__gt=0) if b else queryset.filter(prospections_count=0)

    def filter_has_formations(self, queryset, name, value):
        b = self._bool(value)
        if b is None:
            return queryset
        return queryset.filter(formations_count__gt=0) if b else queryset.filter(formations_count=0)

    def filter_has_candidats(self, queryset, name, value):
        b = self._bool(value)
        if b is None:
            return queryset
        return queryset.filter(candidats_count__gt=0) if b else queryset.filter(candidats_count=0)


@extend_schema_view(
    list=extend_schema(
        summary="Lister les partenaires",
        tags=["Partenaires"],
        responses={200: OpenApiResponse(response=PartenaireSerializer)},
    ),
    retrieve=extend_schema(
        summary="Détail d’un partenaire",
        tags=["Partenaires"],
        responses={200: OpenApiResponse(response=PartenaireSerializer)},
    ),
    create=extend_schema(
        summary="Créer un partenaire",
        tags=["Partenaires"],
        responses={201: OpenApiResponse(description="Création réussie")},
    ),
    update=extend_schema(
        summary="Modifier un partenaire",
        tags=["Partenaires"],
        responses={200: OpenApiResponse(description="Mise à jour réussie")},
    ),
    destroy=extend_schema(
        summary="Archiver un partenaire",
        tags=["Partenaires"],
        responses={204: OpenApiResponse(description="Archivage réussi")},
    ),
)
class PartenaireViewSet(HardDeleteArchivedMixin, ApiResponseMixin, UserVisibilityScopeMixin, viewsets.ModelViewSet):
    """
    ViewSet gérant les partenaires avec opérations CRUD, filtrage,
    permissions de périmètre et actions utilitaires (métadonnées,
    statistiques de relations et export XLSX).
    """

    serializer_class = PartenaireSerializer
    hard_delete_enabled = True
    # ✅ utilise la permission locale pour autoriser la lecture des partenaires attribués via prospection
    permission_classes = [PartenaireAccessPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = InlinePartenaireFilter
    ordering_fields = ["nom", "created_at", "default_centre__nom"]  # ✅ tri centre
    ordering = ["nom"]
    search_fields = [
        "nom",
        "secteur_activite",
        "street_name",
        "zip_code",
        "city",
        "country",
        "contact_nom",
        "contact_poste",
        "contact_email",
        "contact_telephone",
        "website",
        "social_network_url",
        "description",
        "action_description",
        "actions",
        "created_by__first_name",
        "created_by__last_name",
        "created_by__username",
        "default_centre__nom",  # ✅ recherche centre
    ]

    # -------------------- helpers scope --------------------

    def _is_admin_like(self, user) -> bool:
        """
        Indique si un utilisateur est traité comme administrateur
        (superuser ou utilisateur dont la méthode is_admin() renvoie True).
        """
        return is_admin_like(user)

    def _staff_centre_ids(self, user):
        """
        Retourne les identifiants de centres accessibles pour un
        utilisateur staff ou staff_read, ou None pour les profils
        admin-like.
        """
        if self._is_admin_like(user):
            return None  # accès global
        if is_centre_scoped_staff(user):
            return list(user.centres.values_list("id", flat=True))
        return []  # non-staff

    def _scoped_for_user(self, qs, user):
        """
        Restreint un queryset de partenaires au périmètre des centres
        accessibles à l'utilisateur staff ou staff_read.

        Règles (pas d’extension via ``saisi_par``) :
        - appairage / prospection seulement si la **formation** liée a son centre
          dans le périmètre du staff ;
        - ``default_centre`` dans le périmètre ;
        - partenaire créé par l’utilisateur staff.

        Le partage multi-candidats (``saisi_par``) n’ouvre **pas** la fiche côté staff
        hors de ces liens, pour ne pas outrepasser le scope centre.
        """
        centre_ids = self._staff_centre_ids(user)

        # Cas admin/superadmin : accès global
        if centre_ids is None:
            return qs

        # Cas staff sans centre rattaché → uniquement ses créations
        if not centre_ids:
            return qs.filter(created_by=user)

        scoped = qs.filter(
            Q(appairages__formation__centre_id__in=centre_ids)
            | Q(prospections__formation__centre_id__in=centre_ids)
            | Q(default_centre_id__in=centre_ids)
            | Q(created_by=user)
        ).distinct()

        # 🔒 Fiches sans default_centre : exclues si aucun lien de périmètre (voir docstring)
        return scoped.exclude(
            Q(default_centre_id__isnull=True)
            & ~Q(created_by=user)
            & ~Q(appairages__formation__centre_id__in=centre_ids)
            & ~Q(prospections__formation__centre_id__in=centre_ids)
        ).distinct()

    def _user_can_access_partenaire(self, partenaire, user) -> bool:
        """
        Vérifie qu'un utilisateur a le droit d'accéder à un partenaire
        en fonction de son rôle et des centres liés au partenaire.
        """
        if self._is_admin_like(user):
            return True
        if not is_centre_scoped_staff(user):
            # non-staff : déjà géré par permission/queryset
            return True
        centre_ids = set(user.centres.values_list("id", flat=True))
        if not centre_ids:
            return partenaire.created_by_id == user.id
        linked = (
            partenaire.appairages.filter(formation__centre_id__in=centre_ids).exists()
            or partenaire.prospections.filter(formation__centre_id__in=centre_ids).exists()
            or (partenaire.default_centre_id in centre_ids)
        )
        return linked or (partenaire.created_by_id == user.id)

    def _exclut_retrait_liste_perso(self, qs, user):
        """
        Retire les partenaires marqués « hors de ma liste » pour cet utilisateur.

        On évite ``.exclude(retrait_de_vue_par=user)`` sur un queryset déjà annoté
        (Count, OR sur saisi_par, etc.) : certaines bases lèvent une erreur SQL.
        ``exclude(pk__in= sous-requête)`` est suffisant et stable.
        """
        if not getattr(user, "is_authenticated", False) or not getattr(user, "pk", None):
            return qs
        hidden_qs = user.partenaires_masques_pour_moi.values_list("pk", flat=True)
        return qs.exclude(pk__in=hidden_qs)

    # -------------------- queryset --------------------

    def get_queryset(self):
        """
        Retourne les partenaires actifs visibles pour l'utilisateur
        courant, restreints par rôle et périmètre de centres et annotés
        avec des compteurs de relations.
        """
        # ✅ Évite les erreurs DRF Spectacular pendant la génération de schéma
        if getattr(self, "swagger_fake_view", False):
            return Partenaire.objects.none()

        user = self.request.user

        qs = (
            Partenaire.objects.all()
            .select_related("created_by", "default_centre")  # ✅ pas de N+1 sur centre
            .prefetch_related("retrait_de_vue_par")  # ``retrait_dans_ma_liste`` au sérialiseur
            .annotate(
                prospections_count=Count("prospections", distinct=True),
                appairages_count=Count("appairages", distinct=True),
                formations_count=(
                    Count(
                        "appairages__formation",
                        filter=Q(appairages__formation__isnull=False),
                        distinct=True,
                    )
                    + Count(
                        "prospections__formation",
                        filter=Q(prospections__formation__isnull=False),
                        distinct=True,
                    )
                ),
                candidats_count=Count("appairages__candidat", distinct=True),
            )
        )

        include_archived = str(self.request.query_params.get("avec_archivees", "")).lower() in {
            "1",
            "true",
            "yes",
            "on",
        }
        archives_seules = str(self.request.query_params.get("archives_seules", "")).lower() in {
            "1",
            "true",
            "yes",
            "on",
        }

        if archives_seules:
            qs = qs.filter(is_active=False)
        elif not include_archived:
            # Par défaut : fiches actives. Exception : fiches archivées globales visibles pour
            # le créateur **ou** un compte en co-saisie (``saisi_par``) — accès détail / édition
            # + désarchivage, aligné sur la liste « inclure archivés » et l’action ``desarchiver``.
            if getattr(user, "is_authenticated", False) and getattr(user, "pk", None):
                qs = (
                    qs.filter(
                        Q(is_active=True)
                        | (
                            Q(is_active=False)
                            & (Q(created_by=user) | Q(saisi_par=user))
                        )
                    )
                    .distinct()
                )
            else:
                qs = qs.filter(is_active=True)

        # ✅ Admins : tout voir
        if self._is_admin_like(user):
            return qs

        # ✅ Staff ou StaffRead : visibilité restreinte à leur périmètre
        if is_centre_scoped_staff(user):
            return self._scoped_for_user(qs, user)

        # ✅ Candidat·e : créés par lui/elle, prospections, last update, saisi_par
        cand = (
            qs.filter(
                Q(created_by=user)
                | Q(prospections__owner=user)
                | Q(updated_by=user)
                | Q(saisi_par=user)
            )
            .distinct()
        )
        # Retrait de liste perso : uniquement pour la **liste** (et options de filtres), pas pour
        # retrieve/update/… — sinon un GET /partenaires/{id}/ sans ``?avec_archivees=1`` renvoie
        # 404 alors que la même fiche apparaît avec « inclure archivés » (retrait ignoré en liste).
        #
        # Ne pas faire ``or "list"`` : si ``action`` est encore None (certaines actions custom
        # ou le chaînage vers ``get_object``), on ne doit pas appliquer l’exclusion retrait,
        # sinon ``POST …/reafficher-dans-ma-liste/`` ne trouve pas la fiche (404) pour l’utilisateur
        # qui l’a justement retirée de sa liste.
        action = getattr(self, "action", None)
        retrait_pour_cette_action = action in ("list", "filter_options")
        if not include_archived and retrait_pour_cette_action:
            cand = self._exclut_retrait_liste_perso(cand, user)
        return cand

    def get_archived_aware_object(self):
        """
        Récupère un partenaire (y compris archivé) pour ``desarchiver``.

        Pas d’exclusion « retrait de liste perso » ici : un POST sans query string
        doit rester résolvable comme le détail (sinon 404 alors que la ligne est
        visible avec « inclure archivés »).
        """
        lookup_value = self.kwargs.get(self.lookup_url_kwarg or self.lookup_field)
        user = self.request.user
        queryset = (
            Partenaire.objects.all()
            .select_related("created_by", "default_centre")
            .prefetch_related("retrait_de_vue_par")
            .annotate(
                prospections_count=Count("prospections", distinct=True),
                appairages_count=Count("appairages", distinct=True),
                formations_count=(
                    Count(
                        "appairages__formation",
                        filter=Q(appairages__formation__isnull=False),
                        distinct=True,
                    )
                    + Count(
                        "prospections__formation",
                        filter=Q(prospections__formation__isnull=False),
                        distinct=True,
                    )
                ),
                candidats_count=Count("appairages__candidat", distinct=True),
            )
        )

        if self._is_admin_like(user):
            scoped_qs = queryset
        elif is_centre_scoped_staff(user):
            scoped_qs = self._scoped_for_user(queryset, user)
        else:
            scoped_qs = (
                queryset.filter(
                    Q(created_by=user)
                    | Q(prospections__owner=user)
                    | Q(updated_by=user)
                    | Q(saisi_par=user)
                )
                .distinct()
            )

        return get_object_or_404(scoped_qs, **{self.lookup_field: lookup_value})

    # -------------------- endpoints utilitaires --------------------

    @action(detail=False, methods=["get"], url_path="choices")
    def choices(self, request):
        """
        Retourne les valeurs possibles pour les champs de type et
        d'action des partenaires, sérialisées pour le frontend.
        """
        types = [{"value": k, "label": v} for k, v in Partenaire.TYPE_CHOICES]
        actions = [{"value": k, "label": v} for k, v in Partenaire.CHOICES_TYPE_OF_ACTION]
        ser = PartenaireChoicesResponseSerializer(instance={"types": types, "actions": actions})
        return self.success_response(data=ser.data, message="Choix partenaires récupérés avec succès.")

    @extend_schema(summary="🔽 Filtres disponibles pour les partenaires", tags=["Partenaires"])
    @action(detail=False, methods=["get"], url_path="filter-options")
    def filter_options(self, request):
        """
        Fournit les valeurs distinctes utilisables comme options de
        filtre (villes, secteurs, auteurs, centres) pour les partenaires
        visibles de l'utilisateur.
        """
        scoped = self.filter_queryset(self.get_queryset())
        # Ne pas enchaîner .values() / .distinct() sur le queryset scopé avec annotations
        # et jointures (saisi_par, retrait, etc.) : erreurs SQL 500 sur certains moteurs.
        # On repart des pk du périmètre puis un queryset « léger » sur Partenaire.
        base = Partenaire.objects.filter(pk__in=scoped).select_related("created_by", "default_centre")

        villes = (
            base.exclude(city__isnull=True)
            .exclude(city="")  # ✅ évite l’avertissement Pylance
            .values_list("city", flat=True)
            .distinct()
        )
        secteurs = (
            base.exclude(secteur_activite__isnull=True)
            .exclude(secteur_activite="")  # ✅ évite l’avertissement Pylance
            .values_list("secteur_activite", flat=True)
            .distinct()
        )
        users = (
            base.exclude(created_by__isnull=True)
            .values("created_by")
            .distinct()
            .values("created_by", "created_by__first_name", "created_by__last_name")
        )
        # ✅ centres par défaut disponibles
        centres = base.filter(default_centre__isnull=False).values("default_centre_id", "default_centre__nom").distinct()

        return self.success_response(
            data={
                "cities": [{"value": v, "label": v} for v in villes],
                "secteurs": [{"value": s, "label": s} for s in secteurs],
                "users": [
                    {
                        "id": u["created_by"],
                        "full_name": " ".join(
                            filter(None, [u.get("created_by__first_name"), u.get("created_by__last_name")])
                        ).strip(),
                    }
                    for u in users
                    if u["created_by"]
                ],
                "centres": [{"id": c["default_centre_id"], "nom": c["default_centre__nom"]} for c in centres],
            },
            message="Filtres partenaires récupérés avec succès.",
        )

    # -------------------- CRUD --------------------

    def create(self, request, *args, **kwargs):
        """
        Crée un partenaire et renvoie l'enveloppe API standard pour
        homogénéiser le contrat avec les autres endpoints métier.
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        instance = serializer.instance
        headers = self.get_success_headers(serializer.data)
        return Response(
            {
                "success": True,
                "message": "Partenaire créé avec succès.",
                "data": self.get_serializer(instance).data,
            },
            status=status.HTTP_201_CREATED,
            headers=headers,
        )

    def retrieve(self, request, *args, **kwargs):
        """
        Retourne le détail d'un partenaire dans l'enveloppe API
        standard afin d'aligner lecture et écriture.
        """
        instance = self.get_object()
        return self.success_response(
            data=self.get_serializer(instance).data,
            message="Partenaire récupéré avec succès.",
        )

    def perform_create(self, serializer):
        """
        Crée un partenaire en contrôlant le centre par défaut en
        fonction du rôle de l'utilisateur et en renseignant created_by.
        """
        user = self.request.user
        default_centre = None

        # 🔎 Cas 1 — Superadmin / Admin : centre optionnel
        if self._is_admin_like(user):
            default_centre = serializer.validated_data.get("default_centre")

        # 🔎 Cas 2 — Staff / Staff_read : doit avoir un centre autorisé
        elif is_centre_scoped_staff(user):
            centres = list(user.centres.all())
            if not centres:
                raise PermissionDenied("❌ Vous n’êtes rattaché à aucun centre.")

            # Si un centre est envoyé → vérifie qu’il fait partie des siens
            default_centre = serializer.validated_data.get("default_centre") or centres[0]
            if default_centre and default_centre not in centres:
                raise PermissionDenied("❌ Ce centre n’est pas dans votre périmètre autorisé.")

        # 🔎 Cas 3 — Candidat / Stagiaire : centre = celui de sa formation
        else:
            default_centre = serializer.validated_data.get("default_centre") or getattr(user, "centre", None)
            if not default_centre:
                raise PermissionDenied("❌ Impossible de créer un partenaire sans centre associé.")

        # ✅ Vérification finale de sécurité
        if not default_centre:
            raise PermissionDenied("❌ Le centre associé est obligatoire pour créer un partenaire.")

        # ✅ Création de l’objet
        # full_clean() sur le modèle lève des DjangoValidationError (non 400 côté DRF si on ne les trappe pas ici)
        try:
            instance = serializer.save(created_by=user, default_centre=default_centre)
        except DjangoValidationError as exc:
            if getattr(exc, "message_dict", None):
                raise DRFValidationError(detail=exc.message_dict) from exc
            msgs = getattr(exc, "messages", None)
            if msgs is not None:
                raise DRFValidationError(detail=list(msgs)) from exc
            raise DRFValidationError(detail=str(exc)) from exc

        LogUtilisateur.log_action(
            instance=instance,
            action=LogUtilisateur.ACTION_CREATE,
            user=user,
            details="Création d'un partenaire",
        )

    def update(self, request, *args, **kwargs):
        """
        Met à jour un partenaire après contrôle explicite des droits
        d'édition et du centre par défaut associé, puis renvoie
        l'enveloppe API standard.
        """
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        user = request.user

        # 🔒 Candidat : créateur, dernier auteur, ou toute saisie (doublon / co-saisie)
        if not is_centre_scoped_staff(user) and not getattr(user, "is_superuser", False):
            if (
                instance.created_by_id != user.id
                and instance.updated_by_id != user.id
                and not instance.saisi_par.filter(id=user.id).exists()
            ):
                raise PermissionDenied("Vous ne pouvez modifier que vos propres partenaires.")

        if is_centre_scoped_staff(user) and not self._is_admin_like(user):
            if not self._user_can_access_partenaire(instance, user):
                raise PermissionDenied("Partenaire hors de votre périmètre (centres).")

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)

        # ⚙️ Logique ajoutée : remplir le centre manquant automatiquement
        default_centre = instance.default_centre  # valeur actuelle
        if not default_centre:
            if self._is_admin_like(user):
                # Admin : ne rien forcer
                default_centre = serializer.validated_data.get("default_centre")
            elif is_centre_scoped_staff(user):
                centres = list(user.centres.all())
                if not centres:
                    raise PermissionDenied("Vous n’êtes rattaché à aucun centre.")
                default_centre = serializer.validated_data.get("default_centre") or centres[0]
            else:
                default_centre = serializer.validated_data.get("default_centre") or getattr(user, "centre", None)
                if not default_centre:
                    raise PermissionDenied("Impossible de modifier un partenaire sans centre associé.")

        try:
            instance = serializer.save(default_centre=default_centre)
        except DjangoValidationError as exc:
            if getattr(exc, "message_dict", None):
                raise DRFValidationError(detail=exc.message_dict) from exc
            msgs = getattr(exc, "messages", None)
            if msgs is not None:
                raise DRFValidationError(detail=list(msgs)) from exc
            raise DRFValidationError(detail=str(exc)) from exc

        LogUtilisateur.log_action(
            instance=instance,
            action=LogUtilisateur.ACTION_UPDATE,
            user=request.user,
            details="Modification d'un partenaire",
        )

        return self.success_response(
            data=self.get_serializer(instance).data,
            message="Partenaire mis à jour avec succès.",
        )

    def destroy(self, request, *args, **kwargs):
        """
        - Staff / superuser : archivage global (``is_active=False``) selon le périmètre.
        - Créateur (non staff) : archivage global.
        - Autre compte ayant saisi / réouvert le doublon (``saisi_par``) : retrait
          **uniquement** de sa liste (``retrait_de_vue_par``) ; le partenaire reste
          actif et visible pour les autres.
        """
        instance = self.get_object()  # respecte le scope
        user = request.user

        if is_centre_scoped_staff(user) and not self._is_admin_like(user):
            if not self._user_can_access_partenaire(instance, user):
                raise PermissionDenied("Partenaire hors de votre périmètre (centres).")
            instance.is_active = False
            instance.save()
            LogUtilisateur.log_action(
                instance=instance,
                action=LogUtilisateur.ACTION_DELETE,
                user=user,
                details="Archivage logique d'un partenaire",
            )
            return Response(
                {"success": True, "message": "Partenaire archivé avec succès.", "data": None},
                status=status.HTTP_200_OK,
            )

        if getattr(user, "is_superuser", False) or self._is_admin_like(user):
            instance.is_active = False
            instance.save()
            LogUtilisateur.log_action(
                instance=instance,
                action=LogUtilisateur.ACTION_DELETE,
                user=user,
                details="Archivage logique d'un partenaire",
            )
            return Response(
                {"success": True, "message": "Partenaire archivé avec succès.", "data": None},
                status=status.HTTP_200_OK,
            )

        is_creator = instance.created_by_id == user.id
        if not is_creator and not _candidat_lie_pour_fiche_liste(user, instance):
            raise PermissionDenied("Vous ne pouvez pas retirer ce partenaire de votre liste.")

        if is_creator:
            instance.is_active = False
            instance.save()
            LogUtilisateur.log_action(
                instance=instance,
                action=LogUtilisateur.ACTION_DELETE,
                user=user,
                details="Archivage logique d'un partenaire",
            )
            return Response(
                {"success": True, "message": "Partenaire archivé avec succès.", "data": None},
                status=status.HTTP_200_OK,
            )

        # Co-saisie : retrait de liste perso seulement
        instance.retrait_de_vue_par.add(user)
        LogUtilisateur.log_action(
            instance=instance,
            action=LogUtilisateur.ACTION_DELETE,
            user=user,
            details="Retrait de la liste personnelle (fiche partagée toujours active)",
        )
        return Response(
            {
                "success": True,
                "message": "Cette fiche a été retirée de votre liste. Elle reste visible pour les autres comptes concernés.",
                "data": None,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=["post"], url_path="desarchiver")
    def desarchiver(self, request, *args, **kwargs):
        """
        Restaure un partenaire archivé en respectant le même périmètre d'accès
        que les autres opérations du viewset.
        """
        instance = self.get_archived_aware_object()
        user = request.user

        if not is_centre_scoped_staff(user) and not getattr(user, "is_superuser", False):
            is_creator = instance.created_by_id == user.id
            in_saisi = instance.saisi_par.filter(id=user.id).exists()
            if not is_creator and not in_saisi:
                raise PermissionDenied(
                    "Vous ne pouvez restaurer qu'une fiche que vous avez créée "
                    "ou sur laquelle vous êtes en co-saisie."
                )

        if is_centre_scoped_staff(user) and not self._is_admin_like(user):
            if not self._user_can_access_partenaire(instance, user):
                raise PermissionDenied("Partenaire hors de votre périmètre (centres).")

        if instance.is_active:
            return self.success_response(
                data=self.get_serializer(instance).data,
                message="Partenaire déjà actif.",
            )

        instance.is_active = True
        instance.save()
        LogUtilisateur.log_action(
            instance=instance,
            action=LogUtilisateur.ACTION_UPDATE,
            user=request.user,
            details="Désarchivage d'un partenaire",
        )
        return self.success_response(
            data=self.get_serializer(instance).data,
            message="Partenaire désarchivé avec succès.",
        )

    @extend_schema(
        summary="Réafficher une fiche dans ma liste (annuler retrait perso)",
        description="Réservé à un compte listé dans « saisi » (hors seul archivage global) : "
        "retire l’utilisateur de `retrait_de_vue_par` (la fiche reste active pour les autres).",
        tags=["Partenaires"],
    )
    @action(detail=True, methods=["post"], url_path="reafficher-dans-ma-liste")
    def reafficher_dans_ma_liste(self, request, pk=None):
        """
        Après un DELETE « retrait de liste seul », l’enregistrement n’apparaît plus
        au ``get_queryset`` classique : cet endpoint restaure l’affichage pour l’utilisateur.

        Candidat·e : même rattachement que la liste (prospection, saisi, etc.), pas
        seulement co-saisie. Staff / admin : périmètre via ``get_object``.
        """
        user = request.user
        if not (user and getattr(user, "is_authenticated", False)):
            raise PermissionDenied("Authentification requise.")
        p = self.get_object()
        if not p.is_active:
            raise PermissionDenied("Partenaire archivé globalement : voir le désarchivage staff ou créateur.")
        staffish = (
            is_centre_scoped_staff(user) and not is_staff_read(user)
        ) or self._is_admin_like(user) or getattr(user, "is_superuser", False)
        if not staffish and not _candidat_lie_pour_fiche_liste(user, p):
            raise PermissionDenied("Accès refusé.")
        if not p.retrait_de_vue_par.filter(id=user.id).exists():
            return self.success_response(
                data=self.get_serializer(p).data,
                message="Cette fiche est déjà dans votre liste.",
            )
        p.retrait_de_vue_par.remove(user)
        LogUtilisateur.log_action(
            instance=p,
            action=LogUtilisateur.ACTION_UPDATE,
            user=user,
            details="Réaffichage dans la liste personnelle (annulation retrait de vue)",
        )
        return self.success_response(
            data=self.get_serializer(p).data,
            message="Fiche de nouveau affichée dans votre liste.",
        )

    # -------------------- détail enrichi --------------------

    @extend_schema(
        summary="Détail d’un partenaire avec relations",
        description="Statistiques sur prospections, formations (via appairages/prospections), appairages et candidats.",
        tags=["Partenaires"],
        responses={200: PartenaireSerializer},
    )
    @action(detail=True, methods=["get"], url_path="with-relations")
    def retrieve_with_relations(self, request, pk=None):
        """
        Retourne la représentation sérialisée d'un partenaire complétée
        par des compteurs de prospections, appairages, formations et
        candidats associés.
        """
        partenaire = self.get_object()  # déjà scopé
        data = self.get_serializer(partenaire).data

        prospections_count = getattr(partenaire, "prospections_count", None)
        if prospections_count is None:
            prospections_count = partenaire.prospections.count()

        appairages_count = getattr(partenaire, "appairages_count", None)
        if appairages_count is None:
            appairages_count = partenaire.appairages.count()

        formations_count = getattr(partenaire, "formations_count", None)
        if formations_count is None:
            app_ids = set(partenaire.appairages.filter(formation__isnull=False).values_list("formation_id", flat=True))
            pros_ids = set(partenaire.prospections.filter(formation__isnull=False).values_list("formation_id", flat=True))
            formations_count = len(app_ids.union(pros_ids))

        candidats_count = getattr(partenaire, "candidats_count", None)
        if candidats_count is None:
            candidats_count = partenaire.appairages.values("candidat_id").distinct().count()

        data["prospections"] = {"count": prospections_count}
        data["appairages"] = {"count": appairages_count}
        data["formations"] = {"count": formations_count}
        data["candidats"] = {"count": candidats_count}
        return self.success_response(data=data, message="Détail enrichi du partenaire récupéré avec succès.")

    # -------------------- Export Excel --------------------

    @extend_schema(summary="Exporter les partenaires au format XLSX")
    @action(detail=False, methods=["get"], url_path="export-xlsx")
    def export_xlsx(self, request):
        """
        Exporte les partenaires visibles de l'utilisateur dans un
        classeur Excel formaté, en respectant les mêmes règles de
        périmètre et de permissions que la liste.
        """
        qs = self.filter_queryset(self.get_queryset().select_related("default_centre", "created_by"))

        # ==========================================================
        # 📘 Création du classeur
        # ==========================================================
        wb = Workbook()
        ws = wb.active
        ws.title = "Partenaires"

        # ==========================================================
        # 🖼️ Logo Rap_App (affiche en dev et prod)
        # ==========================================================
        try:
            logo_path = Path(settings.BASE_DIR) / "rap_app/static/images/logo.png"
            if logo_path.exists():
                img = XLImage(str(logo_path))
                img.height = 60
                img.width = 60
                ws.add_image(img, "A1")
        except Exception:
            pass

        # ==========================================================
        # 🧾 Titre + date d’export
        # ==========================================================
        ws.merge_cells("B1:AJ1")
        ws["B1"] = "Export des partenaires — Rap_App"
        ws["B1"].font = Font(bold=True, size=14, color="0077CC")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("B2:AJ2")
        ws["B2"] = f"Export réalisé le {dj_timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        ws["B2"].font = Font(italic=True, size=10, color="555555")
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

        ws.append([])  # ligne vide

        # ==========================================================
        # 📋 En-têtes
        # ==========================================================
        headers = [
            # Identité
            "ID",
            "Nom",
            "Type",
            "Secteur d’activité",
            # Adresse
            "Numéro de rue",
            "Adresse",
            "Complément",
            "Code postal",
            "Ville",
            "Pays",
            # Coordonnées
            "Téléphone général",
            "Email général",
            # Contact principal
            "Contact nom",
            "Contact poste",
            "Contact email",
            "Contact téléphone",
            # Employeur
            "SIRET",
            "Type employeur",
            "Employeur spécifique",
            "Code APE",
            "Effectif total",
            "IDCC",
            "Assurance chômage spéciale",
            # Maître 1
            "Maître 1 - Nom de naissance",
            "Maître 1 - Prénom",
            "Maître 1 - Date de naissance",
            "Maître 1 - Courriel",
            "Maître 1 - Emploi occupé",
            "Maître 1 - Diplôme/titre le plus élevé",
            "Maître 1 - Niveau diplôme/titre",
            # Maître 2
            "Maître 2 - Nom de naissance",
            "Maître 2 - Prénom",
            "Maître 2 - Date de naissance",
            "Maître 2 - Courriel",
            "Maître 2 - Emploi occupé",
            "Maître 2 - Diplôme/titre le plus élevé",
            "Maître 2 - Niveau diplôme/titre",
            # Web & Actions
            "Site web",
            "Réseau social",
            "Type d’action",
            "Description action",
            "Description générale",
            # Métadonnées
            "Slug",
            "Centre par défaut",
            "Créé par",
            "Date création",
            # Statistiques
            "Nb prospections",
            "Nb formations",
            "Nb appairages",
        ]
        ws.append(headers)

        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="E9F2FF")

        # ==========================================================
        # 🧮 Données
        # ==========================================================
        def _fmt(val):
            if val is None:
                return ""
            if isinstance(val, datetime.datetime):
                return val.strftime("%d/%m/%Y %H:%M")
            if isinstance(val, datetime.date):
                return val.strftime("%d/%m/%Y")
            return str(val)

        for p in qs:
            prospections_count = getattr(p, "prospections_count", p.nb_prospections)
            formations_count = getattr(p, "formations_count", p.nb_formations)
            appairages_count = getattr(p, "appairages_count", p.nb_appairages)
            ws.append(
                [
                    # Identité
                    p.id,
                    p.nom,
                    p.get_type_display(),
                    p.secteur_activite or "",
                    # Adresse
                    p.street_number or "",
                    p.street_name or "",
                    p.street_complement or "",
                    p.zip_code or "",
                    p.city or "",
                    p.country or "",
                    # Coordonnées générales
                    p.telephone or "",
                    p.email or "",
                    # Contact
                    p.contact_nom or "",
                    p.contact_poste or "",
                    p.contact_email or "",
                    p.contact_telephone or "",
                    # Employeur
                    p.siret or "",
                    p.get_type_employeur_code_display() if p.type_employeur_code else "",
                    p.get_employeur_specifique_code_display() if p.employeur_specifique_code else "",
                    p.code_ape or "",
                    p.effectif_total or "",
                    p.idcc or "",
                    "Oui" if p.assurance_chomage_speciale else "Non",
                    # Maître 1
                    p.maitre1_nom_naissance or "",
                    p.maitre1_prenom or "",
                    _fmt(p.maitre1_date_naissance),
                    p.maitre1_courriel or "",
                    p.maitre1_emploi_occupe or "",
                    p.maitre1_diplome_titre or "",
                    p.get_maitre1_niveau_diplome_code_display() if p.maitre1_niveau_diplome_code else "",
                    # Maître 2
                    p.maitre2_nom_naissance or "",
                    p.maitre2_prenom or "",
                    _fmt(p.maitre2_date_naissance),
                    p.maitre2_courriel or "",
                    p.maitre2_emploi_occupe or "",
                    p.maitre2_diplome_titre or "",
                    p.get_maitre2_niveau_diplome_code_display() if p.maitre2_niveau_diplome_code else "",
                    # Web & Actions
                    p.website or "",
                    p.social_network_url or "",
                    p.get_actions_display() if p.actions else "",
                    p.action_description or "",
                    p.description or "",
                    # Métadonnées
                    p.slug or "",
                    getattr(p.default_centre, "nom", ""),
                    getattr(p.created_by, "username", ""),
                    _fmt(p.created_at),
                    # Stats
                    prospections_count,
                    formations_count,
                    appairages_count,
                ]
            )

        # ==========================================================
        # 📏 Largeurs colonnes + wrap sur texte long
        # ==========================================================
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            if col_letter in ["AV", "AW", "AX", "AY"]:  # Descriptions longues
                ws.column_dimensions[col_letter].width = 80
                for cell in col:
                    cell.alignment = Alignment(wrapText=True, vertical="top")
            else:
                max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                ws.column_dimensions[col_letter].width = min(max_len + 2, 35)

        # ==========================================================
        # 📤 Sauvegarde et réponse HTTP
        # ==========================================================
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        binary_content = buffer.getvalue()

        filename = f'partenaires_{dj_timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            binary_content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = len(binary_content)
        return response
