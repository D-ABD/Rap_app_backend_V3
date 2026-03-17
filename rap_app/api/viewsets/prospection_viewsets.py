# views/prospection.py

import datetime
from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.contrib.auth import get_user_model
from django.db import transaction
from django.db.models import Count, Exists, OuterRef, Q, Subquery
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.templatetags.static import static
from django.utils import timezone as dj_timezone
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import (
    OpenApiExample,
    OpenApiParameter,
    OpenApiResponse,
    extend_schema,
    extend_schema_view,
)
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework import filters, serializers, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.response import Response

from ...api.paginations import RapAppPagination
from ...api.roles import (
    is_admin_like,
    is_candidate,
    is_staff_like,
    is_staff_or_staffread,
    role_of,
    staff_centre_ids,
)
from ...models.candidat import Candidat
from ...models.custom_user import CustomUser
from ...models.formations import Formation
from ...models.logs import LogUtilisateur
from ...models.partenaires import Partenaire
from ...models.prospection import Prospection, ProspectionChoices
from ...models.prospection_comments import ProspectionComment
from ...services.prospection_ownership_service import (
    ProspectionOwnershipService,
    defer_prospection_owner_sync,
)
from ...utils.filters import ProspectionFilterSet
from ..permissions import CanAccessProspectionComment, IsOwnerOrStaffOrAbove
from ..serializers.prospection_serializers import (
    ProspectionChoiceListSerializer,
    ProspectionDetailSerializer,
    ProspectionListSerializer,
    ProspectionSerializer,
)


# -----------------------------------------------------------------------------
# Helpers formation
# -----------------------------------------------------------------------------
def get_candidate_formation(user):
    """
    Retourne la formation associée à un utilisateur candidat ou
    assimilé, ou None si aucune formation n'est liée.
    """
    cand = getattr(user, "candidat_associe", None) or getattr(user, "candidat", None)
    return getattr(cand, "formation", None)


def get_owner_formation(owner):
    """
    Retourne la formation associée à l'utilisateur owner via son
    profil candidat éventuel, ou None.
    """
    if not owner:
        return None
    cand = getattr(owner, "candidat_associe", None) or getattr(owner, "candidat", None)
    return getattr(cand, "formation", None)


def annotate_last_visible_comment(queryset, user):
    """
    Annote un queryset de prospections avec le dernier commentaire
    visible pour l'utilisateur (texte, date, identifiant) et le nombre
    de commentaires visibles en tenant compte du rôle.
    """
    base = ProspectionComment.objects.filter(prospection=OuterRef("pk"))

    if is_staff_like(user):
        visible_sub = base
        comments_filter = Q()
    elif is_candidate(user):
        visible_sub = base.filter(Q(is_internal=False) | Q(created_by=user))
        comments_filter = Q(comments__is_internal=False) | Q(comments__created_by=user)
    else:
        visible_sub = base.filter(is_internal=False)
        comments_filter = Q(comments__is_internal=False)

    return queryset.annotate(
        last_comment=Subquery(visible_sub.order_by("-created_at").values("body")[:1]),
        last_comment_at=Subquery(visible_sub.order_by("-created_at").values("created_at")[:1]),
        last_comment_id=Subquery(visible_sub.order_by("-created_at").values("id")[:1]),
        comments_count=Count("comments", filter=comments_filter, distinct=True),
    )


# -----------------------------------------------------------------------------
# Serializers « créer depuis prospection »
# -----------------------------------------------------------------------------
class PartenaireCreateFromProspectionSerializer(serializers.ModelSerializer):
    """
    Serializer de création de partenaire à partir d'une prospection.
    """

    class Meta:
        model = Partenaire
        fields = [
            "nom",
            "type",
            "secteur_activite",
            "street_name",
            "zip_code",
            "city",
            "country",
            "contact_nom",
            "contact_poste",
            "contact_telephone",
            "contact_email",
            "website",
            "social_network_url",
            "actions",
            "action_description",
            "description",
        ]


class CandidatCreateFromProspectionSerializer(serializers.ModelSerializer):
    """
    Serializer de création d'un candidat minimal à partir d'une
    prospection, avec contrôle sur le champ formation et statuts
    par défaut.
    """

    formation = serializers.PrimaryKeyRelatedField(queryset=Formation.objects.all(), required=False, allow_null=True)

    class Meta:
        model = Candidat
        fields = ["nom", "prenom", "email", "telephone", "ville", "code_postal", "formation", "statut", "cv_statut"]

    def validate_formation(self, value):
        # Seul le staff/admin peut fixer la formation lors de la création
        request = self.context.get("request")
        user = getattr(request, "user", None)
        if value is not None and (not user or getattr(user, "role", None) not in ["admin", "superadmin", "staff"]):
            raise serializers.ValidationError("Seul le staff peut fixer la formation.")
        return value

    def validate(self, attrs):
        # Valeurs par défaut pour les statuts
        attrs.setdefault("statut", getattr(Candidat.StatutCandidat, "ACCOMPAGNEMENT", "accompagnement"))
        attrs.setdefault("cv_statut", getattr(Candidat.CVStatut, "EN_COURS", "en_cours"))
        return attrs


class CandidatReadMinimalSerializer(serializers.ModelSerializer):
    """
    Serializer en lecture seule pour exposer un candidat minimal dans
    les réponses.
    """

    class Meta:
        model = Candidat
        fields = [
            "id",
            "nom",
            "prenom",
            "email",
            "telephone",
            "ville",
            "code_postal",
            "formation",
            "statut",
            "cv_statut",
            "created_at",
            "updated_at",
        ]


@extend_schema_view(
    list=extend_schema(
        summary="📋 Liste des prospections",
        tags=["Prospections"],
        parameters=[
            OpenApiParameter("statut", str, description="Filtrer par statut (prospection)"),
            OpenApiParameter("formation", int, description="Filtrer par formation (id)"),
            OpenApiParameter("partenaire", int, description="Filtrer par partenaire (id)"),
            OpenApiParameter("owner", int, description="Filtrer par responsable (id)"),
            OpenApiParameter("search", str, description="Recherche texte (commentaire, partenaire, etc.)"),
            # ✅ Nouveaux filtres formation
            OpenApiParameter(
                "formation_type_offre", str, description="ID ou liste d’IDs type d’offre (ex: 1 ou 1,2,3)"
            ),
            OpenApiParameter("formation_statut", str, description="ID ou liste d’IDs statut de formation"),
            OpenApiParameter("centre", str, description="ID ou liste d’IDs centre de formation"),
            # (facultatif) autres filtres déjà supportés par DjangoFilterBackend via ProspectionFilterSet
            OpenApiParameter("moyen_contact", str, description="Moyen de contact (email, telephone, visite, reseaux)"),
            OpenApiParameter("type_prospection", str, description="Type de prospection"),
            OpenApiParameter("motif", str, description="Motif de prospection"),
            OpenApiParameter("objectif", str, description="Objectif de prospection"),
        ],
        responses={200: OpenApiResponse(response=ProspectionListSerializer)},
    ),
    retrieve=extend_schema(
        summary="🔍 Détail d’une prospection",
        tags=["Prospections"],
        responses={200: OpenApiResponse(response=ProspectionDetailSerializer)},
    ),
    create=extend_schema(summary="➕ Créer une prospection", tags=["Prospections"]),
    update=extend_schema(summary="✏️ Modifier une prospection", tags=["Prospections"]),
    destroy=extend_schema(summary="🗑️ Annuler une prospection", tags=["Prospections"]),
)
class ProspectionViewSet(viewsets.ModelViewSet):
    """
    ViewSet principal des prospections.

    Source de vérité actuelle :
    - accès protégé par `IsOwnerOrStaffOrAbove`
    - scoping hybride :
      - admins : accès global
      - staff : visibilité centre + certains liens owner/created_by
      - candidats : visibilité centrée sur leurs propres prospections
    - la résolution métier de `owner`, `formation` et `centre_id` lors des
      créations/mises à jour passe par `ProspectionOwnershipService`
    - le fichier reste dense et combine CRUD, filtres, actions métier,
      archivage, choix et export

    La documentation de ce viewset doit être lue avec cette réalité hybride,
    non comme un simple CRUD standard.
    """

    queryset = Prospection.objects.select_related(
        "partenaire",
        "formation",
        "formation__type_offre",
        "formation__statut",
        "formation__centre",
    )
    permission_classes = [IsOwnerOrStaffOrAbove]  # Permission personnalisée (cf. commentaire docstring ci-dessus).
    pagination_class = RapAppPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = ProspectionFilterSet
    ordering_fields = ["created_at", "date_prospection", "owner__username", "last_comment_at", "comments_count"]
    ordering = ["-date_prospection", "-created_at"]

    search_fields = [
        "commentaire",
        "statut",
        "objectif",
        "motif",
        "type_prospection",
        "owner__username",
        "created_by__username",
        "partenaire__nom",
        "partenaire__city",
        "partenaire__zip_code",
        "partenaire__secteur_activite",
        "partenaire__contact_nom",
        "partenaire__contact_email",
        "partenaire__contact_telephone",
        "formation__nom",
        "formation__num_offre",
    ]

    # -------------------------------------------------------------------------
    # helpers scope/permissions
    # -------------------------------------------------------------------------
    def _scoped_for_user(self, qs, user):
        """
        Restreint le queryset pour ne donner à voir que les objets accessibles à l'utilisateur
        en fonction de son rôle et de ses centres.
        - admin_like : accès complet.
        - staff/staffread : voir prospections relatives à leur(s) centre(s), où ils sont owner, ou créées par eux-mêmes,
                            ou où ils ont commenté.
        - candidat : uniquement prospections où il est owner.
        - autre utilisateur : prospections où il est owner ou qu'il a créées.
        - anonymes : rien.
        """
        if not user or not user.is_authenticated:
            return qs.none()

        if is_admin_like(user):
            return qs

        if is_staff_or_staffread(user):
            centre_ids = staff_centre_ids(user) or []
            conds = Q(formation__centre_id__in=centre_ids) if centre_ids else Q()
            return qs.filter(
                conds
                | Q(owner=user)
                | Q(created_by=user)
                | Exists(ProspectionComment.objects.filter(prospection=OuterRef("pk"), created_by=user))
            ).distinct()

        # ✅ candidat : prospections où il est owner (même si créées par staff/admin)
        if is_candidate(user):
            return qs.filter(owner=user)

        return qs.filter(Q(owner=user) | Q(created_by=user))

    def _ensure_staff_can_use_formation(self, user, formation: Formation | None):
        """
        Vérifie que la formation demandée est dans le périmètre de l'utilisateur staff.
        - Les admins ne sont pas restreints.
        - Les autres staff ne peuvent utiliser que les formations de leur(s) centre(s).
        """
        if not formation:
            return
        if is_admin_like(user):
            return
        if is_staff_or_staffread(user):
            allowed = set(user.centres.values_list("id", flat=True))
            if formation.centre_id not in allowed:
                raise PermissionDenied("Formation hors de votre périmètre (centres).")

    # -------------------------------------------------------------------------
    # Queryset & visibilité métier/filtre
    # -------------------------------------------------------------------------
    def get_queryset(self):
        """
        Retourne le queryset filtered approprié, annoté avec le dernier commentaire visible selon l'utilisateur.

        - Applique un filtrage spécifique sur le champ "activite" et gère les paramètres d’inclusion des archives,
          en plus du scoping selon le rôle (voir docstring de classe).
        - Annote chaque prospection sur :
            * dernier commentaire visible,
            * date, id, et nombre de commentaires visibles.
        - Applique filterset_class (si utilisé via DRF filtering) et search_fields/order/pagination via DRF Backends.
        """
        base = super().get_queryset()
        user = getattr(self.request, "user", None)
        qs = annotate_last_visible_comment(base, user)
        qs = self._scoped_for_user(qs, user)

        # 🆕 Gestion du filtre d'activité (active / archivée),
        # via "activite", ou via "inclure_archives" et "avec_archivees"
        activite_param = self.request.query_params.get("activite")
        inclure_archivees = any(
            self.request.query_params.get(k) in ("1", "true", "True", "yes", "oui")
            for k in ("inclure_archives", "avec_archivees")
        )

        if activite_param in ("active", "archivee"):
            qs = qs.filter(activite=activite_param)
        elif not inclure_archivees:
            qs = qs.filter(activite=Prospection.ACTIVITE_ACTIVE)

        # La ligne suivante n'a aucun effet (juste pour debug)
        # (f"[DEBUG get_queryset] action={getattr(self, 'action', '?')} activite={activite_param} ids={ids}")

        return qs

    # -------------------------------------------------------------------------
    # Actions personnalisées
    # -------------------------------------------------------------------------
    @action(detail=False, methods=["get"], url_path="filtres")
    def get_filters(self, request):
        """
        Point d'entrée API : GET /prospections/filtres

        Objectif métier :
            - Fournir les listes de valeurs disponibles et visibles pour l'utilisateur (scopées).
            - À utiliser pour peupler l'UI de filtres avancés.

        Permissions :
            - Accès restreint selon _scoped_for_user (voir docstring classe).

        Entrée : aucun paramètre requis.

        Réponse : format JSON, cf. code ci-dessous ("data": {...}).
        - Impossible de garantir l'exhaustivité du contrat ici si ProspectionChoices.* n'est pas visible.
        """

        def to_choice(queryset, label_attr="nom"):
            return [{"value": obj.id, "label": getattr(obj, label_attr)} for obj in queryset]

        qs = self._scoped_for_user(Prospection.objects.all(), request.user)

        formation_ids = qs.values_list("formation_id", flat=True).distinct()
        partenaire_ids = qs.values_list("partenaire_id", flat=True).distinct()
        owner_ids = qs.values_list("owner_id", flat=True).distinct()

        formations_qs = Formation.objects.filter(id__in=formation_ids).only("id", "nom", "num_offre", "centre_id")
        partenaires = Partenaire.objects.filter(id__in=partenaire_ids)
        owners = CustomUser.objects.filter(id__in=owner_ids)

        formations = [
            {
                "value": f.id,
                "label": f"{f.nom} — {f.num_offre}" if getattr(f, "num_offre", None) else f.nom,
            }
            for f in formations_qs
        ]

        type_offres = (
            formations_qs.values("type_offre_id", "type_offre__nom").exclude(type_offre_id__isnull=True).distinct()
        )
        statuts = formations_qs.values("statut_id", "statut__nom").exclude(statut_id__isnull=True).distinct()
        centres = formations_qs.values("centre_id", "centre__nom").exclude(centre_id__isnull=True).distinct()

        return Response(
            {
                "success": True,
                "message": "Filtres prospections récupérés avec succès.",
                "data": {
                    "formations": formations,
                    "partenaires": to_choice(partenaires),
                    "owners": to_choice(owners, label_attr="username"),
                    "statut": (
                        ProspectionChoices.get_statut_choices()
                        if hasattr(ProspectionChoices, "get_statut_choices")
                        else [
                            {"value": k, "label": str(v)}
                            for k, v in getattr(ProspectionChoices, "PROSPECTION_STATUS_CHOICES", [])
                        ]
                    ),
                    "objectif": (
                        ProspectionChoices.get_objectif_choices()
                        if hasattr(ProspectionChoices, "get_objectif_choices")
                        else [
                            {"value": k, "label": str(v)}
                            for k, v in getattr(ProspectionChoices, "PROSPECTION_OBJECTIF_CHOICES", [])
                        ]
                    ),
                    "motif": (
                        ProspectionChoices.get_motif_choices()
                        if hasattr(ProspectionChoices, "get_motif_choices")
                        else [
                            {"value": k, "label": str(v)}
                            for k, v in getattr(ProspectionChoices, "PROSPECTION_MOTIF_CHOICES", [])
                        ]
                    ),
                    "type_prospection": (
                        ProspectionChoices.get_type_choices()
                        if hasattr(ProspectionChoices, "get_type_choices")
                        else [
                            {"value": k, "label": str(v)}
                            for k, v in getattr(ProspectionChoices, "TYPE_PROSPECTION_CHOICES", [])
                        ]
                    ),
                    "moyen_contact": (
                        ProspectionChoices.get_moyen_contact_choices()
                        if hasattr(ProspectionChoices, "get_moyen_contact_choices")
                        else [
                            {"value": k, "label": str(v)}
                            for k, v in getattr(ProspectionChoices, "MOYEN_CONTACT_CHOICES", [])
                        ]
                    ),
                    "formation_type_offre": [
                        {"value": row["type_offre_id"], "label": row["type_offre__nom"]} for row in type_offres
                    ],
                    "formation_statut": [{"value": row["statut_id"], "label": row["statut__nom"]} for row in statuts],
                    "centres": [{"value": row["centre_id"], "label": row["centre__nom"]} for row in centres],
                    "user_role": getattr(request.user, "role", None),
                }
            }
        )

    @action(detail=True, methods=["post"], url_path="archiver")
    @extend_schema(
        summary="🗃 Archiver une prospection",
        description="Marque une prospection comme archivée (désactivée).",
        tags=["Prospections"],
        responses={200: OpenApiResponse(response=ProspectionSerializer)},
    )
    def archiver(self, request, pk=None):
        """
        Archive une prospection en mettant à jour son activité et en
        journalisant l'action.
        """
        instance = self.get_object()
        resultat = request.data.get("resultat")
        instance.archiver(user=request.user, resultat=resultat)
        LogUtilisateur.log_action(
            instance, LogUtilisateur.ACTION_UPDATE, request.user, f"Archivée ({resultat or 'sans résultat'})"
        )
        serializer = self.get_serializer(instance)
        return Response(
            {"success": True, "message": "Prospection archivée avec succès.", "data": serializer.data},
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=["post"], url_path="desarchiver")
    @extend_schema(
        summary="📂 Désarchiver une prospection",
        description="Rétablit une prospection archivée vers l’état actif.",
        tags=["Prospections"],
        responses={200: OpenApiResponse(response=ProspectionSerializer)},
    )
    def desarchiver(self, request, pk=None):
        """
        Désarchive une prospection archivée et journalise l'action.
        """
        instance = self.get_object()

        if instance.activite != Prospection.ACTIVITE_ARCHIVEE:
            return Response(
                {"success": False, "message": "La prospection n’est pas archivée."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        instance.desarchiver(user=request.user)
        LogUtilisateur.log_action(instance, LogUtilisateur.ACTION_UPDATE, request.user, "Désarchivée")

        serializer = self.get_serializer(instance)
        return Response(
            {"success": True, "message": "Prospection désarchivée avec succès.", "data": serializer.data},
            status=status.HTTP_200_OK,
        )

    def get_object(self):
        """
        Récupère une prospection sans filtrage sur l'activité puis
        applique les permissions sur l'objet.
        """
        queryset = Prospection.objects.all()  # ne filtre pas sur is_active
        obj = get_object_or_404(queryset, pk=self.kwargs["pk"])
        self.check_object_permissions(self.request, obj)
        return obj

    def get_serializer_class(self):
        """
        Retourne le serializer approprié selon l'action (liste, détail,
        création associée).
        """
        if self.action == "list":
            return ProspectionListSerializer
        elif self.action == "retrieve":
            return ProspectionDetailSerializer
        elif self.action == "creer_partenaire":
            return PartenaireCreateFromProspectionSerializer
        elif self.action == "creer_candidat":
            return CandidatCreateFromProspectionSerializer
        return ProspectionSerializer

    # ---------- CREATE / UPDATE ----------
    def perform_create(self, serializer):
        """
        Crée une prospection en appliquant les règles de rôle sur
        l'owner, la formation et le centre, puis journalise l'action.
        """
        user = self.request.user

        # 🧩 Cas 1 : Candidat ou stagiaire → owner forcé à lui-même
        if hasattr(user, "is_candidat_or_stagiaire") and user.is_candidat_or_stagiaire():
            resolved = ProspectionOwnershipService.resolve_and_sync_ownership(
                actor=user,
                validated_data=serializer.validated_data,
            )
            with defer_prospection_owner_sync():
                instance = serializer.save(
                    created_by=user,
                    owner=resolved["owner"],
                    formation=resolved["formation"],
                    centre_id=resolved["centre_id"],
                )
        # 🧩 Cas 2 : Staff / Admin → owner par défaut = user, mais modifiable
        else:
            resolved = ProspectionOwnershipService.resolve_and_sync_ownership(
                actor=user,
                validated_data=serializer.validated_data,
            )

            # 🔒 Contrôle périmètre staff (si non admin)
            self._ensure_staff_can_use_formation(user, resolved["formation"])

            # ✅ Enregistrement final
            with defer_prospection_owner_sync():
                instance = serializer.save(
                    created_by=user,
                    owner=resolved["owner"],
                    formation=resolved["formation"],
                    centre_id=resolved["centre_id"],
                )

        LogUtilisateur.log_action(
            instance, LogUtilisateur.ACTION_CREATE, user, f"Création d’une prospection (owner={instance.owner or '—'})"
        )

    def perform_update(self, serializer):
        """
        Met à jour une prospection en respectant les contraintes de
        rôle sur l'owner, la formation et le centre, puis journalise
        l'opération.
        """
        user = self.request.user
        instance = serializer.instance

        # 🧩 Cas 1 — candidat : restrictions fortes
        if hasattr(user, "is_candidat_or_stagiaire") and user.is_candidat_or_stagiaire():
            # Interdit de changer formation ou owner
            new_form = serializer.validated_data.get("formation")
            if new_form and new_form != instance.formation:
                raise PermissionDenied("Vous n’avez pas le droit de modifier la formation associée.")
            if "owner" in serializer.validated_data and serializer.validated_data["owner"] != instance.owner:
                raise PermissionDenied("Vous n’avez pas le droit de modifier le responsable (owner).")

            data_owner = instance.owner
            data_formation = instance.formation
            centre_id = instance.centre_id

        # 🧩 Cas 2 — staff/admin
        else:
            resolved = ProspectionOwnershipService.resolve_and_sync_ownership(
                actor=user,
                validated_data=serializer.validated_data,
                instance=instance,
            )
            data_owner = resolved["owner"]
            data_formation = resolved["formation"]

            # Vérification périmètre formation (staff non admin)
            self._ensure_staff_can_use_formation(user, data_formation)
            centre_id = resolved["centre_id"]

        # 💾 Sauvegarde finale
        with defer_prospection_owner_sync():
            instance = serializer.save(
                updated_by=user,
                owner=data_owner,
                formation=data_formation,
                centre_id=centre_id,
            )

        LogUtilisateur.log_action(
            instance,
            LogUtilisateur.ACTION_UPDATE,
            user,
            f"Mise à jour d’une prospection (owner={data_owner or '—'})",
        )

    # -------------------------------------------------------------------------
    # DRF actions standards
    # -------------------------------------------------------------------------
    def list(self, request, *args, **kwargs):
        """
        Retourne la liste paginée des prospections visibles avec une
        enveloppe JSON contenant success, message et data.
        """
        qs = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            paginated_response = self.get_paginated_response(serializer.data)
            paginated_response.data["success"] = True
            paginated_response.data["message"] = "Liste paginée des prospections."
            return paginated_response

        serializer = self.get_serializer(qs, many=True)
        return Response(
            {"success": True, "message": "Liste des prospections.", "data": serializer.data},
            status=status.HTTP_200_OK,
        )

    def retrieve(self, request, *args, **kwargs):
        """
        Retourne le détail d’une prospection annotée avec les
        informations de dernier commentaire visible.
        """
        user = request.user
        instance = self.get_object()

        # 🟢 Ré-annote le queryset pour inclure last_comment / last_comment_at / comments_count
        qs = annotate_last_visible_comment(
            Prospection.objects.filter(pk=instance.pk).select_related(
                "partenaire",
                "formation",
                "formation__centre",
                "formation__type_offre",
                "formation__statut",
            ),
            user,
        )

        instance = qs.first()

        serializer = self.get_serializer(instance)
        return Response(
            {
                "success": True,
                "message": "Détail de la prospection",
                "data": serializer.data,
            }
        )

    def create(self, request, *args, **kwargs):
        """
        Crée une prospection après validation et retourne une réponse
        JSON avec success, message et data.
        """
        serializer = self.get_serializer(data=request.data, context={"request": request})
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        output_serializer = self.get_serializer(serializer.instance, context={"request": request})
        return Response(
            {"success": True, "message": "Prospection créée avec succès.", "data": output_serializer.data},
            status=status.HTTP_201_CREATED,
        )

    def update(self, request, *args, **kwargs):
        """
        Met à jour une prospection (PUT ou PATCH) après validation et
        retourne une réponse JSON avec success, message et data.
        """
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial, context={"request": request})
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        output_serializer = self.get_serializer(serializer.instance, context={"request": request})
        return Response(
            {"success": True, "message": "Prospection mise à jour avec succès.", "data": output_serializer.data}
        )

    def destroy(self, request, *args, **kwargs):
        """
        Supprime définitivement une prospection après journalisation et
        retourne un message de confirmation.
        """
        instance = self.get_object()
        instance_id = instance.id
        LogUtilisateur.log_action(
            instance, LogUtilisateur.ACTION_DELETE, request.user, "Suppression définitive de la prospection"
        )
        instance.delete()
        return Response(
            {"success": True, "message": f"Prospection #{instance_id} supprimée définitivement."},
            status=status.HTTP_200_OK,
        )

    # -------------------------------------------------------------------------
    # Actions custom métier
    # -------------------------------------------------------------------------
    @action(detail=True, methods=["post"], url_path="changer-statut")
    @extend_schema(
        summary="🔄 Mettre à jour le statut (et autres champs) d’une prospection",
        description=(
            "Permet de modifier le statut ET tout autre champ éditable de la prospection en une seule requête. "
            "Supporte l’alias `prochain_contact` (équivalent à `relance_prevue`). "
            "Si `relance_prevue` est renseigné et que le statut n’est pas terminal, "
            "la cohérence statut ↔ relance sera appliquée lors de l’enregistrement."
        ),
        tags=["Prospections"],
        request=ProspectionSerializer,
        responses={200: OpenApiResponse(response=ProspectionSerializer)},
    )
    def changer_statut(self, request, pk=None):
        """
        Met à jour le statut et certains champs d'une prospection, peut
        ajouter un commentaire et un historique, puis retourne la
        prospection sérialisée.
        """
        instance = self.get_object()

        ancien_statut = instance.statut

        allowed_keys = {
            "partenaire",
            "formation",
            "owner",
            "date_prospection",
            "type_prospection",
            "motif",
            "statut",
            "objectif",
            "commentaire",
            "relance_prevue",
            "moyen_contact",
        }

        incoming = {k: v for k, v in request.data.items() if k in allowed_keys or k == "prochain_contact"}

        if "relance_prevue" not in incoming and "prochain_contact" in incoming:
            incoming["relance_prevue"] = incoming.pop("prochain_contact")

        moyen_contact_provided = "moyen_contact" in incoming
        moyen_contact_value = incoming.pop("moyen_contact", None) if moyen_contact_provided else None
        commentaire_payload = incoming.get("commentaire") or request.data.get("commentaire") or ""

        serializer = self.get_serializer(instance, data=incoming, partial=True, context={"request": request})
        serializer.is_valid(raise_exception=True)

        # ⚠️ si la formation change, re-vérifier le périmètre staff
        new_form = serializer.validated_data.get("formation", instance.formation)
        self._ensure_staff_can_use_formation(request.user, new_form)

        self.perform_update(serializer)

        # Gestion particulière du champ "moyen_contact" : possible via ce endpoint même si le serializer ne le gère pas
        if moyen_contact_provided and hasattr(instance, "moyen_contact"):
            instance.refresh_from_db()
            if instance.moyen_contact != moyen_contact_value:
                instance.moyen_contact = moyen_contact_value
                if hasattr(instance, "updated_by"):
                    instance.updated_by = request.user
                    instance.save(update_fields=["moyen_contact", "updated_by"])
                else:
                    instance.save(update_fields=["moyen_contact"])

        instance.refresh_from_db()
        nouveau_statut = instance.statut
        relance_prevue_finale = getattr(instance, "relance_prevue", None)

        if commentaire_payload:
            ProspectionComment.objects.create(
                prospection=instance,
                body=commentaire_payload,
                is_internal=False,
                created_by=request.user,
            )

        if hasattr(instance, "creer_historique"):
            instance.creer_historique(
                champ_modifie="statut",
                ancien_statut=ancien_statut,
                nouveau_statut=nouveau_statut,
                type_prospection=instance.type_prospection,
                commentaire=commentaire_payload,
                resultat="Mise à jour via changer-statut",
                moyen_contact=moyen_contact_value if moyen_contact_provided else None,
                user=request.user,
                prochain_contact=relance_prevue_finale,
            )

        LogUtilisateur.log_action(
            instance,
            LogUtilisateur.ACTION_UPDATE,
            request.user,
            f"Mise à jour via changer-statut : {ancien_statut} → {nouveau_statut}",
        )

        return Response(
            {
                "success": True,
                "message": "Prospection mise à jour avec succès.",
                "data": ProspectionSerializer(instance, context={"request": request}).data,
            }
        )

    @action(detail=False, methods=["get"], url_path="choices")
    @extend_schema(
        summary="📚 Choix disponibles (statut, objectif, motif, type_prospection, moyen_contact, responsables, partenaires)",
        tags=["Prospections"],
        responses={200: OpenApiResponse(response=ProspectionChoiceListSerializer)},
    )
    def get_choices(self, request):
        """
        Retourne les valeurs d’énumération (statut, objectif, motif,
        type de prospection, moyen de contact) ainsi que les owners et
        partenaires visibles pour l'utilisateur.
        """

        def fmt(choices):
            return [{"value": k, "label": str(l)} for k, l in choices]

        User = get_user_model()
        user_role = getattr(request.user, "role", None)

        # ✅ restreint aux owners visibles dans le périmètre de l'utilisateur
        scoped_qs = self._scoped_for_user(Prospection.objects.all(), request.user)
        owner_ids = scoped_qs.values_list("owner_id", flat=True).distinct()
        users = User.objects.filter(id__in=owner_ids)
        sorted_users = sorted(users, key=lambda u: (u.get_full_name() or u.username).lower())
        owners = [{"value": u.id, "label": u.get_full_name() or u.username} for u in sorted_users]

        partenaire_ids = scoped_qs.values_list("partenaire_id", flat=True).distinct()
        partenaires = [
            {"value": p.id, "label": p.nom}
            for p in Partenaire.objects.filter(
                id__in=partenaire_ids,
            ).order_by("nom")
        ]

        return Response(
            {
                "success": True,
                "message": "Choix disponibles pour les prospections",
                "data": {
                    "statut": fmt(getattr(ProspectionChoices, "PROSPECTION_STATUS_CHOICES", [])),
                    "objectif": fmt(getattr(ProspectionChoices, "PROSPECTION_OBJECTIF_CHOICES", [])),
                    "motif": fmt(getattr(ProspectionChoices, "PROSPECTION_MOTIF_CHOICES", [])),
                    "type_prospection": fmt(getattr(ProspectionChoices, "TYPE_PROSPECTION_CHOICES", [])),
                    "moyen_contact": fmt(getattr(ProspectionChoices, "MOYEN_CONTACT_CHOICES", [])),
                    "owners": owners,
                    "partenaires": partenaires,
                    "user_role": user_role,
                    # 🆕 Ajout ici :
                    "current_user": {
                        "id": request.user.id,
                        "username": getattr(request.user, "get_full_name", lambda: None)() or request.user.username,
                    },
                },
            }
        )

    @extend_schema(summary="Exporter les prospections au format XLSX")
    @action(detail=False, methods=["get", "post"], url_path="export-xlsx")
    def export_xlsx(self, request):
        """
        Exporte au format Excel les prospections visibles (ou un
        sous-ensemble identifié par des IDs) avec mise en forme, sans
        réponse JSON.
        """
        import mimetypes

        # Fix openpyxl / mimetypes pour certains environnements où .webp n'est pas enregistré
        mimetypes.add_type("image/webp", ".webp")
        user = request.user

        # ==========================================================
        # 🔍 Queryset filtré
        # ==========================================================
        qs = self.filter_queryset(
            self.get_queryset().select_related(
                "formation",
                "formation__centre",
                "formation__type_offre",
                "formation__statut",
                "partenaire",
                "owner",
                "created_by",
            )
        )

        if request.method == "POST":
            ids = request.data.get("ids", [])
            if isinstance(ids, str):
                ids = [int(x) for x in ids.split(",") if x.isdigit()]
            elif isinstance(ids, list):
                ids = [int(x) for x in ids if str(x).isdigit()]
            else:
                ids = []
            if ids:
                qs = qs.filter(id__in=ids)

        # ==========================================================
        # 📘 Workbook
        # ==========================================================
        wb = Workbook()
        ws = wb.active
        ws.title = "Prospections"

        # ==========================================================
        # 🖼️ Logo Rap_App
        # ==========================================================
        try:
            logo_path = Path(settings.BASE_DIR) / "rap_app/static/images/logo.png"
            if logo_path.exists():
                img = XLImage(str(logo_path))
                img.height = 60
                img.width = 60
                ws.add_image(img, "A1")
        except Exception:
            pass

        # ==========================================================
        # 🧾 Titre + date d’export
        # ==========================================================
        ws.merge_cells("B1:Y1")
        ws["B1"] = "Export des prospections — Rap_App"
        ws["B1"].font = Font(bold=True, size=14, color="004C99")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("B2:Y2")
        ws["B2"] = f"Export réalisé le {dj_timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        ws["B2"].font = Font(italic=True, size=10, color="666666")
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

        ws.append([])

        # ==========================================================
        # 📋 En-têtes
        # ==========================================================
        headers = [
            "ID",
            "Date prospection",
            "Statut",
            "Activité",
            "Type prospection",
            "Objectif",
            "Motif",
            "Dernier commentaire",
            "Date Relance prévue",
            "Partenaire",
            "Code postal",
            "Contact",
            "Email",
            "Téléphone",
            "Formation",
            "Centre",
            "Type offre",
            "Statut formation",
            "Date début",
            "Date fin",
            "Numéro offre",
            "Places dispo",
            "Taux saturation (%)",
            "Places totales",
            "Inscrits totaux",
        ]
        ws.append(headers)

        header_row = ws.max_row
        last_col_letter = get_column_letter(len(headers))

        header_fill = PatternFill("solid", fgColor="B7DEE8")
        header_font = Font(bold=True, color="000000")
        border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        for cell in ws[header_row]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            cell.fill = header_fill
            cell.border = border
        ws.row_dimensions[header_row].height = 28

        # ==========================================================
        # 🧮 Données
        # ==========================================================
        def _fmt(val):
            if val is None:
                return ""
            if isinstance(val, datetime.datetime):
                return val.strftime("%d/%m/%Y %H:%M")
            if isinstance(val, datetime.date):
                return val.strftime("%d/%m/%Y")
            if isinstance(val, float):
                return round(val, 2)
            return str(val)

        even_fill = PatternFill("solid", fgColor="EEF3FF")
        odd_fill = PatternFill("solid", fgColor="FAFBFD")
        archived_fill = PatternFill("solid", fgColor="E0E0E0")

        for i, p in enumerate(qs, start=1):
            f = getattr(p, "formation", None)
            part = getattr(p, "partenaire", None)

            taux_pct = getattr(f, "taux_saturation", 0) or 0
            if taux_pct <= 1:
                taux_pct *= 100.0

            last_comment = (getattr(p, "last_comment", "") or "").strip()

            row = [
                p.id,
                _fmt(getattr(p, "date_prospection", None)),
                getattr(p, "statut", ""),
                getattr(p, "get_activite_display", lambda: getattr(p, "activite", ""))(),
                getattr(p, "type_prospection", ""),
                getattr(p, "objectif", ""),
                getattr(p, "motif", ""),
                last_comment,
                _fmt(getattr(p, "relance_prevue", None)),
                getattr(part, "nom", ""),
                getattr(part, "zip_code", ""),
                getattr(part, "contact_nom", ""),
                getattr(part, "contact_email", ""),
                getattr(part, "contact_telephone", ""),
                getattr(f, "nom", "") if f else "",
                getattr(f.centre, "nom", "") if f and f.centre else "",
                getattr(f.type_offre, "nom", "") if f and f.type_offre else "",
                getattr(f.statut, "nom", "") if f and f.statut else "",
                _fmt(getattr(f, "start_date", None)) if f else "",
                _fmt(getattr(f, "end_date", None)) if f else "",
                getattr(f, "num_offre", "") if f else "",
                getattr(f, "places_disponibles", "") if f else "",
                taux_pct,
                getattr(f, "total_places", "") if f else "",
                getattr(f, "total_inscrits", "") if f else "",
            ]
            ws.append(row)

            # Alternance et archive
            fill = (
                archived_fill
                if getattr(p, "activite", "") == Prospection.ACTIVITE_ARCHIVEE
                else (even_fill if i % 2 == 0 else odd_fill)
            )

            for cell in ws[ws.max_row]:
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrapText=True)

            # === KPI couleur sur taux saturation ===
            taux_cell = ws.cell(row=ws.max_row, column=22)
            try:
                taux_value = float(taux_cell.value or 0)
            except ValueError:
                taux_value = 0.0

            if taux_value <= 30:
                taux_color, taux_fill = ("9C0006", "FFC7CE")  # Rouge clair
            elif taux_value <= 50:
                taux_color, taux_fill = ("9C6500", "FFEB9C")  # Orange
            elif taux_value < 100:
                taux_color, taux_fill = ("1F4E79", "BDD7EE")  # Bleu
            else:
                taux_color, taux_fill = ("006100", "C6EFCE")  # Vert

            taux_cell.fill = PatternFill("solid", fgColor=taux_fill)
            taux_cell.font = Font(color=taux_color, bold=True)
            taux_cell.number_format = "0%"

            # Force conversion pour affichage correct
            taux_cell.value = taux_value / 100.0
            ws.row_dimensions[ws.max_row].height = 25

        # ==========================================================
        # 📊 Filtres + mise en page
        # ==========================================================
        end_row = ws.max_row
        if end_row > header_row:
            ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{end_row}"

        ws.freeze_panes = f"A{header_row + 1}"

        # ==========================================================
        # 📏 Largeurs auto-ajustées
        # ==========================================================
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = max((len(str(c.value)) for c in col if c.value), default=0)
            ws.column_dimensions[letter].width = min(max_len + 3, 42)

        # Colonnes spécifiques (meilleur confort lecture)
        ws.column_dimensions["E"].width = 35  # Objectif
        ws.column_dimensions["H"].width = 55  # Dernier commentaire
        ws.column_dimensions["N"].width = 28  # Téléphone
        ws.column_dimensions["Q"].width = 25  # Type offre

        ws.oddFooter.center.text = f"© Rap_App — export du {dj_timezone.now().strftime('%d/%m/%Y %H:%M')}"

        # ==========================================================
        # 📤 Réponse HTTP
        # ==========================================================
        buffer = BytesIO()
        wb.save(buffer)
        binary = buffer.getvalue()
        buffer.close()

        filename = f'prospections_{dj_timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            binary,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = len(binary)
        return response
