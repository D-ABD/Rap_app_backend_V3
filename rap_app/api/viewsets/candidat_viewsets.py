import csv
import logging
from io import BytesIO
from pathlib import Path
from uuid import uuid4

from django.conf import settings
from django.db import transaction
from django.db.models import Count, IntegerField, Prefetch, Q, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone as dj_timezone
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import extend_schema
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework import filters
from rest_framework.decorators import action
from django.core.exceptions import ValidationError as DjangoValidationError
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework.renderers import JSONRenderer
from rest_framework.response import Response
from weasyprint import CSS, HTML

from ...models import atelier_tre
from ...models.appairage import Appairage
from ...models.candidat import (
    NIVEAU_CHOICES,
    Candidat,
    HistoriquePlacement,
    ResultatPlacementChoices,
)
from ...models.commentaires_appairage import CommentaireAppairage
from ...models.centres import Centre
from ...models.custom_user import CustomUser
from ...models.formations import Formation
from ...models.prospection import Prospection
from ...services.candidate_account_service import CandidateAccountService
from ...services.candidate_bulk_service import CandidateBulkService
from ...services.candidate_lifecycle_service import CandidateLifecycleService
from ...utils.filters import CandidatFilter
from ..paginations import RapAppPagination
from ..permissions import CanAccessCandidatObject, IsStaffOrAbove
from ..roles import is_admin_like, is_staff_or_staffread, staff_centre_ids
from ..exception_handler import MESSAGE_ERROR_CODE_MAP
from .scoped_viewset import ScopedModelViewSet
from ..serializers.candidat_serializers import (
    CandidatCreateUpdateSerializer,
    CandidatListSerializer,
    CandidatLiteSerializer,
    CandidatQueryParamsSerializer,
    CandidatSerializer,
)

logger = logging.getLogger("rap_app.candidats")

SENSITIVE_KEYS = {"password", "token", "secret", "api_key", "auth", "credential", "authorization"}


def _parse_candidate_ids(payload) -> list[int]:
    ids = payload.get("candidate_ids") or payload.get("candidats") or []
    if not isinstance(ids, list) or any(not isinstance(i, int) for i in ids):
        raise ValidationError({"candidate_ids": ["Ce champ doit ĂŞtre une liste d'identifiants entiers."]})
    return ids


def _sanitize_dict(d: dict) -> dict:
    out = {}
    for k, v in d.items():
        if any(s in k.lower() for s in SENSITIVE_KEYS):
            out[k] = "***"
        else:
            out[k] = v
    return out


def _extract_validation_payload(exc) -> tuple[str, dict]:
    if hasattr(exc, "message_dict"):
        errors = exc.message_dict
    elif isinstance(getattr(exc, "detail", None), dict):
        errors = exc.detail
    else:
        errors = {"non_field_errors": [str(exc)]}

    if isinstance(errors, dict):
        if errors.get("non_field_errors"):
            message = errors["non_field_errors"][0]
        else:
            first_value = next(iter(errors.values()), [str(exc)])
            if isinstance(first_value, list) and first_value:
                message = str(first_value[0])
            else:
                message = str(first_value)
    else:
        message = str(exc)
        errors = {"non_field_errors": [message]}

    return str(message), errors


def _resolve_error_code(message: str | None) -> str | None:
    if not message:
        return None
    return MESSAGE_ERROR_CODE_MAP.get(message)


def _build_candidat_meta(user=None) -> dict:
    """Construit les mĂŠtadonnĂŠes pour /candidats/meta en respectant le scope de lâutilisateur."""

    if is_admin_like(user):
        centres_qs = Centre.objects.order_by("nom").only("id", "nom")
        formations_qs = (
            Formation.objects.select_related("centre").only("id", "nom", "num_offre", "centre__nom").order_by("nom")
        )
    elif is_staff_or_staffread(user):
        centre_ids = staff_centre_ids(user) or []
        centres_qs = Centre.objects.filter(id__in=centre_ids).order_by("nom").only("id", "nom")
        formations_qs = (
            Formation.objects.select_related("centre")
            .filter(centre_id__in=centre_ids)
            .only("id", "nom", "num_offre", "centre__nom")
            .order_by("nom")
        )
    else:
        centres_qs = Centre.objects.none()
        formations_qs = Formation.objects.none()

    return {
        "statut_choices": [{"value": k, "label": v} for k, v in Candidat.StatutCandidat.choices],
        "parcours_phase_choices": [{"value": k, "label": v} for k, v in Candidat.ParcoursPhase.choices],
        "statut_metier_choices": [{"value": k, "label": v} for k, v in Candidat.StatutMetier.choices],
        "phase_contract": {
            "legacy_status_field": "statut",
            "recommended_phase_field": "parcours_phase",
            "derived_phase_field": "parcours_phase_calculee",
            "business_status_field": "statut_metier_calcule",
            "legacy_status_supported": True,
            "legacy_status_deprecated": True,
            "legacy_status_removal_stage": "post_front_migration",
            "legacy_status_write_locked": True,
        },
        "phase_filter_aliases": {
            "parcours_phase": ["parcours_phase", "parcoursPhase"],
            "legacy_status": ["statut"],
        },
        "phase_ordering_fields": ["parcours_phase", "statut", "date_inscription", "nom", "prenom"],
        "phase_read_only_fields": [
            "parcours_phase_calculee",
            "is_inscrit_valide",
            "is_en_formation_now",
            "is_stagiaire_role_aligned",
            "has_compte_utilisateur",
        ],
        "manual_status_flags": [
            "admissible",
            "inscrit_gespers",
            "en_accompagnement_tre",
            "en_appairage",
        ],
        "phase_transition_actions": [
            {"key": "validate_inscription", "url_name": "candidat-validate-inscription", "method": "POST"},
            {"key": "set_admissible", "url_name": "candidat-set-admissible", "method": "POST"},
            {"key": "clear_admissible", "url_name": "candidat-clear-admissible", "method": "POST"},
            {"key": "set_gespers", "url_name": "candidat-set-gespers", "method": "POST"},
            {"key": "clear_gespers", "url_name": "candidat-clear-gespers", "method": "POST"},
            {"key": "set_accompagnement_tre", "url_name": "candidat-set-accompagnement", "method": "POST"},
            {"key": "clear_accompagnement_tre", "url_name": "candidat-clear-accompagnement", "method": "POST"},
            {"key": "set_appairage", "url_name": "candidat-set-appairage", "method": "POST"},
            {"key": "clear_appairage", "url_name": "candidat-clear-appairage", "method": "POST"},
            {"key": "start_formation", "url_name": "candidat-start-formation", "method": "POST"},
            {"key": "cancel_start_formation", "url_name": "candidat-cancel-start-formation", "method": "POST"},
            {"key": "complete_formation", "url_name": "candidat-complete-formation", "method": "POST"},
            {"key": "abandon", "url_name": "candidat-abandon", "method": "POST"},
        ],
        "rgpd_legal_basis_choices": [{"value": k, "label": v} for k, v in Candidat.RgpdLegalBasis.choices],
        "rgpd_notice_status_choices": [{"value": k, "label": v} for k, v in Candidat.RgpdNoticeStatus.choices],
        "rgpd_creation_source_choices": [{"value": k, "label": v} for k, v in Candidat.RgpdCreationSource.choices],
        "rgpd_consent_fields": [
            "rgpd_consent_obtained",
            "rgpd_consent_obtained_at",
            "rgpd_consent_recorded_by",
        ],
        "cv_statut_choices": [{"value": k, "label": v} for k, v in Candidat.CVStatut.choices],
        "type_contrat_choices": [{"value": k, "label": v} for k, v in Candidat.TypeContrat.choices],
        "disponibilite_choices": [{"value": k, "label": v} for k, v in Candidat.Disponibilite.choices],
        "resultat_placement_choices": [{"value": k, "label": v} for k, v in ResultatPlacementChoices.choices],
        "contrat_signe_choices": [{"value": k, "label": v} for k, v in Candidat.ContratSigne.choices],
        "niveau_choices": [{"value": val, "label": f"{val} â"} for val, _ in NIVEAU_CHOICES],
        "centre_choices": [{"value": c.id, "label": c.nom} for c in centres_qs],
        "formation_choices": [
            {
                "value": f.id,
                "label": f"{f.nom}" + (f" â {f.num_offre}" if f.num_offre else ""),
            }
            for f in formations_qs
        ],
    }


class CandidatViewSet(ScopedModelViewSet):
    """
    ViewSet principal des candidats.

    Source de vĂŠritĂŠ actuelle :
    - accĂ¨s protĂŠgĂŠ par `IsStaffOrAbove` et `CanAccessCandidatObject`
    - scoping centralisĂŠ via `ScopedModelViewSet` avec `scope_mode = "centre"`
      sur `formation__centre_id`
    - sĂŠlection de serializer par action :
      - `list` => `CandidatListSerializer` ou `CandidatLiteSerializer`
      - `create` / `update` / `partial_update` => `CandidatCreateUpdateSerializer`
      - autres actions => `CandidatSerializer`
    - actions mĂŠtier majeures :
      - `meta`
      - `creer-compte`
      - `validate-inscription`
      - `set-admissible` / `clear-admissible`
      - `set-gespers` / `clear-gespers`
      - `set-accompagnement` / `clear-accompagnement`
      - `set-appairage` / `clear-appairage`
      - `start-formation`
      - `cancel-start-formation`
      - `complete-formation`
      - `abandon`
      - `valider-demande-compte`
      - `refuser-demande-compte`
      - `export-xlsx`

    Les actions liĂŠes au compte candidat s'appuient dĂŠsormais sur
    `CandidateAccountService` plutĂ´t que sur les anciens signaux implicites.
    En particulier, `creer-compte` garantit un compte `candidatuser` sans
    promotion immĂŠdiate en `stagiaire` : cette promotion reste rĂŠservĂŠe Ă 
    l'entrĂŠe effective en formation.
    """

    permission_classes = [IsStaffOrAbove, CanAccessCandidatObject]
    pagination_class = RapAppPagination
    scope_mode = "centre"
    centre_lookup_paths = ("formation__centre_id",)

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = CandidatFilter

    search_fields = [
        "nom",
        "prenom",
        "email",
        "telephone",
        "ville",
        "code_postal",
        "origine_sourcing",
        "numero_osia",
        "formation__nom",
        "formation__num_offre",
        "formation__centre__nom",
        "compte_utilisateur__email",
        "placement_appairage__partenaire__nom",
        "cv_statut",
    ]

    ordering_fields = [
        "date_inscription",
        "nom",
        "prenom",
        "statut",
        "parcours_phase",
        "cv_statut",
        "formation",
        "formation__nom",
        "formation__centre__nom",
        "nb_appairages_calc",
        "nb_prospections_calc",
        "date_placement",
        "resultat_placement",
        "contrat_signe",
        "entreprise_placement",
        "entreprise_validee",
        "vu_par",
        "placement_appairage__date_appairage",
        "placement_appairage__partenaire__nom",
    ]
    ordering = ["-date_inscription"]

    def _qp_dict(self, request):
        qp = {}
        for k in request.query_params.keys():
            vals = request.query_params.getlist(k)
            qp[k] = vals if len(vals) > 1 else (vals[0] if vals else None)
        return _sanitize_dict(qp)

    def _log_filters(self, request, base_qs, filtered_qs):
        try:
            logger.debug("query_params=%s", self._qp_dict(request))
            before_count = base_qs.count()
            after_count = filtered_qs.count()
            logger.debug("queryset counts: before=%s after=%s", before_count, after_count)
            try:
                logger.debug("SQL: %s", str(filtered_qs.query))
            except Exception:
                logger.debug("SQL: <unavailable>")

            backend = DjangoFilterBackend()
            fs = backend.get_filterset(request, base_qs, self)
            if fs is not None:
                valid = fs.is_valid()
                form = getattr(fs, "form", None)
                errors = getattr(form, "errors", {})
                cleaned = getattr(form, "cleaned_data", {})
                logger.debug("FilterSet valid=%s errors=%s cleaned=%s", valid, errors, cleaned)
        except Exception:
            logger.exception("Erreur pendant le logging des filtres.")

    def _assert_staff_can_use_formation(self, formation):
        """LĂ¨ve PermissionDenied si staff et formation.centre_id hors de user.centres."""
        if not formation:
            return
        user = self.request.user
        if is_admin_like(user):
            return
        if is_staff_or_staffread(user):
            allowed = set(user.centres.values_list("id", flat=True))
            if getattr(formation, "centre_id", None) not in allowed:
                raise PermissionDenied("Formation hors de votre pĂŠrimĂ¨tre (centre).")

    def get_base_queryset(self):
        """
        Construit le queryset mĂŠtier de base avant application du scope centre.

        Le queryset est enrichi avec :
        - `select_related` sur les relations principales
        - `prefetch_related` pour appairages et ateliers
        - annotations de comptage (`nb_appairages_calc`, `nb_prospections_calc`)
        - flags ateliers via `AtelierTRE.annotate_candidats_with_atelier_flags`
        """
        qs = Candidat.objects.select_related(
            "formation",
            "formation__centre",
            "formation__type_offre",
            "evenement",
            "compte_utilisateur",
            "responsable_placement",
            "vu_par",
            "entreprise_placement",
            "entreprise_validee",
            "placement_appairage",
            "placement_appairage__partenaire",
            "placement_appairage__created_by",
            "placement_appairage__updated_by",
        ).prefetch_related(
            Prefetch(
                "appairages",
                queryset=Appairage.objects.select_related(
                    "partenaire",
                    "created_by",
                    "updated_by",
                )
                .only(
                    "id",
                    "candidat_id",
                    "partenaire_id",
                    "date_appairage",
                    "statut",
                    "retour_partenaire",
                    "date_retour",
                    "created_at",
                    "updated_at",
                    "created_by_id",
                    "updated_by_id",
                    "partenaire__id",
                    "partenaire__nom",
                    "created_by__id",
                    "created_by__first_name",
                    "created_by__last_name",
                    "created_by__email",
                    "created_by__username",
                )
                .prefetch_related(
                    Prefetch(
                        "commentaires",
                        queryset=CommentaireAppairage.objects.select_related("created_by")
                        .only(
                            "id",
                            "appairage_id",
                            "body",
                            "created_at",
                            "updated_at",
                            "created_by_id",
                            "updated_by_id",
                            "is_active",
                            "statut_snapshot",
                            "created_by__id",
                            "created_by__first_name",
                            "created_by__last_name",
                            "created_by__email",
                            "created_by__username",
                        )
                        .order_by("-created_at", "-pk"),
                    ),
                ),
            ),
            Prefetch(
                "ateliers_tre",
                queryset=atelier_tre.AtelierTRE.objects.only("id", "type_atelier"),
            ),
        )

        qs = qs.annotate(nb_appairages_calc=Count("appairages", distinct=True))

        qs = qs.annotate(
            nb_prospections_calc=Coalesce(
                Count("compte_utilisateur__prospections_attribuees", distinct=True),
                Value(0),
                output_field=IntegerField(),
            )
        )

        qs = atelier_tre.AtelierTRE.annotate_candidats_with_atelier_flags(qs)
        return qs

    def list(self, request, *args, **kwargs):
        """
        Liste paginĂŠe des candidats visibles aprĂ¨s validation souple des
        query params et journalisation technique des filtres.
        """
        qp_ser = CandidatQueryParamsSerializer(data=request.query_params)
        qp_ser.is_valid(raise_exception=False)
        logger.debug(
            "qp valid=%s errors=%s cleaned=%s",
            qp_ser.is_valid(),
            qp_ser.errors,
            qp_ser.validated_data,
        )

        base_qs = self.get_queryset()
        filtered_qs = self.filter_queryset(base_qs)
        self._log_filters(request, base_qs, filtered_qs)

        return super().list(request, *args, **kwargs)

    def get_serializer_class(self):
        """
        SĂŠlectionne dynamiquement le serializer en fonction de lâaction et des paramĂ¨tresÂ :
        - list (GET) : CandidatListSerializer, ou CandidatLiteSerializer si lite=1
        - create/update/partial_update : CandidatCreateUpdateSerializer
        - retrieve/autre : CandidatSerializer

        Permet dâadapter le format de rĂŠponse Ă  lâusage interface/front.
        """
        if self.action == "list":
            if self.request.query_params.get("lite") == "1":
                return CandidatLiteSerializer
            return CandidatListSerializer
        elif self.action in ["create", "update", "partial_update"]:
            return CandidatCreateUpdateSerializer
        return CandidatSerializer

    def get_serializer_context(self):
        """
        Ajoute le contexte DRF standard + lâobjet request.
        Important pour lâaccĂ¨s user/request dans les serializers si besoin.
        """
        ctx = super().get_serializer_context()
        ctx["request"] = getattr(self, "request", None)
        return ctx

    def perform_create(self, serializer):
        """ValidationError si formation absente ; _assert_staff_can_use_formation ; transaction.atomic() ; sauvegarde unique."""
        data = serializer.validated_data

        formation = data.get("formation")
        if not formation:
            raise ValidationError({"formation": ["La formation est obligatoire."]})

        self._assert_staff_can_use_formation(formation)

        with transaction.atomic():
            serializer.save(created_by=self.request.user, updated_by=self.request.user)

    def perform_update(self, serializer):
        """VĂŠrifie le pĂŠrimĂ¨tre centre avant mise Ă  jour et cascade les prospections le cas ĂŠchĂŠant."""
        instance = self.get_object()
        old_formation = getattr(instance, "formation", None)
        new_formation = serializer.validated_data.get("formation", old_formation)

        self._assert_staff_can_use_formation(new_formation)

        with transaction.atomic():
            updated = serializer.save(updated_by=self.request.user)

        if getattr(old_formation, "id", None) != getattr(new_formation, "id", None):
            self._cascade_update_prospections_on_formation_change(
                candidat=updated, old_form=old_formation, new_form=new_formation
            )

    @action(detail=True, methods=["post"], url_path="creer-compte")
    def creer_compte(self, request, pk=None):
        """POST : crĂŠe ou lie un compte candidat sans promotion immĂŠdiate en stagiaire."""
        candidat = self.get_object()
        try:
            user = CandidateAccountService.ensure_candidate_user(candidat, actor=request.user)
        except (ValidationError, DjangoValidationError) as e:
            message, errors = _extract_validation_payload(e)
            return self.error_response(
                message=message,
                errors=errors,
                error_code=_resolve_error_code(message),
                status_code=400,
            )

        return Response(
            {
                "success": True,
                "message": "Compte candidat crĂŠĂŠ ou liĂŠ avec succĂ¨s.",
                "user_id": user.id,
                "user_role": user.role,
            }
        )

    @action(detail=True, methods=["post"], url_path="validate-inscription")
    def validate_inscription(self, request, pk=None):
        """POST : valide l'entrĂŠe dans le parcours de recrutement sans forcer l'ĂŠtat GESPERS."""
        candidat = self.get_object()
        try:
            candidat = CandidateLifecycleService.validate_inscription(candidat, actor=request.user)
        except (ValidationError, DjangoValidationError) as e:
            message, errors = _extract_validation_payload(e)
            return self.error_response(
                message=message,
                errors=errors,
                error_code=_resolve_error_code(message),
                status_code=400,
            )

        return Response(
            {
                "success": True,
                "message": "EntrĂŠe dans le parcours de recrutement validĂŠe.",
                "data": {
                    "candidat_id": candidat.id,
                    "parcours_phase": candidat.parcours_phase,
                    "date_validation_inscription": candidat.date_validation_inscription,
                },
            }
        )

    @action(detail=True, methods=["post"], url_path="set-admissible")
    def set_admissible(self, request, pk=None):
        """POST : active manuellement l'ĂŠtat admissible."""
        candidat = self.get_object()
        candidat = CandidateLifecycleService.set_admissible(candidat, actor=request.user)
        return self.success_response(
            data={"candidat_id": candidat.id, "admissible": candidat.admissible},
            message="Statut 'Candidat admissible' enregistrĂŠ.",
        )

    @action(detail=True, methods=["post"], url_path="clear-admissible")
    def clear_admissible(self, request, pk=None):
        """POST : retire manuellement l'ĂŠtat admissible."""
        candidat = self.get_object()
        candidat = CandidateLifecycleService.clear_admissible(candidat, actor=request.user)
        return self.success_response(
            data={"candidat_id": candidat.id, "admissible": candidat.admissible},
            message="Statut 'Candidat admissible' retirĂŠ.",
        )

    @action(detail=True, methods=["post"], url_path="set-gespers")
    def set_gespers(self, request, pk=None):
        """POST : marque manuellement le candidat comme inscrit GESPERS."""
        candidat = self.get_object()
        candidat = CandidateLifecycleService.mark_gespers(candidat, actor=request.user)
        return self.success_response(
            data={"candidat_id": candidat.id, "inscrit_gespers": candidat.inscrit_gespers},
            message="Inscription GESPERS enregistrĂŠe.",
        )

    @action(detail=True, methods=["post"], url_path="clear-gespers")
    def clear_gespers(self, request, pk=None):
        """POST : annule manuellement l'inscription GESPERS."""
        candidat = self.get_object()
        candidat = CandidateLifecycleService.clear_gespers(candidat, actor=request.user)
        return self.success_response(
            data={"candidat_id": candidat.id, "inscrit_gespers": candidat.inscrit_gespers},
            message="Inscription GESPERS annulĂŠe.",
        )

    @action(detail=True, methods=["post"], url_path="set-accompagnement")
    def set_accompagnement(self, request, pk=None):
        """POST : positionne manuellement le candidat en accompagnement TRE."""
        candidat = self.get_object()
        candidat = CandidateLifecycleService.set_accompagnement(candidat, actor=request.user)
        return self.success_response(
            data={
                "candidat_id": candidat.id,
                "en_accompagnement_tre": candidat.en_accompagnement_tre,
                "statut": candidat.statut,
            },
            message="Statut 'En accompagnement TRE' enregistrĂŠ.",
        )

    @action(detail=True, methods=["post"], url_path="clear-accompagnement")
    def clear_accompagnement(self, request, pk=None):
        """POST : retire le statut manuel d'accompagnement TRE."""
        candidat = self.get_object()
        candidat = CandidateLifecycleService.clear_accompagnement(candidat, actor=request.user)
        return self.success_response(
            data={
                "candidat_id": candidat.id,
                "en_accompagnement_tre": candidat.en_accompagnement_tre,
                "statut": candidat.statut,
            },
            message="Statut 'En accompagnement TRE' retirĂŠ.",
        )

    @action(detail=True, methods=["post"], url_path="set-appairage")
    def set_appairage(self, request, pk=None):
        """POST : positionne manuellement le candidat en appairage."""
        candidat = self.get_object()
        candidat = CandidateLifecycleService.set_appairage(candidat, actor=request.user)
        return self.success_response(
            data={"candidat_id": candidat.id, "en_appairage": candidat.en_appairage, "statut": candidat.statut},
            message="Statut 'En appairage' enregistrĂŠ.",
        )

    @action(detail=True, methods=["post"], url_path="clear-appairage")
    def clear_appairage(self, request, pk=None):
        """POST : retire le statut manuel d'appairage."""
        candidat = self.get_object()
        candidat = CandidateLifecycleService.clear_appairage(candidat, actor=request.user)
        return self.success_response(
            data={"candidat_id": candidat.id, "en_appairage": candidat.en_appairage, "statut": candidat.statut},
            message="Statut 'En appairage' retirĂŠ.",
        )

    @action(detail=True, methods=["post"], url_path="start-formation")
    def start_formation(self, request, pk=None):
        """POST : positionne la phase mĂŠtier sur `stagiaire_en_formation` et aligne le rĂ´le si possible."""
        candidat = self.get_object()
        try:
            candidat = CandidateLifecycleService.start_formation(candidat, actor=request.user)
        except (ValidationError, DjangoValidationError) as e:
            message, errors = _extract_validation_payload(e)
            return self.error_response(
                message=message,
                errors=errors,
                error_code=_resolve_error_code(message),
                status_code=400,
            )

        compte = getattr(candidat, "compte_utilisateur", None)
        return Response(
            {
                "success": True,
                "message": "EntrĂŠe en formation enregistrĂŠe.",
                "data": {
                    "candidat_id": candidat.id,
                    "parcours_phase": candidat.parcours_phase,
                    "date_entree_formation_effective": candidat.date_entree_formation_effective,
                    "user_role": getattr(compte, "role", None),
                },
            }
        )

    @action(detail=True, methods=["post"], url_path="cancel-start-formation")
    def cancel_start_formation(self, request, pk=None):
        """POST : annule une entrĂŠe en formation enregistrĂŠe par erreur."""
        candidat = self.get_object()
        try:
            candidat = CandidateLifecycleService.cancel_start_formation(candidat, actor=request.user)
        except (ValidationError, DjangoValidationError) as e:
            message, errors = _extract_validation_payload(e)
            return self.error_response(
                message=message,
                errors=errors,
                error_code=_resolve_error_code(message),
                status_code=400,
            )

        return Response(
            {
                "success": True,
                "message": "EntrĂŠe en formation annulĂŠe.",
                "data": {
                    "candidat_id": candidat.id,
                    "parcours_phase": candidat.parcours_phase,
                    "date_entree_formation_effective": candidat.date_entree_formation_effective,
                    "date_sortie_formation": candidat.date_sortie_formation,
                },
            }
        )

    @action(detail=True, methods=["post"], url_path="complete-formation")
    def complete_formation(self, request, pk=None):
        """POST : positionne la phase mĂŠtier sur `sorti` sans casser le contrat legacy."""
        candidat = self.get_object()
        try:
            candidat = CandidateLifecycleService.complete_formation(candidat, actor=request.user)
        except (ValidationError, DjangoValidationError) as e:
            message, errors = _extract_validation_payload(e)
            return self.error_response(
                message=message,
                errors=errors,
                error_code=_resolve_error_code(message),
                status_code=400,
            )

        return Response(
            {
                "success": True,
                "message": "Sortie de formation enregistrĂŠe.",
                "data": {
                    "candidat_id": candidat.id,
                    "parcours_phase": candidat.parcours_phase,
                    "date_sortie_formation": candidat.date_sortie_formation,
                },
            }
        )

    @action(detail=True, methods=["post"], url_path="abandon")
    def abandon(self, request, pk=None):
        """POST : enregistre un abandon sur la nouvelle phase et le statut legacy."""
        candidat = self.get_object()
        try:
            candidat = CandidateLifecycleService.abandon(candidat, actor=request.user)
        except (ValidationError, DjangoValidationError) as e:
            message, errors = _extract_validation_payload(e)
            return self.error_response(
                message=message,
                errors=errors,
                error_code=_resolve_error_code(message),
                status_code=400,
            )

        return Response(
            {
                "success": True,
                "message": "Abandon enregistrĂŠ.",
                "data": {
                    "candidat_id": candidat.id,
                    "parcours_phase": candidat.parcours_phase,
                    "statut": candidat.statut,
                    "date_sortie_formation": candidat.date_sortie_formation,
                },
            }
        )

    @action(detail=True, methods=["post"], url_path="valider-demande-compte")
    def valider_demande_compte(self, request, pk=None):
        """POST : demande_compte en_attente et pas de compte liĂŠ ; creer_ou_lier_compte_utilisateur puis statut acceptee."""
        candidat = self.get_object()

        try:
            user = CandidateAccountService.approve_account_request(candidat, actor=request.user)
        except (ValidationError, DjangoValidationError) as e:
            message, errors = _extract_validation_payload(e)
            return self.error_response(
                message=message,
                errors=errors,
                error_code=_resolve_error_code(message),
                status_code=400,
            )

        return Response(
            {
                "success": True,
                "message": "Demande de compte validĂŠe et compte utilisateur crĂŠĂŠ ou liĂŠ.",
                "user_id": user.id,
                "user_email": user.email,
            }
        )

    @action(detail=True, methods=["post"], url_path="refuser-demande-compte")
    def refuser_demande_compte(self, request, pk=None):
        """POST : demande_compte en_attente ; passe statut Ă  refusee."""
        candidat = self.get_object()

        try:
            CandidateAccountService.reject_account_request(candidat, actor=request.user)
        except (ValidationError, DjangoValidationError) as e:
            message, errors = _extract_validation_payload(e)
            return self.error_response(
                message=message,
                errors=errors,
                error_code=_resolve_error_code(message),
                status_code=400,
            )

        return Response(
            {
                "success": True,
                "message": "Demande de compte refusĂŠe.",
            }
        )

    def _cascade_update_prospections_on_formation_change(self, candidat, old_form, new_form):
        """Met Ă  jour les Prospections (owner=candidat.compte_utilisateur) lorsqu'une formation change.

        Seules les prospections qui n'ont pas de formation ou qui ont l'ancienne formation sont mises Ă  jour.
        """
        if not candidat or not getattr(candidat, "compte_utilisateur_id", None):
            return

        qs = Prospection.objects.filter(owner_id=candidat.compte_utilisateur_id)

        if old_form is not None:
            qs = qs.filter(Q(formation__isnull=True) | Q(formation=old_form))
        else:
            qs = qs.filter(formation__isnull=True)

        if new_form is not None:
            centre_id = getattr(new_form, "centre_id", None)
            qs.update(formation=new_form, centre_id=centre_id)
        else:
            qs.update(formation=None)

    @extend_schema(responses=None)
    @action(
        detail=False,
        methods=["get"],
        url_path="meta",
        url_name="meta",
        renderer_classes=[JSONRenderer],
        permission_classes=[IsStaffOrAbove],
    )
    def meta(self, request):
        """
        Endpoint API GET /candidats/meta/

        - But mĂŠtier : Retourner toutes les valeurs de rĂŠfĂŠrence nĂŠcessaires pour afficher des menus filtres sur le front (statuts, choix enum, centres/formations accessibles, etc).
        - Permissions : IsAuthenticated (utilisateur connectĂŠ requis), mais filtrage mĂŠtier dans le code pour ne retourner que le pĂŠrimĂ¨tre accessible (staff, admin...).
        - ResponseâŻ: application/json ; voir la structure exacte dans _build_candidat_meta.
        - Exemple clĂŠ du JSON : {"statut_choices": [...], ...}

        NoteÂ : la structure exacte dĂŠpend du code in _build_candidat_meta/user scope.
        """
        logger.debug("/candidats/meta called")
        data = _build_candidat_meta(request.user)
        logger.debug("/candidats/meta keys=%s", list(data.keys()))
        return Response(
            {
                "success": True,
                "message": "MĂŠtadonnĂŠes candidats rĂŠcupĂŠrĂŠes avec succĂ¨s.",
                "data": data,
            }
        )

    @action(detail=False, methods=["post"], url_path="bulk/validate-inscription")
    def bulk_validate_inscription(self, request):
        candidate_ids = _parse_candidate_ids(request.data)
        qs = self.filter_queryset(self.get_queryset()).filter(id__in=candidate_ids)
        result = CandidateBulkService.bulk_validate_inscription(qs, actor=request.user)
        return self.success_response(
            data=result,
            message="Transition bulk 'inscription validĂŠe' exĂŠcutĂŠe.",
        )

    @action(detail=False, methods=["post"], url_path="bulk/start-formation")
    def bulk_start_formation(self, request):
        candidate_ids = _parse_candidate_ids(request.data)
        qs = self.filter_queryset(self.get_queryset()).filter(id__in=candidate_ids)
        result = CandidateBulkService.bulk_start_formation(qs, actor=request.user)
        return self.success_response(
            data=result,
            message="Transition bulk 'entrĂŠe en formation' exĂŠcutĂŠe.",
        )

    @action(detail=False, methods=["post"], url_path="bulk/abandon")
    def bulk_abandon(self, request):
        candidate_ids = _parse_candidate_ids(request.data)
        qs = self.filter_queryset(self.get_queryset()).filter(id__in=candidate_ids)
        result = CandidateBulkService.bulk_abandon(qs, actor=request.user)
        return self.success_response(
            data=result,
            message="Transition bulk 'abandon' exĂŠcutĂŠe.",
        )

    @action(detail=False, methods=["post"], url_path="bulk/assign-atelier-tre")
    def bulk_assign_atelier_tre(self, request):
        candidate_ids = _parse_candidate_ids(request.data)
        atelier_id = request.data.get("atelier_tre_id")
        if not isinstance(atelier_id, int):
            raise ValidationError({"atelier_tre_id": ["Ce champ est obligatoire et doit ĂŞtre un entier."]})

        atelier_qs = atelier_tre.AtelierTRE.objects.select_related("centre")
        user = request.user
        if not is_admin_like(user):
            centre_ids = staff_centre_ids(user) or []
            atelier_qs = atelier_qs.filter(centre_id__in=centre_ids)
        atelier = atelier_qs.filter(id=atelier_id).first()
        if not atelier:
            raise PermissionDenied("Atelier TRE hors de votre pĂŠrimĂ¨tre.")

        qs = self.filter_queryset(self.get_queryset()).filter(id__in=candidate_ids)
        result = CandidateBulkService.bulk_assign_atelier_tre(qs, atelier=atelier)
        return self.success_response(
            data=result,
            message="Affectation bulk Ă  l'atelier TRE exĂŠcutĂŠe.",
        )

    @action(detail=False, methods=["get"], url_path="export-xlsx")
    def export_xlsx(self, request):
        """GET : export de la liste filtrĂŠe en XLSX (attachment). MĂŞme queryset et filtres que list."""
        qs = self.filter_queryset(self.get_queryset())
        logger.debug("export XLSX candidats params=%s rows=%d", self._qp_dict(request), qs.count())

        wb = Workbook()
        ws = wb.active
        ws.title = "Candidats"

        try:
            logo_path = Path(settings.BASE_DIR) / "rap_app/static/images/logo.png"
            if logo_path.exists():
                img = XLImage(str(logo_path))
                img.height = 60
                img.width = 120
                ws.add_image(img, "A1")
        except Exception:
            pass

        ws.merge_cells("B1:AF1")
        ws["B1"] = "Export complet des candidats â Rap_App"
        ws["B1"].font = Font(name="Calibri", bold=True, size=15, color="004C99")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("B2:AF2")
        ws["B2"] = f"Export rĂŠalisĂŠ le {dj_timezone.now().strftime('%d/%m/%Y Ă  %H:%M')}"
        ws["B2"].font = Font(name="Calibri", italic=True, size=10, color="666666")
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

        ws.append([])
        ws.append([])

        sep_row = ws.max_row + 1
        ws.append(["" for _ in range(10)])
        for cell in ws[sep_row]:
            cell.fill = PatternFill("solid", fgColor="BDD7EE")
        ws.row_dimensions[sep_row].height = 5

        ws.append([])

        headers = [
            "ID",
            "Sexe",
            "Nom de naissance",
            "Nom dâusage",
            "PrĂŠnom",
            "Date de naissance",
            "DĂŠpartement de naissance",
            "Commune de naissance",
            "Pays de naissance",
            "NationalitĂŠ",
            "NIR",
            "Ăge",
            "Email",
            "TĂŠlĂŠphone",
            "NumĂŠro de voie",
            "Nom de la rue",
            "ComplĂŠment dâadresse",
            "Code postal",
            "Ville",
            "Statut",
            "CV",
            "Type de contrat",
            "DisponibilitĂŠ",
            "Entretien rĂŠalisĂŠ",
            "Test dâentrĂŠe OK",
            "RQTH",
            "Permis B",
            "Dernier diplĂ´me prĂŠparĂŠ",
            "DiplĂ´me/titre obtenu",
            "DerniĂ¨re classe frĂŠquentĂŠe",
            "IntitulĂŠ diplĂ´me prĂŠparĂŠ",
            "Situation avant contrat",
            "RĂŠgime social",
            "Sportif de haut niveau",
            "Ăquivalence jeunes",
            "Extension BOE",
            "Situation actuelle",
            "Lien reprĂŠsentant",
            "Nom naissance reprĂŠsentant",
            "PrĂŠnom reprĂŠsentant",
            "Email reprĂŠsentant",
            "Adresse reprĂŠsentant",
            "CP reprĂŠsentant",
            "Ville reprĂŠsentant",
            "Formation",
            "Num offre",
            "Centre formation",
            "Type formation",
            "Origine sourcing",
            "Date inscription",
            "RĂŠsultat placement",
            "Contrat signĂŠ",
            "Date placement",
            "Entreprise placement",
            "Entreprise validĂŠe",
            "Responsable placement",
            "Vu par (staff)",
            "Nb appairages",
            "Nb prospections",
            "Inscrit GESPERS",
            "Courrier rentrĂŠe envoyĂŠ",
            "Date rentrĂŠe",
            "Admissible",
            "OSIA",
            "Communication â",
            "ExpĂŠrience â",
            "CSP â",
            "Projet crĂŠation entreprise",
            "Notes",
        ]
        ws.append(headers)

        header_row = ws.max_row
        header_fill = PatternFill("solid", fgColor="DCE6F1")
        border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        for cell in ws[header_row]:
            cell.font = Font(name="Calibri", bold=True, color="002060")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            cell.fill = header_fill
            cell.border = border
        ws.row_dimensions[header_row].height = 28

        even_fill = PatternFill("solid", fgColor="F8FBFF")
        odd_fill = PatternFill("solid", fgColor="FFFFFF")

        for i, c in enumerate(qs, start=1):
            ws.append(
                [
                    c.id,
                    c.sexe or "",
                    c.nom_naissance or "",
                    c.nom or "",
                    c.prenom or "",
                    c.date_naissance.strftime("%d/%m/%Y") if c.date_naissance else "",
                    c.departement_naissance or "",
                    c.commune_naissance or "",
                    c.pays_naissance or "",
                    c.nationalite or "",
                    c.nir or "",
                    c.age or "",
                    c.email or "",
                    c.telephone or "",
                    c.street_number or "",
                    c.street_name or "",
                    c.street_complement or "",
                    c.code_postal or "",
                    c.ville or "",
                    c.get_statut_display() if hasattr(c, "get_statut_display") else c.statut,
                    c.get_cv_statut_display() if hasattr(c, "get_cv_statut_display") else c.cv_statut,
                    c.get_type_contrat_display() if hasattr(c, "get_type_contrat_display") else c.type_contrat,
                    c.get_disponibilite_display() if hasattr(c, "get_disponibilite_display") else c.disponibilite,
                    "Oui" if c.entretien_done else "Non",
                    "Oui" if c.test_is_ok else "Non",
                    "Oui" if c.rqth else "Non",
                    "Oui" if c.permis_b else "Non",
                    c.dernier_diplome_prepare or "",
                    c.diplome_plus_eleve_obtenu or "",
                    c.derniere_classe or "",
                    c.intitule_diplome_prepare or "",
                    c.situation_avant_contrat or "",
                    c.regime_social or "",
                    "Oui" if c.sportif_haut_niveau else "Non",
                    "Oui" if c.equivalence_jeunes else "Non",
                    "Oui" if c.extension_boe else "Non",
                    c.situation_actuelle or "",
                    c.representant_lien or "",
                    c.representant_nom_naissance or "",
                    c.representant_prenom or "",
                    c.representant_email or "",
                    c.representant_street_name or "",
                    c.representant_zip_code or "",
                    c.representant_city or "",
                    getattr(c.formation, "nom", "") if c.formation else "",
                    getattr(c.formation, "num_offre", "") if c.formation else "",
                    getattr(getattr(c.formation, "centre", None), "nom", "") if c.formation else "",
                    getattr(getattr(c.formation, "type_offre", None), "nom", "") if c.formation else "",
                    c.origine_sourcing or "",
                    c.date_inscription.strftime("%d/%m/%Y") if c.date_inscription else "",
                    (
                        c.get_resultat_placement_display()
                        if hasattr(c, "get_resultat_placement_display")
                        else c.resultat_placement
                    ),
                    c.get_contrat_signe_display() if hasattr(c, "get_contrat_signe_display") else c.contrat_signe,
                    c.date_placement.strftime("%d/%m/%Y") if c.date_placement else "",
                    getattr(c.entreprise_placement, "nom", ""),
                    getattr(c.entreprise_validee, "nom", ""),
                    getattr(c.responsable_placement, "username", ""),
                    getattr(c.vu_par, "username", ""),
                    getattr(c, "nb_appairages_calc", 0),
                    getattr(c, "nb_prospections_calc", 0),
                    "Oui" if c.inscrit_gespers else "Non",
                    "Oui" if c.courrier_rentree else "Non",
                    c.date_rentree.strftime("%d/%m/%Y") if c.date_rentree else "",
                    "Oui" if c.admissible else "Non",
                    c.numero_osia or "",
                    c.communication or "",
                    c.experience or "",
                    c.csp or "",
                    "Oui" if c.projet_creation_entreprise else "Non",
                    (c.notes or "").replace("\n", " "),
                ]
            )

            fill = even_fill if i % 2 == 0 else odd_fill
            for j, cell in enumerate(ws[ws.max_row], start=1):
                cell.fill = fill
                cell.border = border
                cell.font = Font(name="Calibri", size=10, color="333333")
                cell.alignment = Alignment(vertical="top", wrapText=True)

            ws.row_dimensions[ws.max_row].height = 22

        # ==========================================================
        # đ Filtres + gel dâen-tĂŞte
        # ==========================================================
        end_row = ws.max_row
        last_col_letter = get_column_letter(len(headers))
        if end_row > header_row:
            ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{end_row}"
        ws.freeze_panes = f"A{header_row + 1}"

        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 45)

        ws.append([])
        ws.append([""])
        ws.append([f"Nombre total de candidats exportĂŠs : {qs.count()}"])
        ws[ws.max_row][0].font = Font(name="Calibri", bold=True, color="004C99", size=11)

        ws.oddFooter.center.text = f"ÂŠ Rap_App â export gĂŠnĂŠrĂŠ le {dj_timezone.now().strftime('%d/%m/%Y %H:%M')}"

        buffer = BytesIO()
        wb.save(buffer)
        binary_content = buffer.getvalue()

        filename = f'candidats_{dj_timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            binary_content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = len(binary_content)
        return response
