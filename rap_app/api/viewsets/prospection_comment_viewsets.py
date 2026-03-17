# apps/.../api/viewsets/prospection_comment_viewsets.py
import datetime
import logging
from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.db.models import Q
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.templatetags.static import static
from django.utils import timezone as dj_timezone
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import (
    OpenApiParameter,
    OpenApiResponse,
    extend_schema,
    extend_schema_view,
)
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.response import Response
from weasyprint import HTML

from ...models.prospection import HistoriqueProspection
from ...models.prospection_comments import ProspectionComment
from ..paginations import RapAppPagination
from ..permissions import IsOwnerOrStaffOrAbove
from ..roles import (
    is_admin_like,
    is_candidate,
    is_staff_like,
    is_staff_or_staffread,
    role_of,
    staff_centre_ids,
)
from ..serializers.prospection_comment_serializers import ProspectionCommentSerializer

logger = logging.getLogger("PROSPECTION_COMMENT")


@extend_schema_view(
    list=extend_schema(
        summary="📋 Liste des commentaires de prospection",
        tags=["ProspectionComments"],
        parameters=[
            OpenApiParameter("prospection", int, description="Filtrer par prospection ID"),
            OpenApiParameter("is_internal", bool, description="Filtrer par interne/public"),
            OpenApiParameter("created_by", int, description="Filtrer par auteur (ID)"),
            OpenApiParameter("partenaire_nom", str, description="Filtrer par nom de partenaire (icontains)"),
            OpenApiParameter("formation_nom", str, description="Filtrer par nom de formation (icontains)"),
            OpenApiParameter("created_by_username", str, description="Filtrer par username auteur (icontains)"),
            OpenApiParameter("search", str, description="Recherche (body, auteur, partenaire, formation)"),
            OpenApiParameter("ordering", str, description="created_at, -created_at, id, -id"),
        ],
        responses={200: OpenApiResponse(response=ProspectionCommentSerializer(many=True))},
    ),
    retrieve=extend_schema(summary="🔍 Détail d’un commentaire", tags=["ProspectionComments"]),
    create=extend_schema(summary="➕ Créer un commentaire", tags=["ProspectionComments"]),
    update=extend_schema(summary="✏️ Modifier un commentaire", tags=["ProspectionComments"]),
    partial_update=extend_schema(summary="✏️ Modifier partiellement un commentaire", tags=["ProspectionComments"]),
    destroy=extend_schema(summary="🗑️ Supprimer un commentaire", tags=["ProspectionComments"]),
)
class ProspectionCommentViewSet(viewsets.ModelViewSet):
    """
    ViewSet des commentaires de prospection.

    Source de vérité actuelle :
    - accès protégé par `IsOwnerOrStaffOrAbove`
    - visibilité dépendante du rôle, du caractère interne du commentaire,
      et des liens owner/created_by/prospection
    - actions custom dédiées au filtrage, à l'archivage et aux exports
    - plusieurs réponses d'actions custom sortent volontairement du CRUD
      strict pour exposer soit du JSON métier, soit des téléchargements
    """

    queryset = ProspectionComment.objects.select_related(
        "prospection",
        "prospection__partenaire",
        "prospection__formation",
        "prospection__owner",
        "prospection__created_by",
        "created_by",
    )
    serializer_class = ProspectionCommentSerializer
    permission_classes = [IsOwnerOrStaffOrAbove]
    pagination_class = RapAppPagination

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = [
        "prospection",
        "is_internal",
        "created_by",
        "prospection__owner",
        "prospection__partenaire",
    ]
    search_fields = [
        "body",
        "created_by__username",
        "prospection__partenaire__nom",
        "prospection__formation__nom",
        "prospection__owner__username",
    ]
    ordering_fields = ["created_at", "updated_at", "id"]
    ordering = ["-updated_at", "-created_at"]

    def get_queryset(self):
        """
        Retourne les commentaires visibles pour l'utilisateur courant en
        appliquant les règles de rôle, le filtre d'archivage et les filtres
        additionnels de requête.
        """
        base = ProspectionComment.objects.select_related(
            "prospection",
            "prospection__owner",
            "prospection__created_by",
            "prospection__partenaire",
            "prospection__formation",
            "prospection__formation__centre",
            "created_by",
        )

        u = getattr(self.request, "user", None)
        if not getattr(u, "is_authenticated", False):
            return base.none()

        qs = base

        # 🔹 Filtrage par rôle
        if is_candidate(u):
            qs = qs.filter(prospection__owner_id=u.id).filter(Q(is_internal=False) | Q(created_by_id=u.id))
        elif is_staff_like(u):
            centre_ids = staff_centre_ids(u) or []
            if centre_ids:
                qs = qs.filter(
                    Q(prospection__formation__centre_id__in=centre_ids)
                    | Q(prospection__owner=u)
                    | Q(prospection__created_by=u)
                    | Q(created_by=u)
                ).distinct()
            else:
                qs = qs.filter(Q(prospection__owner=u) | Q(prospection__created_by=u) | Q(created_by=u)).distinct()
        elif not is_admin_like(u):
            return base.none()

        # ✅ Ne pas filtrer sur "actif" pour la vue de détail
        # (permet d'afficher les commentaires archivés)
        if getattr(self, "action", None) == "retrieve":
            return qs.order_by("-created_at", "-id").distinct()

        # 🔹 Filtre supplémentaire : ?est_archive=true|false|both
        qp = self.request.query_params
        est_archive = qp.get("est_archive")

        if est_archive is None:
            # ✅ Par défaut : uniquement les actifs
            qs = qs.filter(statut_commentaire="actif")
        else:
            val = est_archive.lower()
            if val in ("both", "all", "tous", "tout"):
                # ✅ Inclure actifs + archivés
                pass  # aucun filtre
            elif val in ("1", "true", "yes", "oui", "archive", "archived"):
                # ✅ Seulement les archivés
                qs = qs.filter(statut_commentaire="archive")
            elif val in ("0", "false", "no", "non", "actif", "active"):
                # ✅ Seulement les actifs
                qs = qs.filter(statut_commentaire="actif")
            else:
                # Par sécurité : défaut = actifs
                qs = qs.filter(statut_commentaire="actif")

        # 🔹 Filtres additionnels
        part_nom = (qp.get("partenaire_nom") or "").strip()
        form_nom = (qp.get("formation_nom") or "").strip()
        author_username = (qp.get("created_by_username") or "").strip()

        if part_nom:
            qs = qs.filter(prospection__partenaire__nom__icontains=part_nom)
        if form_nom:
            qs = qs.filter(prospection__formation__nom__icontains=form_nom)
        if author_username:
            qs = qs.filter(created_by__username__icontains=author_username)

        owner_id = qp.get("prospection_owner")
        partenaire_id = qp.get("prospection_partenaire")
        if owner_id:
            qs = qs.filter(prospection__owner_id=owner_id)
        if partenaire_id:
            qs = qs.filter(prospection__partenaire_id=partenaire_id)

        return qs.order_by("-updated_at", "-created_at", "-id").distinct()

    @action(detail=False, methods=["get"], url_path="filter-options")
    def filter_options(self, request):
        """
        Retourne les valeurs distinctes utilisables comme options de
        filtre pour les commentaires (formations, partenaires, auteurs,
        centres et propriétaires).
        """
        base_qs = ProspectionComment.objects.select_related(
            "prospection__formation__centre",
            "prospection__partenaire",
            "prospection__owner",
            "created_by",
        )

        # Récup formations
        formations = (
            base_qs.filter(prospection__formation__isnull=False)
            .values_list("prospection__formation__nom", flat=True)
            .distinct()
        )
        formations_choices = [{"value": f, "label": f} for f in formations if f]
        ("[DEBUG filter_options] formations =", list(formations))

        # Récup partenaires
        partenaires = (
            base_qs.filter(prospection__partenaire__isnull=False)
            .values_list("prospection__partenaire__nom", flat=True)
            .distinct()
        )
        partenaires_choices = [{"value": p, "label": p} for p in partenaires if p]
        ("[DEBUG filter_options] partenaires =", list(partenaires))

        # Récup auteurs
        authors = base_qs.filter(created_by__isnull=False).values_list("created_by__username", flat=True).distinct()
        authors_choices = [{"value": a, "label": a} for a in authors if a]
        ("[DEBUG filter_options] authors =", list(authors))

        # Récup centres
        centres = (
            base_qs.filter(prospection__formation__centre__isnull=False)
            .values_list("prospection__formation__centre__nom", flat=True)
            .distinct()
        )
        centres_choices = [{"value": c, "label": c} for c in centres if c]
        ("[DEBUG filter_options] centres =", list(centres))

        # 🆕 Récup owners (id + username)
        owners = (
            base_qs.filter(prospection__owner__isnull=False)
            .values_list("prospection__owner__id", "prospection__owner__username")
            .distinct()
        )
        owners_list = list(owners)
        ("[DEBUG filter_options] owners =", owners_list)

        owners_choices = [{"value": oid, "label": uname} for oid, uname in owners_list if oid and uname]

        return Response(
            {
                "formations": formations_choices,
                "partenaires": partenaires_choices,
                "authors": authors_choices,
                "centres": centres_choices,
                "owners": owners_choices,
            }
        )

    def list(self, request, *args, **kwargs):
        """
        Liste paginée des commentaires accessibles, enrichis avec les
        informations de centre et type d'offre de la formation liée.
        """
        qs = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            data = serializer.data

            # enrichissement pour chaque item de la page
            for item, obj in zip(data, page):
                formation = getattr(obj.prospection, "formation", None)
                if formation:
                    item["formation_centre_nom"] = getattr(getattr(formation, "centre", None), "nom", None)
                    item["formation_type_offre_nom"] = getattr(getattr(formation, "type_offre", None), "nom", None)
                else:
                    item["formation_centre_nom"] = None
                    item["formation_type_offre_nom"] = None

            response = self.get_paginated_response(data)

        else:
            # fallback si pagination désactivée, mais on garde la même forme
            serializer = self.get_serializer(qs, many=True)
            data = serializer.data
            for item, obj in zip(data, qs):
                formation = getattr(obj.prospection, "formation", None)
                if formation:
                    item["formation_centre_nom"] = getattr(getattr(formation, "centre", None), "nom", None)
                    item["formation_type_offre_nom"] = getattr(getattr(formation, "type_offre", None), "nom", None)
                else:
                    item["formation_centre_nom"] = None
                    item["formation_type_offre_nom"] = None

            response = Response(
                {
                    "count": len(data),
                    "next": None,
                    "previous": None,
                    "results": data,
                }
            )

        # extra headers si tu veux les garder
        response["X-PC-Role"] = role_of(request.user)
        response["X-PC-User"] = str(getattr(request.user, "id", "anon"))
        response["X-PC-Count"] = str(len(data))

        return response

    def retrieve(self, request, *args, **kwargs):
        """
        Retourne le détail d'un commentaire et ajoute les informations
        de centre et type d'offre de la formation associée.
        """
        obj: ProspectionComment = self.get_object()
        serializer = self.get_serializer(obj)
        data = serializer.data

        # idem : enrichissement uniquement pour centre + type_offre
        formation = getattr(obj.prospection, "formation", None)
        if formation:
            data["formation_centre_nom"] = getattr(getattr(formation, "centre", None), "nom", None)
            data["formation_type_offre_nom"] = getattr(getattr(formation, "type_offre", None), "nom", None)
        else:
            data["formation_centre_nom"] = None
            data["formation_type_offre_nom"] = None

        return Response(data)

    def perform_create(self, serializer):
        """
        Crée un commentaire en appliquant les règles de rôle
        (candidat, staff, admin) et en positionnant created_by.
        """
        u = self.request.user
        prosp = serializer.validated_data["prospection"]

        if is_candidate(u):
            if prosp.owner_id != u.id:
                raise PermissionDenied("Vous ne pouvez commenter que vos propres prospections.")
            serializer.validated_data["is_internal"] = False
        elif is_staff_like(u) and not is_admin_like(u):
            cids = staff_centre_ids(u) or []
            form_centre_id = getattr(getattr(prosp, "formation", None), "centre_id", None)
            if form_centre_id and cids and form_centre_id not in cids:
                raise PermissionDenied("Prospection hors de votre périmètre (centres).")

        serializer.save(created_by=u)

    def perform_update(self, serializer):
        """
        Met à jour un commentaire en contrôlant les droits des
        candidats et du staff sur la prospection et le champ interne.
        """
        u = self.request.user
        obj: ProspectionComment = self.get_object()

        if is_candidate(u):
            if obj.prospection.owner_id != u.id:
                raise PermissionDenied("Accès refusé.")
            if obj.created_by_id != u.id:
                raise PermissionDenied("Vous ne pouvez modifier que vos propres commentaires.")
            if serializer.validated_data.get("is_internal", obj.is_internal):
                raise PermissionDenied("Un candidat ne peut pas rendre un commentaire interne.")
            new_prosp = serializer.validated_data.get("prospection")
            if new_prosp and new_prosp.id != obj.prospection_id:
                raise PermissionDenied("Vous ne pouvez pas changer la prospection d'un commentaire.")
            serializer.save()
            return

        if is_staff_like(u) and not is_admin_like(u):
            new_prosp = serializer.validated_data.get("prospection", obj.prospection)
            cids = staff_centre_ids(u) or []
            form_centre_id = getattr(getattr(new_prosp, "formation", None), "centre_id", None)
            if form_centre_id and cids and form_centre_id not in cids:
                raise PermissionDenied("Prospection hors de votre périmètre (centres).")

        serializer.save()

    def destroy(self, request, *args, **kwargs):
        """
        Supprime un commentaire après vérification que le candidat
        respecte les contraintes de propriété et de visibilité.
        """
        u = request.user
        obj: ProspectionComment = self.get_object()

        if is_candidate(u):
            if obj.prospection.owner_id != u.id or obj.created_by_id != u.id or obj.is_internal:
                raise PermissionDenied("Vous ne pouvez pas supprimer ce commentaire.")
        return super().destroy(request, *args, **kwargs)

    # ------------------------------------------------------------------
    # 🔒 ARCHIVER / DÉSARCHIVER un commentaire
    # ------------------------------------------------------------------

    @action(detail=True, methods=["post"], url_path="archiver")
    def archiver(self, request, pk=None):
        """
        Archive logiquement un commentaire en mettant à jour son statut.
        """
        comment = self.get_object()
        if comment.est_archive:
            return Response({"detail": "Déjà archivé."}, status=status.HTTP_200_OK)

        comment.archiver(save=True)
        logger.info("Commentaire #%s archivé par %s", comment.pk, request.user)
        return Response({"detail": "Commentaire archivé."}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="desarchiver")
    def desarchiver(self, request, pk=None):
        """
        Désarchive un commentaire précédemment archivé.
        """
        comment = self.get_object()
        if not comment.est_archive:
            return Response({"detail": "Déjà actif."}, status=status.HTTP_200_OK)

        comment.desarchiver(save=True)
        logger.info("Commentaire #%s désarchivé par %s", comment.pk, request.user)
        return Response({"detail": "Commentaire désarchivé."}, status=status.HTTP_200_OK)

    # ------------------------------------------------------------------
    # 📊 EXPORT EXCEL — Commentaires de prospection
    # ------------------------------------------------------------------
    @extend_schema(summary="Exporter les commentaires de prospection au format XLSX")
    @action(detail=False, methods=["get"], url_path="export-xlsx")
    def export_xlsx(self, request):
        """
        Exporte au format Excel les commentaires visibles pour
        l'utilisateur, ou renvoie une réponse vide si aucun résultat.
        """
        qs = self.get_queryset()  # ✅ pas de filter_queryset()

        if not qs.exists():
            return Response({"detail": "Aucun commentaire à exporter."}, status=204)

        # ==========================================================
        # 📘 Création du classeur
        # ==========================================================
        wb = Workbook()
        ws = wb.active
        ws.title = "Commentaires prospection"

        # ==========================================================
        # 🖼️ Logo Rap_App (visible en dev & prod)
        # ==========================================================
        try:
            logo_path = Path(settings.BASE_DIR) / "rap_app/static/images/logo.png"
            if logo_path.exists():
                img = XLImage(str(logo_path))
                img.height = 45
                img.width = 120
                ws.add_image(img, "A1")
        except Exception:
            pass

        # ==========================================================
        # 🧾 Titre + date d’export
        # ==========================================================
        ws.merge_cells("B1:H1")
        ws["B1"] = "Commentaires de prospection — Rap_App"
        ws["B1"].font = Font(bold=True, size=14, color="0077CC")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("B2:H2")
        ws["B2"] = f"Export réalisé le {dj_timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        ws["B2"].font = Font(italic=True, size=10, color="555555")
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

        ws.append([])  # ligne vide

        # ==========================================================
        # 📋 En-têtes
        # ==========================================================
        headers = [
            "ID",
            "Statut du commentaire",
            "Prospection",
            "Partenaire",
            "Formation",
            "Auteur",
            "Commentaire",
            "Créé le",
        ]
        ws.append(headers)

        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="E9F2FF")

        # ==========================================================
        # 🧮 Données
        # ==========================================================
        for c in qs:
            statut_display = dict(c.STATUT_CHOICES).get(c.statut_commentaire, c.statut_commentaire)
            statut_color = "C8E6C9" if c.statut_commentaire == "actif" else "E0E0E0"  # vert / gris

            row = [
                c.id,
                str(statut_display),  # ✅ conversion explicite en string
                str(getattr(c.prospection, "id", "—")),
                str(getattr(getattr(c.prospection, "partenaire", None), "nom", "—")),
                str(getattr(getattr(c.prospection, "formation", None), "nom", "—")),
                str(getattr(c.created_by, "username", "—")),
                str(c.body or ""),
                c.created_at.strftime("%d/%m/%Y %H:%M") if c.created_at else "—",
            ]
            ws.append(row)

            # Couleur du statut
            statut_cell = ws[f"B{ws.max_row}"]
            statut_cell.fill = PatternFill("solid", fgColor=statut_color)
            statut_cell.alignment = Alignment(horizontal="center", vertical="center")
            statut_cell.font = Font(bold=True)

        # ==========================================================
        # 📏 Largeur des colonnes + wrap texte
        # ==========================================================
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            if col_letter == "G":  # Colonne "Commentaire"
                ws.column_dimensions[col_letter].width = 80
                for cell in col:
                    cell.alignment = Alignment(wrapText=True, vertical="top")
            else:
                max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        # ==========================================================
        # 📤 Sauvegarde et réponse HTTP
        # ==========================================================
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        binary = buffer.getvalue()

        filename = f'prospection_commentaires_{dj_timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            binary,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = len(binary)
        return response

    # ------------------------------------------------------------------
    # 📄 EXPORT PDF — Commentaires de prospection
    # ------------------------------------------------------------------
    @extend_schema(summary="Exporter les commentaires de prospection au format PDF")
    @action(detail=False, methods=["get"], url_path="export-pdf")
    def export_pdf(self, request):
        """
        Exporte en PDF les commentaires visibles pour l'utilisateur en
        utilisant un template HTML, ou renvoie 204 s'il n'y a aucun
        commentaire.
        """
        qs = self.get_queryset()  # ✅ pas de filter_queryset()

        if not qs.exists():
            return Response({"detail": "Aucun commentaire à exporter."}, status=204)

        # ==========================================================
        # 🧩 Préparation du logo
        # ==========================================================
        try:
            logo_url = request.build_absolute_uri(static("images/logo.png"))
        except Exception:
            logo_path = settings.BASE_DIR / "rap_app/static/images/logo.png"
            logo_url = f"file://{logo_path}"

        # ==========================================================
        # 🧮 Construction des données pour le template
        # ==========================================================
        commentaires_data = []
        for c in qs:
            prosp = getattr(c, "prospection", None)
            formation = getattr(prosp, "formation", None)
            partenaire = getattr(prosp, "partenaire", None)
            statut_display = dict(c.STATUT_CHOICES).get(c.statut_commentaire, c.statut_commentaire)

            commentaires_data.append(
                {
                    "id": c.id,
                    "statut_commentaire": statut_display,
                    "body": c.body or "",
                    "created_by": getattr(c.created_by, "username", "—"),
                    "created_at": c.created_at,
                    "activite": getattr(c, "activite", ""),
                    "prospection": {
                        "id": getattr(prosp, "id", None),
                        "date_prospection": getattr(prosp, "date_prospection", None),
                        "statut": getattr(prosp, "statut_display", getattr(prosp, "statut", None)),
                    },
                    "formation": {
                        "nom": getattr(formation, "nom", None),
                        "centre_nom": getattr(getattr(formation, "centre", None), "nom", None),
                        "start_date": getattr(formation, "date_debut", getattr(formation, "start_date", None)),
                        "end_date": getattr(formation, "date_fin", getattr(formation, "end_date", None)),
                        "type_offre": getattr(getattr(formation, "type_offre", None), "nom", None),
                    },
                    "partenaire": {
                        "nom": getattr(partenaire, "nom", None),
                    },
                }
            )

        # ==========================================================
        # 📦 Contexte du template
        # ==========================================================
        context = {
            "commentaires": commentaires_data,
            "now": dj_timezone.now(),
            "logo_url": logo_url,
            "user": request.user,
        }

        # ==========================================================
        # 🧾 Rendu HTML -> PDF
        # ==========================================================
        html_string = render_to_string("exports/prospection_commentaires_pdf.html", context)
        pdf = HTML(string=html_string, base_url=request.build_absolute_uri("/")).write_pdf()

        # ==========================================================
        # 📤 Réponse HTTP
        # ==========================================================
        filename = f'prospection_commentaires_{dj_timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = len(pdf)
        return response
