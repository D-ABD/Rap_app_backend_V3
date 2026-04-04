"""ViewSet des objectifs Prépa."""

from collections import OrderedDict
from io import BytesIO

from django.http import HttpResponse
from drf_spectacular.utils import (
    OpenApiParameter,
    extend_schema,
    extend_schema_view,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from ...api.roles import (
    is_admin_like,
    is_candidate,
    is_prepa_staff,
    is_staff_read,
    is_staff_standard,
)
from ...models.centres import Centre
from ...models.prepa import ObjectifPrepa, Prepa
from ..mixins import ApiResponseMixin, HardDeleteArchivedMixin
from ..permissions import IsPrepaStaffOrAbove
from ..serializers.prepa_objectifs_serializers import ObjectifPrepaSerializer


@extend_schema_view(
    list=extend_schema(
        summary="Lister tous les objectifs Prépa",
        description="Retourne la liste paginée des objectifs Prépa (format DRF standard).",
        parameters=[
            OpenApiParameter(name="annee", description="Filtrer par année", required=False, type=int),
            OpenApiParameter(
                name="centre_id", description="Filtrer par identifiant de centre", required=False, type=int
            ),
            OpenApiParameter(
                name="departement", description="Filtrer par code département (ex: 59, 75)", required=False, type=str
            ),
        ],
        responses={200: ObjectifPrepaSerializer(many=True)},
    ),
    destroy=extend_schema(
        summary="Archiver un objectif Prépa",
        description="Archive logiquement un objectif Prépa en le désactivant (`is_active = False`).",
        responses={200: ObjectifPrepaSerializer},
    ),
)
class ObjectifPrepaViewSet(HardDeleteArchivedMixin, ApiResponseMixin, viewsets.ModelViewSet):
    """
    ViewSet CRUD des objectifs annuels Prépa.

    Le périmètre est restreint par centre selon les rôles Prépa/admin/staff :
    `prepa_staff` reste le rôle métier principal, `staff` et `staff_read`
    gardent un accès transverse, tandis que `commercial` et
    `charge_recrutement` restent exclus de ce bloc spécialisé.

    Le fichier expose aussi des actions de filtres, synthèse et export Excel.
    """

    serializer_class = ObjectifPrepaSerializer
    permission_classes = [IsAuthenticated, IsPrepaStaffOrAbove]
    queryset = ObjectifPrepa.objects.select_related("centre").all()
    hard_delete_enabled = True
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["centre__nom", "centre__code_postal", "annee"]
    ordering_fields = ["annee", "centre__nom"]

    # -----------------------------------------------------------
    # 🔹 Helpers de périmètre (scope)
    # -----------------------------------------------------------
    def _centre_ids_for_user(self, user):
        # Retourne les identifiants de centres accessibles en fonction
        # du rôle Prépa de l'utilisateur. Les règles détaillées sont
        # déléguées aux helpers de rôles.
        if is_admin_like(user):
            return None
        if is_prepa_staff(user):
            centres = getattr(user, "centres_acces", None) or getattr(user, "centres", None)
            return list(centres.values_list("id", flat=True)) if centres else []
        if is_staff_standard(user) or is_staff_read(user):
            try:
                return list(user.centres.values_list("id", flat=True))
            except Exception:
                return []
        return []

    def _scope_qs_to_user_centres(self, qs):
        # Restreint un queryset ObjectifPrepa au périmètre de centres
        # autorisés pour l'utilisateur courant, ou renvoie qs.none()
        # pour les profils non autorisés.
        user = self.request.user
        if not user.is_authenticated or is_candidate(user):
            return qs.none()
        centre_ids = self._centre_ids_for_user(user)
        if centre_ids is None:
            return qs
        if centre_ids:
            return qs.filter(centre_id__in=centre_ids).distinct()
        return qs.none()

    def _assert_user_can_use_centre(self, centre):
        # Lève PermissionDenied si le centre fourni n'est pas inclus
        # dans le périmètre de centres autorisés pour l'utilisateur.
        if not centre:
            return
        user = self.request.user
        if is_admin_like(user):
            return
        allowed = set(self._centre_ids_for_user(user) or [])
        if getattr(centre, "id", None) not in allowed:
            raise PermissionDenied("Centre hors de votre périmètre d'accès.")

    # -----------------------------------------------------------
    # 🔹 Queryset principal
    # -----------------------------------------------------------
    def get_queryset(self):
        # Retourne les ObjectifPrepa actifs visibles pour l'utilisateur,
        # filtrés par centre, année, identifiant de centre et
        # département, puis triés par année décroissante et nom de centre.
        user = self.request.user
        qs = ObjectifPrepa.objects.select_related("centre").filter(is_active=True)
        qs = self._scope_qs_to_user_centres(qs)

        params = self.request.query_params
        annee = params.get("annee")
        centre_id = params.get("centre_id")
        departement = params.get("departement")

        if annee:
            qs = qs.filter(annee=annee)
        if centre_id:
            qs = qs.filter(centre_id=centre_id)
        if departement:
            qs = qs.filter(centre__code_postal__startswith=departement)

        return qs.order_by("-annee", "centre__nom")

    # -----------------------------------------------------------
    # 🔹 list() — pagination DRF standard
    # -----------------------------------------------------------
    def list(self, request, *args, **kwargs):
        """
        Liste les objectifs Prépa accessibles à l'utilisateur, en
        appliquant les filtres de requête et la pagination DRF
        standard si elle est configurée.
        """
        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return self.success_response(data=serializer.data, message="Liste des objectifs Prépa récupérée avec succès.")

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return self.success_response(data=serializer.data, message="Objectif Prépa récupéré avec succès.")

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return self.success_response(
            data=self.get_serializer(serializer.instance).data,
            message="Objectif Prépa créé avec succès.",
            status_code=status.HTTP_201_CREATED,
        )

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return self.success_response(
            data=self.get_serializer(serializer.instance).data,
            message="Objectif Prépa mis à jour avec succès.",
        )

    def partial_update(self, request, *args, **kwargs):
        kwargs["partial"] = True
        return self.update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        """
        Conserve `DELETE` pour compatibilité mais remplace la
        suppression physique par un archivage logique.
        """
        instance = self.get_object()
        if not instance.is_active:
            serializer = self.get_serializer(instance)
            return self.success_response(
                data=serializer.data,
                message="Objectif Prépa déjà archivé.",
                status_code=status.HTTP_200_OK,
            )

        instance.is_active = False
        instance.save(user=request.user, update_fields=["is_active"])
        return self.success_response(
            data=self.get_serializer(instance).data,
            message="Objectif Prépa archivé avec succès.",
            status_code=status.HTTP_200_OK,
        )

    # -----------------------------------------------------------
    # 🔹 create / update sécurisés
    # -----------------------------------------------------------
    def perform_create(self, serializer):
        # Crée un objectif Prépa puis vérifie que le centre associé fait
        # partie du périmètre de l'utilisateur avant de sauvegarder avec
        # ou sans paramètre user selon la signature du modèle.
        self._assert_user_can_use_centre(serializer.validated_data.get("centre"))
        serializer.save()

    def perform_update(self, serializer):
        # Met à jour un objectif Prépa après contrôle de la portée sur
        # le centre éventuellement modifié, puis sauvegarde avec ou sans
        # paramètre user selon la signature du modèle.
        current = serializer.instance
        new_centre = serializer.validated_data.get("centre", getattr(current, "centre", None))
        self._assert_user_can_use_centre(new_centre)
        serializer.save()

    # -----------------------------------------------------------
    # 🔹 Filtres
    # -----------------------------------------------------------
    @action(detail=False, methods=["get"], url_path="filters")
    def filters(self, request):
        """
        Retourne les valeurs disponibles pour les filtres année, centre
        et département, limitées aux objectifs visibles de
        l'utilisateur.
        """
        qs = self._scope_qs_to_user_centres(ObjectifPrepa.objects.select_related("centre"))

        annees = qs.order_by("-annee").values_list("annee", flat=True).distinct()
        centres = qs.values("centre__id", "centre__nom", "centre__code_postal").distinct()
        departements = sorted({(c["centre__code_postal"] or "")[:2] for c in centres if c.get("centre__code_postal")})

        data = OrderedDict(
            annee=[{"value": a, "label": str(a)} for a in annees],
            centre=[
                {
                    "value": c["centre__id"],
                    "label": f'{c["centre__nom"]} ({c["centre__code_postal"]})',
                }
                for c in centres
            ],
            departement=[{"value": d, "label": f"Département {d}"} for d in departements],
        )
        return self.success_response(data=data, message="Filtres objectifs Prépa récupérés avec succès.")

    # -----------------------------------------------------------
    # 🔹 Synthèse
    # -----------------------------------------------------------
    @action(detail=False, methods=["get"], url_path="synthese")
    def synthese(self, request):
        """
        Retourne la liste des synthèses calculées par
        obj.synthese_globale() pour chaque objectif Prépa accessible à
        l'utilisateur.
        """
        qs = self.get_queryset()
        data = [obj.synthese_globale() for obj in qs]
        return self.success_response(data=data, message="Synthèse objectifs Prépa récupérée avec succès.")

    # -----------------------------------------------------------
    # 🔹 Export Excel
    # -----------------------------------------------------------
    @action(detail=False, methods=["get"], url_path="export-xlsx")
    def export_xlsx(self, request):
        """
        Exporte en fichier Excel les objectifs Prépa filtrés et
        accessibles à l'utilisateur, ou renvoie une erreur 404 s'il n'y
        a aucun objectif à exporter.
        """
        qs = self.get_queryset()
        if not qs.exists():
            return Response(
                {"success": False, "message": "Aucun objectif à exporter.", "data": None},
                status=404,
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Objectifs Prépa"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        # === En-têtes ===
        headers = [
            "Centre",
            "Département",
            "Année",
            "Objectif",
            "Réalisé (entrées Atelier 1)",
            "Adhésions",
            "Taux prescription (%)",
            "Taux présence IC (%)",
            "Taux adhésion IC (%)",
            "Taux atteinte (%)",
            "Taux rétention (%)",
            "Reste à faire",
        ]

        ws.append(headers)
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # === Données ===
        for obj in qs:
            # La structure de data dépend du retour de obj.synthese_globale().
            data = obj.synthese_globale()
            ws.append(
                [
                    data["centre"],
                    obj.centre.departement,
                    data["annee"],
                    data["objectif"],
                    data["realise"],
                    data["adhesions"],
                    data["taux_prescription"],
                    data["taux_presence"],
                    data["taux_adhesion"],
                    data["taux_atteinte"],
                    data.get("taux_retention", 0),
                    data["reste_a_faire"],
                ]
            )

        # === Ajustement largeur colonnes ===
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="objectifs_prepa.xlsx"'
        return response
