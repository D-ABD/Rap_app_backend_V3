from io import BytesIO

from django.db.models import Q
from django.http import HttpResponse
from django.utils.timezone import localdate
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from ...models.centres import Centre
from ...models.declic import Declic, ParticipantDeclic
from ..permissions import IsDeclicStaffOrAbove
from ..roles import is_admin_like, is_candidate, is_declic_staff, is_staff_or_staffread
from ..serializers.declic_serializers import ParticipantDeclicSerializer


class ParticipantDeclicViewSet(viewsets.ModelViewSet):
    """
    CRUD et exports du suivi nominatif des participants Déclic.
    """

    serializer_class = ParticipantDeclicSerializer
    permission_classes = [IsDeclicStaffOrAbove]
    queryset = ParticipantDeclic.objects.select_related(
        "centre", "declic_origine", "declic_origine__centre"
    ).all()

    def _admin_like(self, user) -> bool:
        return is_admin_like(user)

    def _accessible_centre_ids(self, user):
        if self._admin_like(user):
            return None
        if is_declic_staff(user) or is_staff_or_staffread(user):
            centres = getattr(user, "centres_acces", None) or getattr(user, "centres", None)
            if not centres or not centres.exists():
                return []
            return list(centres.values_list("id", flat=True))
        return []

    def _scope_qs(self, qs):
        user = self.request.user
        if not user.is_authenticated or is_candidate(user):
            return qs.none()

        centre_ids = self._accessible_centre_ids(user)
        if centre_ids is None:
            return qs
        if not centre_ids:
            return qs.none()

        model = getattr(qs, "model", None)
        if model and getattr(model._meta, "model_name", "") == "centre":
            return qs.filter(id__in=centre_ids)

        return qs.filter(centre_id__in=centre_ids)

    def get_queryset(self):
        qs = self._scope_qs(
            ParticipantDeclic.objects.select_related("centre", "declic_origine", "declic_origine__centre")
        )
        params = self.request.query_params

        search = params.get("search")
        centre = params.get("centre")
        declic_origine = params.get("declic_origine")
        type_declic = params.get("type_declic")
        annee = params.get("annee")
        present = params.get("present")
        ordering = params.get("ordering") or "nom"

        if search:
            qs = qs.filter(
                Q(nom__icontains=search)
                | Q(prenom__icontains=search)
                | Q(telephone__icontains=search)
                | Q(email__icontains=search)
                | Q(centre__nom__icontains=search)
            )
        if centre:
            qs = qs.filter(centre_id=centre)
        if declic_origine:
            qs = qs.filter(declic_origine_id=declic_origine)
        if type_declic:
            qs = qs.filter(declic_origine__type_declic=type_declic)
        if annee:
            qs = qs.filter(declic_origine__date_declic__year=annee)
        if present in {"true", "false"}:
            qs = qs.filter(present=(present == "true"))

        if ordering in {"nom", "-nom", "prenom", "-prenom", "updated_at", "-updated_at"}:
            qs = qs.order_by(ordering, "prenom", "id")
        else:
            qs = qs.order_by("nom", "prenom", "id")

        return qs

    @action(detail=False, methods=["get"], url_path="meta")
    def meta(self, request):
        centres = self._scope_qs(Centre.objects.all()).order_by("nom")
        annees = (
            self.get_queryset()
            .order_by()
            .values_list("declic_origine__date_declic__year", flat=True)
            .distinct()
        )
        annees = sorted([a for a in annees if a], reverse=True)
        declics_origine = (
            self._scope_qs(Declic.objects.select_related("centre").all())
            .order_by("-date_declic", "-id")[:200]
        )

        return Response(
            {
                "centres": [
                    {"id": c.id, "nom": c.nom, "departement": c.departement, "code_postal": c.code_postal}
                    for c in centres
                ],
                "presence_choices": [
                    {"value": "true", "label": "Présent"},
                    {"value": "false", "label": "Absent / non présent"},
                ],
                "type_declic": [
                    {"value": value, "label": label} for value, label in Declic.TypeDeclic.choices
                ],
                "declics_origine": [
                    {
                        "id": d.id,
                        "label": f"{d.get_type_declic_display()} du {d.date_declic:%d/%m/%Y}"
                        + (f" - {d.centre.nom}" if d.centre else ""),
                    }
                    for d in declics_origine
                ],
                "annees": annees or [localdate().year],
            }
        )

    def _filtered_export_qs(self, request):
        qs = self.filter_queryset(self.get_queryset())
        if request.method.lower() == "post":
            ids = request.data.get("ids") or []
            if ids:
                qs = qs.filter(id__in=ids)
        return qs

    @action(detail=False, methods=["get", "post"], url_path="export-xlsx")
    def export_xlsx(self, request):
        qs = self._filtered_export_qs(request)

        wb = Workbook()
        ws = wb.active
        ws.title = "Participants Declic"
        ws.append(
            [
                "Nom",
                "Prénom",
                "Téléphone",
                "Email",
                "Centre",
                "Séance Déclic",
                "Présent",
                "Commentaire",
            ]
        )

        for obj in qs:
            ws.append(
                [
                    obj.nom,
                    obj.prenom,
                    obj.telephone or "",
                    obj.email or "",
                    getattr(obj.centre, "nom", ""),
                    str(obj.declic_origine) if obj.declic_origine else "",
                    "Oui" if obj.present else "Non",
                    obj.commentaire_presence or "",
                ]
            )

        self._style_sheet(ws)
        return self._xlsx_response(wb, "participants_declic.xlsx")

    @action(detail=False, methods=["get", "post"], url_path="export-presence-xlsx")
    def export_presence_xlsx(self, request):
        qs = self._filtered_export_qs(request)
        wb = Workbook()
        ws = wb.active
        ws.title = "Presence Declic"
        title = "Feuille de présence - Participants Déclic"

        declic_origine = request.query_params.get("declic_origine")
        if declic_origine:
            declic = self._scope_qs(Declic.objects.select_related("centre").filter(id=declic_origine)).first()
            if declic:
                title = f"{title} - {declic.get_type_declic_display()} du {declic.date_declic:%d/%m/%Y}"

        ws.merge_cells("A1:G1")
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.append([])
        ws.append(["Nom", "Prénom", "Centre", "Téléphone", "Email", "Présent", "Observation"])

        for obj in qs:
            ws.append(
                [
                    obj.nom,
                    obj.prenom,
                    getattr(obj.centre, "nom", ""),
                    obj.telephone or "",
                    obj.email or "",
                    "",
                    "",
                ]
            )

        self._style_sheet(ws, header_row=3)
        return self._xlsx_response(wb, "participants_declic_presence.xlsx")

    @action(detail=False, methods=["get", "post"], url_path="export-emargement-xlsx")
    def export_emargement_xlsx(self, request):
        qs = self._filtered_export_qs(request)
        wb = Workbook()
        ws = wb.active
        ws.title = "Emargement Declic"
        title = "Feuille d'émargement - Participants Déclic"

        declic_origine = request.query_params.get("declic_origine")
        if declic_origine:
            declic = self._scope_qs(Declic.objects.select_related("centre").filter(id=declic_origine)).first()
            if declic:
                title = f"{title} - {declic.get_type_declic_display()} du {declic.date_declic:%d/%m/%Y}"

        ws.merge_cells("A1:H1")
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.append([])
        ws.append(["Nom", "Prénom", "Centre", "Téléphone", "Email", "Présent", "Heure", "Signature"])

        for obj in qs:
            ws.append(
                [
                    obj.nom,
                    obj.prenom,
                    getattr(obj.centre, "nom", ""),
                    obj.telephone or "",
                    obj.email or "",
                    "",
                    "",
                    "",
                ]
            )

        self._style_sheet(ws, header_row=3)
        return self._xlsx_response(wb, "participants_declic_emargement.xlsx")

    def _style_sheet(self, ws, header_row=1):
        fill = PatternFill("solid", fgColor="DCE6F1")
        border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="002060")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top")

        for col in ws.columns:
            if not col:
                continue
            letter = get_column_letter(col[0].column)
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[letter].width = min(max_len + 2, 40)

    def _xlsx_response(self, wb, filename):
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
