"""Statistiques agrégées sur le périmètre Déclic."""

from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.db.models import F, Sum, Value
from django.db.models.functions import Coalesce, Substr
from django.http import HttpResponse
from django.utils import timezone as dj_timezone
from django.utils.timezone import localdate
from drf_spectacular.utils import (
    OpenApiExample,
    OpenApiParameter,
    OpenApiResponse,
    extend_schema,
)
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from ....models.declic import Declic, ObjectifDeclic
from ...paginations import RapAppPagination
from ...permissions import IsDeclicStaffOrAbove
from ...roles import (
    build_department_scope_q,
    get_staff_centre_ids_cached,
    get_staff_department_codes_cached,
    is_admin_like,
    use_department_stats_scope,
)
from ...serializers.base_serializers import EmptySerializer


@extend_schema(tags=["Déclic - Statistiques"])
class DeclicStatsViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Reporting read-only sur les indicateurs Déclic.

    Les données sont restreintes par `_filtered_qs()` selon le rôle courant,
    puis exposées via des actions orientées dashboard :
    - `grouped` pour les agrégats par centre, département ou type ;
    - `synthese` pour les objectifs annuels ;
    - `resume` pour les KPI synthétiques ;
    - `export_xlsx` pour l'export tabulaire.

    Les endpoints renvoient soit des payloads JSON construits à la main, soit
    un fichier XLSX pour l'export.
    """

    serializer_class = EmptySerializer
    permission_classes = [IsDeclicStaffOrAbove]
    pagination_class = RapAppPagination
    queryset = Declic.objects.select_related("centre").all()

    # ----------------------------------------------------------
    # Filtre avec scope / année / centre / type
    # ----------------------------------------------------------
    def _filtered_qs(self, request):
        """
        Fonction utilitaire appliquant le filtrage business sur les objets Déclic,
        en fonction des droits d'accès de l'utilisateur et des paramètres de requête.

        - L'accès est plus étendu si l'utilisateur est "admin_like" (cf. roles.is_admin_like).
        - Pour les non-admin, scoping par centres de rattachement (soit via staff_centre_ids soit par département).
        - Si aucun centre/département ne peut être déterminé : retourne un queryset vide pour l'utilisateur courant.
        - Les filtres possibles par query_params : annee (par défaut année courante), centre, type_declic.
        """
        qs = self.queryset
        user = request.user

        annee = int(request.query_params.get("annee", localdate().year))
        centre_param = request.query_params.get("centre")
        type_declic = request.query_params.get("type_declic")
        department_scope = use_department_stats_scope(request)

        if is_admin_like(user):
            qs = qs.filter(date_declic__year=annee)
        else:
            centre_ids = list(get_staff_centre_ids_cached(request) or [])
            departements = set(get_staff_department_codes_cached(request) or [])

            if centre_param:
                qs = qs.filter(centre_id=centre_param)
            elif department_scope and departements:
                qs = qs.filter(build_department_scope_q("centre__code_postal", departements))
            elif centre_ids:
                qs = qs.filter(centre_id__in=centre_ids)
            elif departements:
                qs = qs.filter(build_department_scope_q("centre__code_postal", departements))
            else:
                # Aucun accès explicite autorisé pour cet utilisateur
                return Declic.objects.none()

            qs = qs.filter(date_declic__year=annee)

        if type_declic:
            qs = qs.filter(type_declic=type_declic)

        return qs

    # ----------------------------------------------------------
    # 1️⃣ GROUPED
    # ----------------------------------------------------------
    @action(detail=False, methods=["get"], url_path="grouped")
    def grouped(self, request):
        """
        Endpoint métier : /api/declic/stats/grouped
        Objectif :
            - Fournir des agrégats statistiques sur les séances Déclic groupées par :
              - centre (par défaut), OU département, OU type_declic
        Permissions :
            - Dépendent de IsDeclicStaffOrAbove + business _filtered_qs (voir docstring de la classe)
            - Prend en compte les restrictions d'accès utilisateur selon centre/département
        Paramètre GET attendu :
            - by = [centre|departement|type_declic] (par défaut centre)
        Format de la réponse JSON :
            - cf. docstring de la classe. Structure explicite dans le code ci-dessous.
        """
        by = request.query_params.get("by", "centre")
        qs = self._filtered_qs(request)

        if by == "centre":
            qs = qs.annotate(
                group_key=F("centre__nom"),
                centre_id_ref=F("centre__id"),
            )
            group_fields = ["centre_id_ref", "group_key"]

        elif by == "departement":
            qs = qs.annotate(group_key=Substr(Coalesce("centre__code_postal", Value("")), 1, 2))
            group_fields = ["group_key"]

        elif by == "type_declic":
            qs = qs.annotate(group_key=F("type_declic"))
            group_fields = ["group_key"]

        else:
            return Response({"success": False, "message": "Paramètre 'by' invalide", "data": None}, status=400)

        data = (
            qs.values(*group_fields)
            .annotate(
                nb_inscrits_declic=Sum("nb_inscrits_declic"),
                nb_presents_declic=Sum("nb_presents_declic"),
                nb_absents_declic=Sum("nb_absents_declic"),
            )
            .order_by("group_key")
        )

        results = []
        for d in data:
            insc = d["nb_inscrits_declic"] or 0
            pres = d["nb_presents_declic"] or 0
            absn = d["nb_absents_declic"] or 0

            taux_presence = round(pres / (pres + absn) * 100, 1) if pres + absn > 0 else None

            taux_retention = round(pres / insc * 100, 1) if insc > 0 else None

            results.append(
                {
                    "id": d.get("centre_id_ref"),
                    "group_key": d["group_key"],
                    "nb_inscrits_declic": insc,
                    "nb_presents_declic": pres,
                    "nb_absents_declic": absn,
                    "taux_presence_declic": taux_presence,
                    "taux_retention": taux_retention,
                }
            )

        return Response({"by": by, "results": results})

    # ----------------------------------------------------------
    # 2️⃣ SYNTHÈSE ANNUELLE
    # ----------------------------------------------------------
    @action(detail=False, methods=["get"], url_path="synthese")
    def synthese(self, request):
        """
        Endpoint métier : /api/declic/stats/synthese
        Objectif :
            - Retourner la synthèse annuelle pour Déclic (atteinte des objectifs, récapitulatif).
        Permissions :
            - Dépendent de IsDeclicStaffOrAbove + _filtered_qs (cf. docstring de la classe)
        Paramètres GET pris en compte :
            - annee = <int>, défaut = année courante
        Structure de la réponse :
            - Non documentée depuis ce code — contient ce que retourne Declic.synthese_objectifs(annee).
        """
        annee = int(request.query_params.get("annee", localdate().year))
        return Response(Declic.synthese_objectifs(annee))

    # ----------------------------------------------------------
    # 3️⃣ RESUME (dashboard)
    # ----------------------------------------------------------
    @action(detail=False, methods=["get"], url_path="resume")
    def resume(self, request):
        """
        Endpoint métier : /api/declic/stats/resume
        Objectif :
            - Dashboard global (pour un admin ou staff centré) : fournit un résumé agrégé,
              incluant le nombre d'inscrits, présents, absents, taux de présence/rétention,
              objectifs de l'année sélectionnée et reste à faire.
        Permissions :
            - Dépendent de IsDeclicStaffOrAbove + filtrage _filtered_qs (cf. docstring de la classe)
        Paramètres GET utilisables :
            - annee = <int>, défaut année courante
            - centre = <id centre> (optionnel)
        Structure de la réponse JSON :
            - Voir ci-dessous, contrat explicitement visible.
        """
        qs = self._filtered_qs(request)

        agg = qs.aggregate(
            inscrits=Sum("nb_inscrits_declic"),
            presents=Sum("nb_presents_declic"),
            absents=Sum("nb_absents_declic"),
        )

        inscrits = agg["inscrits"] or 0
        pres = agg["presents"] or 0
        absn = agg["absents"] or 0

        taux_presence = round(pres / (pres + absn) * 100, 1) if pres + absn > 0 else 0

        taux_retention = round(pres / inscrits * 100, 1) if inscrits > 0 else 0

        # Objectifs →
        annee = int(request.query_params.get("annee", localdate().year))
        centre_param = request.query_params.get("centre")
        department_scope = use_department_stats_scope(request)

        if is_admin_like(request.user):
            objectifs = ObjectifDeclic.objects.filter(annee=annee)
        else:
            centre_ids = list(get_staff_centre_ids_cached(request) or [])
            departements = set(get_staff_department_codes_cached(request) or [])
            objectifs = ObjectifDeclic.objects.filter(annee=annee)
            if department_scope and departements:
                objectifs = objectifs.filter(build_department_scope_q("centre__code_postal", departements))
            else:
                objectifs = objectifs.filter(centre_id__in=centre_ids)

        if centre_param:
            objectifs = objectifs.filter(centre_id=centre_param)

        objectif = objectifs.aggregate(total=Sum("valeur_objectif"))["total"] or 0

        realise_total = pres
        reste = objectif - realise_total
        taux_atteinte = round((realise_total / objectif) * 100, 1) if objectif else 0

        return Response(
            {
                "annee": annee,
                "objectif_total": objectif,
                "realise_total": realise_total,
                "taux_atteinte_total": taux_atteinte,
                "reste_a_faire_total": reste,
                "inscrits_total": inscrits,
                "taux_presence_declic": taux_presence,
                "taux_retention": taux_retention,
            }
        )

    # ==========================================================
    # 📤 4️⃣ Export Excel unifié
    # ==========================================================
    @extend_schema(
        summary="Export Excel Déclic",
        description="Génère un export XLSX complet regroupant les séances Déclic et les indicateurs clés.",
        responses={200: OpenApiResponse(description="Fichier Excel généré avec succès")},
    )
    @action(detail=False, methods=["get"], url_path="export-xlsx")
    def export_xlsx(self, request):
        """
        Endpoint métier : /api/declic/stats/export-xlsx
        Objectif :
            - Générer un export Excel (XLSX) complet des participations Déclic filtrées pour un utilisateur/annee/droits déterminés.
            - Ajoute logo, en-tête, lignes de données avec calculs de taux de présence/rétention.
        Permissions :
            - Dépendent de IsDeclicStaffOrAbove et filtrage _filtered_qs.
        Paramètres GET utilisables :
            - annee = <int>, centre = <id centre>, type_declic = <valeur>
        Structure de la réponse :
            - Fichier binaire XLSX. Ce endpoint ne retourne PAS de JSON.
            - Entêtes HTTP : Content-Disposition: 'attachment; filename=...xlsx'
        """
        qs = self._filtered_qs(request)
        annee = int(request.query_params.get("annee", localdate().year))

        wb = Workbook()
        ws = wb.active
        ws.title = f"Déclic {annee}"

        # ----------------------------------------------------------
        # Logo et titres
        # ----------------------------------------------------------
        try:
            logo_path = Path(settings.BASE_DIR) / "rap_app/static/images/logo.png"
            if logo_path.exists():
                img = XLImage(str(logo_path))
                img.height = 60
                img.width = 120
                ws.add_image(img, "A1")
        except Exception:
            pass

        ws.merge_cells("B1:H1")
        ws["B1"] = f"Export Déclic {annee} — RAP_APP"
        ws["B1"].font = Font(name="Calibri", bold=True, size=15, color="004C99")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("B2:H2")
        ws["B2"] = f"Généré le {dj_timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        ws["B2"].font = Font(name="Calibri", italic=True, size=10, color="666666")
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.append([])
        ws.append([])

        # ----------------------------------------------------------
        # En-têtes
        # ----------------------------------------------------------
        headers = [
            "Centre",
            "Type Déclic",
            "Date",
            "Inscrits (Atelier)",
            "Présents (Atelier)",
            "Absents (Atelier)",
            "Taux Présence Atelier %",
            "Reste à faire",
            "Taux rétention (%)",
        ]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="DCE6F1")
        border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        for cell in ws[ws.max_row]:
            cell.font = Font(name="Calibri", bold=True, color="002060")
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ----------------------------------------------------------
        # Données
        # ----------------------------------------------------------
        even_fill = PatternFill("solid", fgColor="F8FBFF")
        odd_fill = PatternFill("solid", fgColor="FFFFFF")

        for i, s in enumerate(qs, start=1):

            # 🧮 Calcul du taux de rétention (Atelier 1 → Atelier 6)
            taux_retention = (
                round(s.nb_presents_declic / s.nb_inscrits_declic * 100, 1) if s.nb_inscrits_declic > 0 else None
            )

            ws.append(
                [
                    getattr(s.centre, "nom", ""),
                    s.get_type_declic_display(),
                    s.date_declic.strftime("%d/%m/%Y") if s.date_declic else "",
                    s.nb_inscrits_declic,
                    s.nb_presents_declic,
                    s.nb_absents_declic,
                    s.taux_presence_declic,
                    s.reste_a_faire,
                    taux_retention,  # ✅ ici c’est la vraie valeur calculée
                ]
            )

            fill = even_fill if i % 2 == 0 else odd_fill
            for cell in ws[ws.max_row]:
                cell.fill = fill
                cell.border = border
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(vertical="center")

        # ----------------------------------------------------------
        # Auto-filter + ajustement colonnes
        # ----------------------------------------------------------
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        ws.freeze_panes = "A2"
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 35)

        # ----------------------------------------------------------
        # Fichier final
        # ----------------------------------------------------------
        buffer = BytesIO()
        wb.save(buffer)
        content = buffer.getvalue()
        filename = f"declic_stats_{annee}_{dj_timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        response = HttpResponse(
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = len(content)
        return response
