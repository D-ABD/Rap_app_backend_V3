# rap_app/api/viewsets/prepa_stats_viewset.py

from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.db.models import F, Sum, Value
from django.db.models.functions import Coalesce, Substr
from django.http import HttpResponse
from django.utils import timezone as dj_timezone
from django.utils.timezone import localdate
from drf_spectacular.utils import (
    OpenApiExample,
    OpenApiParameter,
    OpenApiResponse,
    extend_schema,
)
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from ....models.prepa import Prepa
from ...paginations import RapAppPagination
from ...permissions import IsPrepaStaffOrAbove
from ...serializers.base_serializers import EmptySerializer


# ==========================================================
#  📊 VIEWSET — PREPA STATS
# ==========================================================
@extend_schema(tags=["Prépa - Statistiques"])
class PrepaStatsViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Reporting read-only sur le module Prépa.

    Toutes les actions reposent sur `_filtered_qs()` pour appliquer le
    périmètre Prépa staff/admin et les filtres métier (`centre`,
    `departement`, `type_prepa`, `annee`).

    Le viewset expose surtout des endpoints dashboard et d'export ; les
    réponses sont soit des payloads JSON construits à la main, soit des
    exports binaires.
    """

    serializer_class = EmptySerializer
    permission_classes = [IsPrepaStaffOrAbove]
    pagination_class = RapAppPagination
    queryset = Prepa.objects.select_related("centre").all()

    # ==========================================================
    # 🧩 0️⃣ Options de filtres — utilisé par le frontend
    # ==========================================================
    @extend_schema(
        summary="Options de filtres Prepa (stats)",
        description="Retourne les années, centres, départements, types Prepa accessibles.",
        responses={200: dict},
    )
    @action(detail=False, methods=["get"], url_path="filters")
    def filters(self, request):
        """
        Point d’entrée API : GET /api/prepa-stats/filters/

        ──────────────────────────────────────────────────────────────
        OBJECTIF MÉTIER :
        ──────────────────────────────────────────────────────────────
        - Permet à l’UI de proposer des filtres dynamiques pour les statistiques
        - Fournit uniquement les années (données présentes), les centres et départements effectivement accessibles
          pour l’utilisateur courant, et tous les types de Prépa possibles

        ──────────────────────────────────────────────────────────────
        PERMISSIONS :
        ──────────────────────────────────────────────────────────────
        - Identiques au reste du ViewSet (voir classe): seuls utilisateurs “Prepa Staff” ou “admin” selon `IsPrepaStaffOrAbove`

        ──────────────────────────────────────────────────────────────
        STRUCTURE DE RÉPONSE (exemple observable dans le code — pas de schéma universel DRF) :
        ──────────────────────────────────────────────────────────────
        {
            "annees": [2023,2022,...],
            "centres": [
                {"value": 4, "label": "Nom du centre", "code_postal": "59000", "departement": "59"}
            ],
            "departements": [
                {"value": "59", "label": "Département 59"}
            ],
            "type_prepa": [
                {"value": "info_collective", "label": "Information collective"}
            ]
        }
        """
        # Années (uniques)
        annees = Prepa.objects.order_by().values_list("date_prepa__year", flat=True).distinct()
        annees = sorted([a for a in annees if a], reverse=True)

        # Centres visibles selon le scope utilisateur
        centres_qs = self._filtered_qs(request).values("centre_id", "centre__nom", "centre__code_postal").distinct()

        centres = []
        departements = set()

        for c in centres_qs:
            cp = c.get("centre__code_postal") or ""
            dep = cp[:2] if len(cp) >= 2 else None
            if dep:
                departements.add(dep)

            centres.append(
                {
                    "value": c["centre_id"],
                    "label": c["centre__nom"],
                    "code_postal": cp,
                    "departement": dep,
                }
            )

        # Départements
        departements_list = [{"value": d, "label": f"Département {d}"} for d in sorted(departements)]

        # Types Prépa
        types = [{"value": t[0], "label": t[1]} for t in Prepa.TypePrepa.choices]

        return Response(
            {
                "annees": annees,
                "centres": centres,
                "departements": departements_list,
                "type_prepa": types,
            }
        )

    # ==========================================================
    # 🔍 Helper — filtrage contextuel avec scope par rôle
    # ==========================================================
    def _filtered_qs(self, request):
        """
        Helper privé.
        Retourne un queryset Prépa filtré selon :
            - les droits utilisateur :
                - admin/superadmin accède à toutes les lignes de l’année demandée
                - staff ou “prepa_staff” :
                    • accède à ses centres, ou à défaut aux départements associés à ses centres
                    • un filtre explicite “centre” prime sur la logique de scope staff
            - Les filtres GET : centre, type_prepa, année

        PERMISSION LOGIQUE : croisement du scope utilisateur + filtres explicites.


        ATTENTION : Cette méthode ne remplace pas get_queryset standard mais est appelée explicitement
        par les actions custom.
        """
        qs = self.queryset
        user = request.user

        from ...roles import is_admin_like  # Le détail de cette fonction n’est pas ici

        # 🔹 Paramètres communs
        annee = int(request.query_params.get("annee", localdate().year))
        centre_param = request.query_params.get("centre")
        type_prepa = request.query_params.get("type_prepa")

        # ------------------------------------------------------
        # 🔹 1️⃣ ADMIN / SUPERADMIN → accès complet
        # ------------------------------------------------------
        if is_admin_like(user):
            qs = qs.filter(date_prepa__year=annee)

        # ------------------------------------------------------
        # 🔹 2️⃣ STAFF / STAFF_READ / PREPA_STAFF → scope restreint
        # ------------------------------------------------------
        else:
            # --- 1) Centres autorisés (via attribut custom du user : staff_centre_ids)
            centre_ids = set(getattr(user, "staff_centre_ids", []) or [])

            # Ajout des centres M2M si existants (user.centre m2m)
            if hasattr(user, "centres") and user.centres.exists():
                for c in user.centres.all():
                    centre_ids.add(c.id)

            # --- 2) Si aucun centre explicite → fallback départements associés
            departements = set()
            if not centre_ids and hasattr(user, "centres") and user.centres.exists():
                for c in user.centres.all():
                    cp = getattr(c, "code_postal", "") or ""
                    if len(cp) >= 2:
                        departements.add(cp[:2])

            # --- 3) Filtre “centre” explicite (query param) → prioritaire
            if centre_param:
                try:
                    centre_id = int(str(centre_param).strip())
                    qs = qs.filter(centre_id=centre_id)
                except ValueError:
                    qs = qs.filter(centre__nom__iexact=str(centre_param).strip())

            # --- 4) Accès centres autorisés du staff
            elif centre_ids:
                qs = qs.filter(centre_id__in=centre_ids)

            # --- 5) Accès fallback : département
            elif departements:
                qs = qs.filter(centre__code_postal__startswith=tuple(departements))

            # --- 6) Aucun centre autorisé : accès vide
            else:
                qs = qs.none()

            qs = qs.filter(date_prepa__year=annee)

        # ------------------------------------------------------
        # 🔹 3️⃣ Filtre optionnel : type_prepa
        # ------------------------------------------------------
        if type_prepa:
            qs = qs.filter(type_prepa=type_prepa)

        return qs

    # ==========================================================
    # 📊 1️⃣ Regroupement détaillé
    # ==========================================================
    @extend_schema(
        summary="Statistiques groupées Prépa",
        description="Retourne les totaux et taux par centre, département (via code postal) ou type d’activité.",
        parameters=[
            OpenApiParameter(
                name="by",
                description="Type de regroupement (centre | departement | type_prepa)",
                required=True,
                type=str,
            ),
            OpenApiParameter(
                name="annee",
                description="Filtrer par année (ex: 2025)",
                required=False,
                type=int,
            ),
            OpenApiParameter(
                name="centre",
                description="Filtrer par identifiant de centre",
                required=False,
                type=str,
            ),
            OpenApiParameter(
                name="type_prepa",
                description="Filtrer par type d’activité (info_collective, atelier_1, ...)",
                required=False,
                type=str,
            ),
        ],
        responses={
            200: OpenApiResponse(
                description="Regroupement statistique des données Prépa.",
                examples=[
                    OpenApiExample(
                        name="Exemple regroupement",
                        value={
                            "by": "centre",
                            "results": [
                                {
                                    "id": 2,
                                    "group_key": "Centre Lille",
                                    "total": 220,
                                    "nb_presents_info": 120,
                                    "nb_absents_info": 30,
                                    "nb_adhesions": 40,
                                    "nb_inscrits_prepa": 90,
                                    "nb_presents_prepa": 80,
                                    "nb_absents_prepa": 10,
                                    "taux_presence_info": 80.0,
                                    "taux_adhesion": 33.3,
                                    "taux_presence_prepa": 88.9,
                                }
                            ],
                        },
                        response_only=True,
                    )
                ],
            )
        },
    )
    @action(detail=False, methods=["get"], url_path="grouped")
    def grouped(self, request):
        """
        Point d’entrée API : GET /api/prepa-stats/grouped/

        ──────────────────────────────────────────────────────────────
        OBJECTIF MÉTIER :
        ──────────────────────────────────────────────────────────────
        - Fournit des agrégations statistiques groupées
          par centre, département ou type de prépa pour l’UI ou un export/reporting
        - Repose sur le scope dynamique (_filtered_qs) → filtre auto selon droits utilisateur

        ──────────────────────────────────────────────────────────────
        PARAMS :
        - by (obligatoire) : centre, departement, type_prepa
        - année, centre, type_prepa : optionnels (filtrage)
        ──────────────────────────────────────────────────────────────

        ──────────────────────────────────────────────────────────────
        RÉPONSE JSON :
        ──────────────────────────────────────────────────────────────
        • List d’objets contenant pour chaque groupe :
            - group_key : nom du centre, numéro de département, ou valeur type_prepa, selon “by”
            - total, nb_presents_info, nb_absents_info, nb_adhesions...
            - taux_presence_info, taux_adhesion, taux_presence_prepa, taux_retention (calculés)
        - champ “by” rappelle le mode de regroupement.

        Voir l’exemple explicite dans @extend_schema → pas de serializer classique.
        """
        by = request.query_params.get("by", "centre")
        qs = self._filtered_qs(request)

        # 🧩 Définition du champ de regroupement
        if by == "centre":
            qs = qs.annotate(
                group_key=F("centre__nom"),
                centre_id_ref=F("centre__id"),
            )
            group_fields = ["centre_id_ref", "group_key"]

        elif by == "departement":
            qs = qs.annotate(group_key=Substr(Coalesce("centre__code_postal", Value("")), 1, 2))
            group_fields = ["group_key"]

        elif by == "type_prepa":
            qs = qs.annotate(group_key=F("type_prepa"))
            group_fields = ["group_key"]

        else:
            return Response(
                {"detail": "Paramètre 'by' invalide. Valeurs possibles : centre, departement, type_prepa."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # 🧮 Agrégations
        data = (
            qs.values(*group_fields)
            .annotate(
                nb_presents_info=Sum("nb_presents_info"),
                nb_absents_info=Sum("nb_absents_info"),
                nb_adhesions=Sum("nb_adhesions"),
                nb_inscrits_prepa=Sum("nb_inscrits_prepa"),
                nb_presents_prepa=Sum("nb_presents_prepa"),
                nb_absents_prepa=Sum("nb_absents_prepa"),
            )
            .order_by("group_key")
        )

        results = []
        for d in data:
            key = d.get("group_key") or "—"
            centre_id = d.get("centre_id_ref")

            p_info = d["nb_presents_info"] or 0
            a_info = d["nb_absents_info"] or 0
            adh = d["nb_adhesions"] or 0
            insc = d["nb_inscrits_prepa"] or 0
            p_prepa = d["nb_presents_prepa"] or 0
            a_prepa = d["nb_absents_prepa"] or 0

            taux_presence_info = round(p_info / (p_info + a_info) * 100, 1) if (p_info + a_info) > 0 else None
            taux_adhesion = round(adh / p_info * 100, 1) if p_info > 0 else None
            taux_presence_prepa = round(p_prepa / (p_prepa + a_prepa) * 100, 1) if (p_prepa + a_prepa) > 0 else None

            total = p_prepa

            taux_retention = round((p_prepa - a_prepa) / p_prepa * 100, 1) if p_prepa > 0 else None

            results.append(
                {
                    "id": centre_id,
                    "group_key": key,
                    "total": total,
                    "nb_presents_info": p_info,
                    "nb_absents_info": a_info,
                    "nb_adhesions": adh,
                    "nb_inscrits_prepa": insc,
                    "nb_presents_prepa": p_prepa,
                    "nb_absents_prepa": a_prepa,
                    "taux_presence_info": taux_presence_info,
                    "taux_adhesion": taux_adhesion,
                    "taux_presence_prepa": taux_presence_prepa,
                    "taux_retention": taux_retention,
                }
            )

        return Response({"by": by, "results": results})

    # ==========================================================
    # 📈 2️⃣ Synthèse globale annuelle
    # ==========================================================
    @extend_schema(
        summary="Synthèse annuelle Prépa",
        description="Retourne les totaux et taux d’atteinte globaux de l’année.",
    )
    @action(detail=False, methods=["get"], url_path="synthese")
    def synthese(self, request):
        """
        Point d’entrée API : GET /api/prepa-stats/synthese/

        ──────────────────────────────────────────────────────────────
        OBJECTIF MÉTIER :
        ──────────────────────────────────────────────────────────────
        - Fournit la synthèse de toute l’année souhaitée : objectifs globaux, réalisé, taux...

        ──────────────────────────────────────────────────────────────
        PERMISSION :
        ──────────────────────────────────────────────────────────────
        - Identique au reste du ViewSet : staff Prépa ou admin

        ──────────────────────────────────────────────────────────────
        FILTRAGE :
        - Année, selon paramètre GET

        ──────────────────────────────────────────────────────────────
        RÉPONSE JSON :
        - Directement ce que retourne Prepa.synthese_objectifs(annee) (le détail du format dépend du modèle :
          n’est PAS documenté ici car dépend de la méthode de modèle NON visible dans ce fichier – ⚠️ Contrat de structure inconnu sans explorer le modèle)

        """
        annee = int(request.query_params.get("annee", localdate().year))
        data = Prepa.synthese_objectifs(annee)
        return Response(data)

    # ==========================================================
    # 🧾 3️⃣ Résumé rapide (dashboard)
    # ==========================================================
    @action(detail=False, methods=["get"], url_path="resume")
    def resume(self, request):
        """
        Point d’entrée API : GET /api/prepa-stats/resume/

        ──────────────────────────────────────────────────────────────
        OBJECTIF MÉTIER :
        ──────────────────────────────────────────────────────────────
        - Fournit en un seul appel les principales statistiques de l’année, liées à l’utilisateur courant :
            objectifs réalisés, taux globaux, taux de présence IC/atelier, prescriptions
        - Compatible avec filtres centre/département

        ──────────────────────────────────────────────────────────────
        PERMISSION :
        ──────────────────────────────────────────────────────────────
        - Identique au reste du ViewSet : staff Prépa ou admin

        ──────────────────────────────────────────────────────────────
        FILTRAGE :
        - Année (GET)
        - centre (GET)
        - departement (GET)

        ──────────────────────────────────────────────────────────────
        RÉPONSE JSON :
        ──────────────────────────────────────────────────────────────
        - Structure observable :
            {
                "annee": ...,
                "objectif_total": ...,
                "realise_total": ...,
                "taux_atteinte_total": ...,
                "reste_a_faire_total": ...,
                "nb_prescriptions": ...,
                "taux_prescription": ...,
                "presents_info": ...,
                "absents_info": ...,
                "taux_presence_ic": ...,
                "presents_ateliers": ...,
                "absents_ateliers": ...,
                "taux_presence_ateliers": ...,
                "par_centre": [
                    {"centre_id": ..., "centre__nom": "...", "total": ...}
                ],
                "par_departement": [
                    {"departement": "...", "total": ...}
                ]
            }
        - Toute la structure est codée ici explicitement ; si le calcul doit évoluer il faudra
          maintenir la description.
        """
        from django.db.models import Sum, Value
        from django.db.models.functions import Coalesce, Substr
        from django.utils.timezone import localdate

        from ....models.prepa import ObjectifPrepa, Prepa
        from ...roles import is_admin_like

        user = request.user
        annee = int(request.query_params.get("annee", localdate().year))
        centre_param = request.query_params.get("centre")
        departement_param = request.query_params.get("departement")

        # ------------------------------------------------------
        # 🔍 1) Scope complet via _filtered_qs
        # ------------------------------------------------------
        qs = self._filtered_qs(request).filter(date_prepa__year=annee)

        if centre_param:
            try:
                centre_id = int(str(centre_param).strip())
                qs = qs.filter(centre_id=centre_id)
            except ValueError:
                qs = qs.filter(centre__nom__icontains=str(centre_param).strip())

        if departement_param:
            qs = qs.filter(centre__code_postal__startswith=str(departement_param).strip())

        # ------------------------------------------------------
        # 🎯 2) Objectifs accessibles selon le scope
        # ------------------------------------------------------
        if is_admin_like(user):
            objectifs_qs = ObjectifPrepa.objects.filter(annee=annee)
        else:
            centre_ids = list(getattr(user, "staff_centre_ids", []) or [])

            # Ajout centres M2M
            if hasattr(user, "centres") and user.centres.exists():
                for c in user.centres.all():
                    centre_ids.append(c.id)

            departements = set()
            if hasattr(user, "centres") and user.centres.exists():
                for c in user.centres.all():
                    cp = getattr(c, "code_postal", "") or ""
                    if len(cp) >= 2:
                        departements.add(cp[:2])

            objectifs_qs = ObjectifPrepa.objects.filter(annee=annee)
            if centre_ids:
                objectifs_qs = objectifs_qs.filter(centre_id__in=centre_ids)
            elif departements:
                objectifs_qs = objectifs_qs.filter(centre__code_postal__startswith=tuple(departements))
            else:
                objectifs_qs = ObjectifPrepa.objects.none()

        if centre_param:
            try:
                centre_id = int(str(centre_param).strip())
                objectifs_qs = objectifs_qs.filter(centre_id=centre_id)
            except ValueError:
                objectifs_qs = objectifs_qs.filter(centre__nom__icontains=str(centre_param).strip())

        elif departement_param:
            objectifs_qs = objectifs_qs.filter(centre__code_postal__startswith=str(departement_param).strip())

        # ------------------------------------------------------
        # 📊 3) Totaux et Taux Globaux
        # ------------------------------------------------------
        objectif_total = objectifs_qs.aggregate(total=Sum("valeur_objectif"))["total"] or 0
        realise_total = qs.aggregate(total=Sum("nb_presents_prepa"))["total"] or 0

        reste_a_faire_total = objectif_total - realise_total
        taux_atteinte_total = round((realise_total / objectif_total) * 100, 1) if objectif_total > 0 else None

        # ----------------------------------------------
        # 🟦  PRESCRIPTIONS (IC)
        # ----------------------------------------------
        nb_prescriptions = qs.aggregate(total=Sum("nombre_prescriptions"))["total"] or 0
        places_ouvertes = qs.aggregate(total=Sum("nombre_places_ouvertes"))["total"] or 0

        taux_prescription = round(nb_prescriptions / places_ouvertes * 100, 1) if places_ouvertes > 0 else None

        # ----------------------------------------------
        # 🟩  PRÉSENCE INFORMATION COLLECTIVE
        # ----------------------------------------------
        presents_info = qs.aggregate(total=Sum("nb_presents_info"))["total"] or 0
        absents_info = qs.aggregate(total=Sum("nb_absents_info"))["total"] or 0

        taux_presence_ic = (
            round(presents_info / (presents_info + absents_info) * 100, 1)
            if (presents_info + absents_info) > 0
            else None
        )

        # ----------------------------------------------
        # 🟪  PRÉSENCE ATELIERS PRÉPA
        # ----------------------------------------------
        presents_ateliers = qs.aggregate(total=Sum("nb_presents_prepa"))["total"] or 0
        absents_ateliers = qs.aggregate(total=Sum("nb_absents_prepa"))["total"] or 0

        taux_presence_ateliers = (
            round(presents_ateliers / (presents_ateliers + absents_ateliers) * 100, 1)
            if (presents_ateliers + absents_ateliers) > 0
            else None
        )

        # ------------------------------------------------------
        # 📌 4) Détail par centre
        # ------------------------------------------------------
        par_centre_qs = (
            qs.values("centre__id", "centre__nom").annotate(total=Sum("nb_presents_prepa")).order_by("centre__nom")
        )
        par_centre = [
            {
                "centre_id": r["centre__id"],
                "centre__nom": r["centre__nom"] or "—",
                "total": r["total"] or 0,
            }
            for r in par_centre_qs
        ]

        # ------------------------------------------------------
        # 📌 5) Détail par département
        # ------------------------------------------------------
        par_departement_qs = (
            qs.annotate(departement=Substr(Coalesce("centre__code_postal", Value("")), 1, 2))
            .values("departement")
            .annotate(total=Sum("nb_presents_prepa"))
            .order_by("departement")
        )
        par_departement = [
            {"departement": r["departement"] or "—", "total": r["total"] or 0} for r in par_departement_qs
        ]

        # ------------------------------------------------------
        # 🟧 6) RÉPONSE COMPLÈTE
        # ------------------------------------------------------
        return Response(
            {
                "annee": annee,
                "objectif_total": objectif_total,
                "realise_total": realise_total,
                "taux_atteinte_total": taux_atteinte_total,
                "reste_a_faire_total": reste_a_faire_total,
                # ---- PRESCRIPTIONS ----
                "nb_prescriptions": nb_prescriptions,
                "taux_prescription": taux_prescription,
                # ---- INFO COLLECTIVE ----
                "presents_info": presents_info,
                "absents_info": absents_info,
                "taux_presence_ic": taux_presence_ic,
                # ---- ATELIERS ----
                "presents_ateliers": presents_ateliers,
                "absents_ateliers": absents_ateliers,
                "taux_presence_ateliers": taux_presence_ateliers,
                # ---- GROUPES ----
                "par_centre": par_centre,
                "par_departement": par_departement,
            }
        )

    # ==========================================================
    # 📤 4️⃣ Export Excel unifié
    # ==========================================================
    @extend_schema(
        summary="Export Excel Prépa",
        description="Génère un export XLSX complet regroupant les séances Prépa et les indicateurs clés.",
        responses={200: OpenApiResponse(description="Fichier Excel généré avec succès")},
    )
    @action(detail=False, methods=["get"], url_path="export-xlsx")
    def export_xlsx(self, request):
        """
        Point d’entrée API : GET /api/prepa-stats/export-xlsx/

        ──────────────────────────────────────────────────────────────
        OBJECTIF MÉTIER :
        ──────────────────────────────────────────────────────────────
        - Permet aux utilisateurs autorisés d’exporter sous Excel la liste détaillée des séances Prépa
          et tous les indicateurs principaux, pour l’année/filtres voulus

        ──────────────────────────────────────────────────────────────
        PERMISSIONS :
        ──────────────────────────────────────────────────────────────
        - Identique au reste du ViewSet : staff ou admin selon `IsPrepaStaffOrAbove`

        ──────────────────────────────────────────────────────────────
        FILTRAGE :
        - Année (GET)

        ──────────────────────────────────────────────────────────────
        RÉPONSE :
        - Fichier .xlsx généré dynamiquement en output HTTP (voir Content-Type)
        - Le format exact du tableur est observable dans le code, ligne par ligne
        - Pas de contrat de schéma DRF en JSON : pure réponse binaire, à interpréter comme un fichier Excel

        ATTENTION :
        - Si vous ajoutez des colonnes, il faut aussi ajuster la structure d’entête ici
        - Ce endpoint ne retourne JAMAIS de JSON (ni erreur structurée DRF), mais un fichier à télécharger
        """
        qs = self._filtered_qs(request)
        annee = int(request.query_params.get("annee", localdate().year))

        wb = Workbook()
        ws = wb.active
        ws.title = f"Prépa {annee}"

        try:
            logo_path = Path(settings.BASE_DIR) / "rap_app/static/images/logo.png"
            if logo_path.exists():
                img = XLImage(str(logo_path))
                img.height = 60
                img.width = 120
                ws.add_image(img, "A1")
        except Exception:
            pass

        ws.merge_cells("B1:H1")
        ws["B1"] = f"Export Prépa {annee} — RAP_APP"
        ws["B1"].font = Font(name="Calibri", bold=True, size=15, color="004C99")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("B2:H2")
        ws["B2"] = f"Généré le {dj_timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        ws["B2"].font = Font(name="Calibri", italic=True, size=10, color="666666")
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.append([])
        ws.append([])

        headers = [
            "Centre",
            "Type Prépa",
            "Date",
            "Présents (IC)",
            "Absents (IC)",
            "Adhésions",
            "Inscrits (Atelier)",
            "Présents (Atelier)",
            "Absents (Atelier)",
            "Taux Présence IC %",
            "Taux Adhésion %",
            "Taux Présence Atelier %",
            "Reste à faire",
            "Taux rétention (%)",
        ]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="DCE6F1")
        border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        for cell in ws[ws.max_row]:
            cell.font = Font(name="Calibri", bold=True, color="002060")
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        even_fill = PatternFill("solid", fgColor="F8FBFF")
        odd_fill = PatternFill("solid", fgColor="FFFFFF")

        for i, s in enumerate(qs, start=1):

            taux_retention = (
                round((s.nb_presents_prepa - s.nb_absents_prepa) / s.nb_presents_prepa * 100, 1)
                if s.nb_presents_prepa > 0
                else None
            )

            ws.append(
                [
                    getattr(s.centre, "nom", ""),
                    s.get_type_prepa_display(),
                    s.date_prepa.strftime("%d/%m/%Y") if s.date_prepa else "",
                    s.nb_presents_info,
                    s.nb_absents_info,
                    s.nb_adhesions,
                    s.nb_inscrits_prepa,
                    s.nb_presents_prepa,
                    s.nb_absents_prepa,
                    s.taux_presence_info,
                    s.taux_adhesion,
                    s.taux_presence_prepa,
                    s.reste_a_faire,
                    taux_retention,
                ]
            )

            fill = even_fill if i % 2 == 0 else odd_fill
            for cell in ws[ws.max_row]:
                cell.fill = fill
                cell.border = border
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(vertical="center")

        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        ws.freeze_panes = "A2"

        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 35)

        buffer = BytesIO()
        wb.save(buffer)
        content = buffer.getvalue()
        filename = f"prepa_stats_{annee}_{dj_timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        response = HttpResponse(
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = len(content)
        return response
