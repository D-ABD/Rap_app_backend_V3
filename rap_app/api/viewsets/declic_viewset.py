# rap_app/api/viewsets/declic_viewset.py

import logging
from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone as dj_timezone
from django.utils.timezone import localdate
from drf_spectacular.utils import OpenApiParameter, extend_schema, extend_schema_view
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.response import Response

from ...api.roles import (
    build_department_scope_q,
    get_staff_department_codes_cached,
    is_admin_like,
    is_candidate,
    is_declic_staff,
    is_staff_or_staffread,
    use_department_stats_scope,
)
from ...models.centres import Centre
from ...models.declic import Declic, ObjectifDeclic
from ..mixins import ApiResponseMixin, HardDeleteArchivedMixin
from ..permissions import IsDeclicStaffOrAbove
from ..serializers.declic_serializers import DeclicSerializer

logger = logging.getLogger(__name__)

# =====================================================================================
# 🔧 HELPERS SCOPE — (MANQUAIENT DANS TON FICHIER) → POUR L'AUTORISATION PAR CENTRES
# =====================================================================================


class ScopeMixin:
    """
    Mixin local de scoping par centre pour le module Déclic.

    Il conserve une implémentation historique autonome et ne remplace pas la
    stratégie plus récente portée ailleurs par `ScopedModelViewSet`.
    """

    def _centre_ids_for_user(self, user):
        """
        Retourne la liste des centres accessibles selon le rôle.

        - Si l'utilisateur est admin (cf. is_admin_like), retourne None (accès complet sans filtre).
        - Sinon, tente de récupérer les centres via user.centres_acces ou user.centres et retourne leurs ids.
        - Si cette relation n'existe pas, retourne liste vide (aucun accès).
        """
        if is_admin_like(user):
            return None  # accès complet

        # déclic_staff = accès restreint à ses centres
        centres = getattr(user, "centres_acces", None) or getattr(user, "centres", None)

        if centres is not None:
            return list(centres.values_list("id", flat=True))

        return []  # par défaut aucun centre

    def _scope_qs_to_user_centres(self, qs):
        """
        Filtre un queryset sur les centres autorisés pour l'utilisateur courant.

        - Si l'utilisateur n'est pas authentifié OU est candidat : aucun résultat.
        - Si l'utilisateur est admin-like (cf. _centre_ids_for_user), aucun filtre appliqué.
        - Sinon, filtre les objets rattachés aux centres autorisés pour l'utilisateur.
        - Si aucun centre autorisé : aucun résultat.
        """
        user = self.request.user

        if not user.is_authenticated or is_candidate(user):
            return qs.none()

        centre_ids = self._centre_ids_for_user(user)

        if centre_ids is None:  # superadmin/admin
            return qs

        if len(centre_ids) == 0:
            return qs.none()

        model = getattr(qs, "model", None)
        if model and getattr(getattr(model, "_meta", None), "model_name", None) == "centre":
            return qs.filter(id__in=centre_ids)

        return qs.filter(centre_id__in=centre_ids)


# =====================================================================================
# 📊 DÉCLIC VIEWSET — ATELIERS UNIQUEMENT
# =====================================================================================


@extend_schema_view(
    list=extend_schema(
        summary="Lister tous les ateliers Déclic",
        description="Retourne uniquement les ateliers Déclic (Atelier 1 → 6 + Autre).",
        parameters=[
            OpenApiParameter(name="annee", type=int),
            OpenApiParameter(name="centre", type=int),
            OpenApiParameter(name="departement", type=str),
            OpenApiParameter(name="type_declic", type=str),
            OpenApiParameter(name="date_min", type=str),
            OpenApiParameter(name="date_max", type=str),
            OpenApiParameter(name="search", type=str),
        ],
        responses={200: DeclicSerializer},
    ),
    destroy=extend_schema(
        summary="Archiver une séance Déclic",
        description="Archive logiquement une séance Déclic en la désactivant (`is_active = False`).",
        responses={200: DeclicSerializer},
    ),
)
class DeclicViewSet(HardDeleteArchivedMixin, ApiResponseMixin, ScopeMixin, viewsets.ModelViewSet):
    """
    ViewSet CRUD du module Déclic.

    Le fichier conserve une logique de scope locale via `ScopeMixin` et une
    documentation historiquement très détaillée. La source de vérité actuelle
    est plus simple :
    - permission principale : `IsDeclicStaffOrAbove`
    - visibilité filtrée par centres accessibles, sauf profils admin-like
    - filtres manuels dans `get_queryset()`
    - actions custom de filtres, statistiques et export XLSX
    """

    serializer_class = DeclicSerializer
    permission_classes = [IsDeclicStaffOrAbove]
    hard_delete_enabled = True
    queryset = Declic.objects.select_related("centre").prefetch_related("participants_declic").all()

    # -------------------------------------------------------------------------
    # 🎛️ OPTIONS FILTRES
    # -------------------------------------------------------------------------
    @action(detail=False, methods=["get"], url_path="filters")
    def filters(self, request):
        """
        Point d'entrée : GET /declic/filters/

        Objectif métier : Fournir la liste ordonnée :
            - des années rencontrées dans date_declic des ateliers
            - des centres accessibles par l'utilisateur, chacun enrichi du code département et code postal
            - des départements rencontrés, pour menu déroulant
            - des types d'ateliers possibles, pour affichage

        Permissions & visibilité : cette action applique la restriction par centre via ScopeMixin, donc
        - un admin-like voit la liste complète ;
        - un staff ou assimilé ne voit que ses centres/autorisations ;
        - un "candidat" ne verra rien.

        Réponse : JSON contenant "annees", "departements", "centres", "type_declic"
        (voir exemple dans docstring de la classe).
        """
        annees = self._scope_qs_to_user_centres(Declic.objects.all()).order_by().values_list("date_declic__year", flat=True).distinct()
        annees = sorted([a for a in annees if a], reverse=True)

        centres_qs = self._scope_qs_to_user_centres(Centre.objects.all())

        centres = []
        deps = set()

        for c in centres_qs.order_by("nom"):
            # Le code département est soit le champ 'departement', soit les 2 premiers chiffres du code postal.
            dep = getattr(c, "departement", "") or (c.code_postal[:2] if c.code_postal else "")
            if dep:
                deps.add(dep)

            centres.append(
                {
                    "value": c.id,
                    "label": c.nom,
                    "departement": dep,
                    "code_postal": c.code_postal,
                }
            )

        # Liste des types atelier uniquement (tuple formaté value/label)
        types = [{"value": t[0], "label": t[1]} for t in Declic.TypeDeclic.choices]

        return self.success_response(
            data={
                "annees": annees,
                "departements": [{"value": d, "label": f"Département {d}"} for d in sorted(deps)],
                "centres": centres,
                "type_declic": types,
            },
            message="Filtres Déclic récupérés avec succès.",
        )

    # -------------------------------------------------------------------------
    # 🔍 QUERYSET PRINCIPAL
    # -------------------------------------------------------------------------
    def get_queryset(self):
        """
        Génère le queryset principal des séances Déclic visibles pour
        toutes les actions standard et d'export Excel/statistiques.

        Filtrage combiné :
        - Restriction par centres autorisés via ScopeMixin._scope_qs_to_user_centres, selon le rôle utilisateur
          (cf. docstring de la classe, dépendance à is_admin_like/is_candidate/custom logic).
        - Filtrages "manuels" sur paramètres query_string :
            * annee: filtrage par année (date_declic__year)
            * centre: filtrage par identifiant exact de centre
            * departement: filtre sur 'centre__departement' OU 'centre__code_postal', si commence par la valeur
            * type_declic: filtrage type atelier
            * date_min/date_max: inclusives sur date_declic
            * search: recherche sur centre__nom ou commentaire
            * ordering: tri principal (défaut: -date_declic, puis -id).

        Retourne uniquement les objets autorisés et filtrés.
        """
        qs = Declic.objects.select_related("centre").prefetch_related("participants_declic")
        qs = self._scope_qs_to_user_centres(qs)

        p = self.request.query_params
        annee = p.get("annee")
        centre_id = p.get("centre")
        departement = p.get("departement")
        type_declic = p.get("type_declic")
        date_min = p.get("date_min")
        date_max = p.get("date_max")
        search = p.get("search")
        ordering = p.get("ordering")
        avec_archivees = str(p.get("avec_archivees", "")).lower() in {"1", "true", "yes", "on"}
        archives_seules = str(p.get("archives_seules", "")).lower() in {"1", "true", "yes", "on"}

        if archives_seules:
            qs = qs.filter(is_active=False)
        elif not avec_archivees:
            qs = qs.filter(is_active=True)

        if annee:
            qs = qs.filter(date_declic__year=annee)
        if centre_id:
            qs = qs.filter(centre_id=centre_id)
        if type_declic:
            qs = qs.filter(type_declic=type_declic)

        if departement:
            departement = str(departement)
            qs = qs.filter(
                Q(centre__departement__startswith=departement) | Q(centre__code_postal__startswith=departement)
            )

        if date_min:
            qs = qs.filter(date_declic__gte=date_min)
        if date_max:
            qs = qs.filter(date_declic__lte=date_max)

        if search:
            qs = qs.filter(Q(centre__nom__icontains=search) | Q(commentaire__icontains=search))

        return qs.order_by(ordering or "-date_declic", "-id")

    def get_archived_aware_object(self):
        """
        Retourne une séance Déclic y compris archivée, en réappliquant le scope
        utilisateur courant.
        """
        lookup_field = getattr(self, "lookup_field", "pk")
        lookup_url_kwarg = getattr(self, "lookup_url_kwarg", None) or lookup_field
        lookup_value = self.kwargs.get(lookup_url_kwarg)
        queryset = Declic.objects.select_related("centre").prefetch_related("participants_declic")
        queryset = self._scope_qs_to_user_centres(queryset)
        return get_object_or_404(queryset, **{lookup_field: lookup_value})

    def destroy(self, request, *args, **kwargs):
        """
        Conserve `DELETE` pour compatibilité mais remplace la
        suppression physique par un archivage logique.
        """
        instance = self.get_object()
        if not instance.is_active:
            return self.success_response(
                data=self.get_serializer(instance).data,
                message="Séance Déclic déjà archivée.",
                status_code=status.HTTP_200_OK,
            )

        instance.is_active = False
        instance.save(user=request.user, update_fields=["is_active"])
        return self.success_response(
            data=self.get_serializer(instance).data,
            message="Séance Déclic archivée avec succès.",
            status_code=status.HTTP_200_OK,
        )

    @extend_schema(summary="Restaurer une séance Déclic archivée")
    @action(detail=True, methods=["post"], url_path="desarchiver")
    def desarchiver(self, request, pk=None):
        """
        Restaure une séance Déclic archivée et renvoie l'enveloppe API standard.
        """
        instance = self.get_archived_aware_object()
        if instance.is_active:
            return self.error_response(message="Séance Déclic déjà active.", status_code=status.HTTP_400_BAD_REQUEST)

        instance.is_active = True
        instance.save(user=request.user, update_fields=["is_active"])
        return self.success_response(
            data=self.get_serializer(instance).data,
            message="Séance Déclic désarchivée avec succès.",
            status_code=status.HTTP_200_OK,
        )

    def create(self, request, *args, **kwargs):
        """
        Crée une séance Déclic et renvoie l'enveloppe API standard afin
        d'aligner le contrat CRUD avec les autres modules métier.
        """
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            logger.warning("Validation Déclic échouée à la création: %s", serializer.errors)
            return self.error_response(
                message="Erreur de validation.",
                errors=serializer.errors,
                status_code=status.HTTP_400_BAD_REQUEST,
            )
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(
            {
                "success": True,
                "message": "Séance Déclic créée avec succès.",
                "data": self.get_serializer(serializer.instance).data,
            },
            status=status.HTTP_201_CREATED,
            headers=headers,
        )

    def retrieve(self, request, *args, **kwargs):
        """
        Retourne le détail d'une séance Déclic dans l'enveloppe API standard.
        """
        instance = self.get_object()
        return self.success_response(
            data=self.get_serializer(instance).data,
            message="Séance Déclic récupérée avec succès.",
        )

    def update(self, request, *args, **kwargs):
        """
        Met à jour une séance Déclic et renvoie l'enveloppe API standard.
        """
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if not serializer.is_valid():
            logger.warning("Validation Déclic échouée à la mise à jour #%s: %s", instance.pk, serializer.errors)
            return self.error_response(
                message="Erreur de validation.",
                errors=serializer.errors,
                status_code=status.HTTP_400_BAD_REQUEST,
            )
        self.perform_update(serializer)
        return self.success_response(
            data=self.get_serializer(serializer.instance).data,
            message="Séance Déclic mise à jour avec succès.",
        )

    # -------------------------------------------------------------------------
    # 💾 CREATE / UPDATE
    # -------------------------------------------------------------------------
    def perform_create(self, serializer):
        """
        Méthode d'infrastructure DRF pour la création (POST).

        Ajoute la sauvegarde de l'utilisateur ayant effectué l'action via save(user=...).
        Permissions : soumises au décorateur classe (cf. IsDeclicStaffOrAbove et ScopeMixin).
        """
        self._assert_user_can_use_centre(serializer.validated_data.get("centre"))
        serializer.save()

    def perform_update(self, serializer):
        """
        Méthode d'infrastructure DRF pour la mise à jour (PUT/PATCH).

        Ajoute la sauvegarde de l'utilisateur ayant effectué l'action via save(user=...).
        """
        current = serializer.instance
        new_centre = serializer.validated_data.get("centre", getattr(current, "centre", None))
        self._assert_user_can_use_centre(new_centre)
        serializer.save()

    def _assert_user_can_use_centre(self, centre):
        if not centre:
            return

        allowed_ids = self._centre_ids_for_user(self.request.user)
        if allowed_ids is None:
            return
        if getattr(centre, "id", None) not in set(allowed_ids):
            raise PermissionDenied("Centre hors de votre périmètre d'accès.")

    # -------------------------------------------------------------------------
    # 📊 STATISTIQUES
    # -------------------------------------------------------------------------
    @action(detail=False, methods=["get"], url_path="stats-centres")
    def stats_centres(self, request):
        """
        Point d'entrée : GET /declic/stats-centres/?annee=YYYY

        Objectif métier : Statistiques sur le nombre de candidats accueillis, agrégées par centre,
        pour l'année demandée (ou année courante par défaut).

        Restrictions de visibilité : appliquées automatiquement via ScopeMixin et permissions.
        Structure de réponse : dépendante du résultat de Declic.accueillis_par_centre(annee).
        Aucun format JSON garanti visible ici.
        """
        annee = int(request.query_params.get("annee", localdate().year))
        department_scope = use_department_stats_scope(request)
        centres_qs = self._scope_qs_to_user_centres(Centre.objects.all())
        sessions_qs = self._scope_qs_to_user_centres(Declic.objects.filter(date_declic__year=annee))
        if department_scope and not is_admin_like(request.user):
            departements = get_staff_department_codes_cached(request) or []
            if departements:
                dep_q = build_department_scope_q("code_postal", departements)
                centres_qs = Centre.objects.filter(dep_q)
                sessions_qs = Declic.objects.filter(date_declic__year=annee).filter(
                    build_department_scope_q("centre__code_postal", departements)
                )
        centres_qs = centres_qs.order_by("nom")
        totals_by_centre = {
            row["centre_id"]: row["total"] or 0
            for row in sessions_qs.values("centre_id").annotate(total=Sum("nb_presents_declic"))
        }
        data = {centre.nom: totals_by_centre.get(centre.id, 0) for centre in centres_qs}
        return self.success_response(data=data, message="Statistiques Déclic par centre récupérées avec succès.")

    @action(detail=False, methods=["get"], url_path="stats-departements")
    def stats_departements(self, request):
        """
        Point d'entrée : GET /declic/stats-departements/?annee=YYYY

        Objectif métier : Statistiques analogues au endpoint ci-dessus, mais agrégées par département.
        Restrictions et structure : idem.
        """
        annee = int(request.query_params.get("annee", localdate().year))
        sessions_qs = self._scope_qs_to_user_centres(Declic.objects.filter(date_declic__year=annee).select_related("centre"))
        data = {}
        for session in sessions_qs:
            centre = getattr(session, "centre", None)
            dep = getattr(centre, "departement", None) or ((getattr(centre, "code_postal", "") or "")[:2] or None)
            if dep:
                data[dep] = data.get(dep, 0) + (session.nb_presents_declic or 0)
        data = dict(sorted(data.items()))
        return self.success_response(data=data, message="Statistiques Déclic par département récupérées avec succès.")

    # -------------------------------------------------------------------------
    # 📤 EXPORT EXCEL ATELIERS UNIQUEMENT
    # -------------------------------------------------------------------------
    @action(detail=False, methods=["get"], url_path="export-xlsx")
    def export_xlsx(self, request):
        """
        Point d'entrée : GET /declic/export-xlsx/?annee=YYYY

        Objectif métier : Exporter tous les ateliers accessibles, filtrés selon les droits et paramètres query string,
        dans un fichier Excel unique.

        Permissions : mêmes contrôles que le ViewSet principal.

        Détail Excel :
        - Lignes : 1 en-tête fixe puis une ligne par atelier, colonnes
          [ID, Atelier, Date, Centre, Inscrits, Présents, Absents, Taux présence, Objectif annuel,
           Ateliers cumulés, Taux atteinte, Reste à faire, Commentaire].
        - Statut HTTP : 200, Content-Type =
          application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
        - Content-Disposition : attachment ; nom reprenant l'année exportée.

        Structure Excel stricte ; aucun retour JSON ni structure DRF dédiée.
        """
        qs = self.filter_queryset(self.get_queryset())
        annee = int(request.query_params.get("annee", localdate().year))

        wb = Workbook()
        ws = wb.active
        ws.title = "Ateliers Déclic"

        headers = [
            "ID",
            "Atelier",
            "Date",
            "Centre",
            "Inscrits",
            "Présents",
            "Absents",
            "Taux présence (%)",
            "Objectif annuel",
            "Ateliers cumulés",
            "Taux atteinte (%)",
            "Reste à faire",
            "Commentaire",
        ]
        ws.append(headers)

        border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        for s in qs:
            objectif = s.objectif_annuel

            realise = (
                Declic.objects.filter(
                    centre=s.centre,
                    date_declic__year=s.date_declic.year,
                ).aggregate(
                    total=Sum("nb_presents_declic")
                )["total"]
                or 0
            )

            taux = round((realise / objectif) * 100, 1) if objectif else 0

            ws.append(
                [
                    s.id,
                    s.get_type_declic_display(),
                    s.date_declic.strftime("%d/%m/%Y"),
                    s.centre.nom if s.centre else "",
                    s.nb_inscrits_declic,
                    s.nb_presents_declic,
                    s.nb_absents_declic,
                    s.taux_presence_declic,
                    objectif,
                    realise,
                    taux,
                    max(objectif - realise, 0),
                    (s.commentaire or "").replace("\n", " "),
                ]
            )

        buffer = BytesIO()
        wb.save(buffer)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="declic_ateliers_{annee}.xlsx"'
        return response
