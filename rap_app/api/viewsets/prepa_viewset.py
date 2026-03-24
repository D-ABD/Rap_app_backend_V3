# rap_app/api/viewsets/prepa_viewset.py

from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.utils import timezone as dj_timezone
from django.utils.timezone import localdate
from drf_spectacular.utils import (
    OpenApiExample,
    OpenApiParameter,
    extend_schema,
    extend_schema_view,
)
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.response import Response

from ...models.centres import Centre
from ...models.prepa import ObjectifPrepa, Prepa
from ..mixins import ApiResponseMixin
from ..permissions import IsPrepaStaffOrAbove
from ..roles import (
    is_admin_like,
    is_candidate,
    is_prepa_staff,
    is_staff_or_staffread,
)
from ..serializers.prepa_serializers import PrepaSerializer


@extend_schema_view(
    list=extend_schema(
        summary="Lister toutes les séances Prépa",
        description="Retourne la liste complète des séances Prépa (IC et ateliers).",
        parameters=[
            OpenApiParameter(name="annee", description="Filtrer par année", required=False, type=int),
            OpenApiParameter(name="centre", description="Filtrer par centre (ID)", required=False, type=int),
            OpenApiParameter(
                name="departement",
                description="Filtrer par département (préfixe CP ou champ centre.departement)",
                required=False,
                type=str,
            ),
            OpenApiParameter(
                name="type_prepa", description="Filtrer par type d’activité Prépa", required=False, type=str
            ),
            OpenApiParameter(name="date_min", description="Date minimale (YYYY-MM-DD)", required=False, type=str),
            OpenApiParameter(name="date_max", description="Date maximale (YYYY-MM-DD)", required=False, type=str),
            OpenApiParameter(
                name="search", description="Recherche sur le nom du centre ou le commentaire", required=False, type=str
            ),
            OpenApiParameter(
                name="ordering", description="Tri (ex: -date_prepa, centre__nom)", required=False, type=str
            ),
        ],
        responses={200: PrepaSerializer},
    ),
)
class PrepaViewSet(ApiResponseMixin, viewsets.ModelViewSet):
    """
    ViewSet principal du module Prépa.

    Il gère les séances Prépa visibles pour l'utilisateur courant avec un
    helper de scope local par centre, des actions de métadonnées/front, des
    statistiques et un export Excel.
    """

    serializer_class = PrepaSerializer
    permission_classes = [IsPrepaStaffOrAbove]

    # Par défaut, queryset complet, filtré ensuite dynamiquement (cf. get_queryset)
    queryset = Prepa.objects.select_related("centre").prefetch_related("stagiaires_prepa").all()

    # ---------------------------------------------------
    # 🔹 Options de filtres pour le frontend
    # ---------------------------------------------------
    @extend_schema(
        summary="Options de filtres disponibles pour Prépa",
        description="Retourne les années, départements, centres et types Prépa disponibles pour les filtres du module Prépa.",
        responses={200: dict},
    )
    @action(detail=False, methods=["get"], url_path="filters")
    def filters(self, request):
        """
        Retourne les valeurs disponibles pour les filtres du module
        Prépa (années, départements, centres et types de Prépa) en
        fonction du périmètre de centres accessible.
        """
        from ...models.centres import (
            Centre,  # import local pour éviter les imports circulaires
        )

        # (aucun changement fonctionnel, docstring uniquement)
        # Cf. code source pour détail total des valeurs retournées.
        annees = self._scope_qs_to_user_centres(Prepa.objects.all()).order_by().values_list("date_prepa__year", flat=True).distinct()
        annees = sorted([a for a in annees if a is not None], reverse=True)

        centres_qs = self._scope_qs_to_user_centres(Centre.objects.all())
        centres_data, departements_set = [], set()

        for c in centres_qs.order_by("nom"):
            code_postal = getattr(c, "code_postal", "") or ""
            dep = getattr(c, "departement", "") or (code_postal[:2] if len(code_postal) >= 2 else "")
            if dep:
                departements_set.add(dep)

            centres_data.append(
                {
                    "value": c.id,
                    "label": c.nom,
                    "departement": dep,
                    "code_postal": code_postal,
                }
            )

        departements = [{"value": d, "label": f"Département {d}"} for d in sorted(departements_set)]
        types = [{"value": t[0], "label": t[1]} for t in Prepa.TypePrepa.choices]

        return self.success_response(
            data={
                "annees": annees,
                "departements": departements,
                "centres": centres_data,
                "type_prepa": types,
            },
            message="Filtres Prépa récupérés avec succès.",
        )

    # ---------------------------------------------------
    # 🔹 Filtrage principal (get_queryset)
    # ---------------------------------------------------
    def get_queryset(self):
        """
        Retourne les séances Prépa visibles pour l'utilisateur après
        application du scope local par centre puis des filtres métier.
        """
        qs = Prepa.objects.select_related("centre").prefetch_related("stagiaires_prepa")
        qs = self._scope_qs_to_user_centres(qs)

        params = self.request.query_params
        only = params.get("only")
        if only == "ateliers":
            qs = qs.filter(type_prepa__startswith="atelier") | qs.filter(type_prepa="autre")
        elif only == "ic":
            qs = qs.filter(type_prepa=Prepa.TypePrepa.INFO_COLLECTIVE)

        annee = params.get("annee")
        centre_id = params.get("centre")
        departement = params.get("departement")
        date_min = params.get("date_min")
        date_max = params.get("date_max")
        search = params.get("search")
        ordering = params.get("ordering")

        if annee:
            qs = qs.filter(date_prepa__year=annee)
        if centre_id:
            qs = qs.filter(centre_id=centre_id)

        # Supporte type_prepa=xx, type_prepa[]=xx, et plusieurs valeurs
        type_prepa_list = params.getlist("type_prepa") + params.getlist("type_prepa[]")

        # Nettoyage
        type_prepa_list = [t for t in type_prepa_list if t]

        if type_prepa_list:
            qs = qs.filter(type_prepa__in=type_prepa_list)

        if departement:
            departement = str(departement).strip()
            qs = qs.filter(centre__code_postal__startswith=departement)

        if date_min:
            qs = qs.filter(date_prepa__gte=date_min)
        if date_max:
            qs = qs.filter(date_prepa__lte=date_max)

        if search:
            qs = qs.filter(Q(centre__nom__icontains=search) | Q(commentaire__icontains=search))

        qs = qs.order_by(ordering or "-date_prepa", "-id")

        return qs

    # ---------------------------------------------------
    # 🔹 Métadonnées pour le front (usePrepaMeta)
    # ---------------------------------------------------
    @extend_schema(
        summary="Métadonnées du module Prépa",
        description="Retourne les choix statiques ou contextuels pour le module Prépa (types, centres, etc.).",
        responses={200: dict},
    )
    @action(detail=False, methods=["get"], url_path="meta")
    def meta(self, request):
        """
        Retourne les choix pour le module Prépa (types d'activité et
        centres accessibles) destinés à alimenter les formulaires et
        filtres frontend.
        """
        centres_qs = self._scope_qs_to_user_centres(Centre.objects.all())

        centres = [
            {"id": c.id, "nom": c.nom, "departement": c.departement, "code_postal": c.code_postal}
            for c in centres_qs.order_by("nom")
        ]

        types = [{"value": t[0], "label": t[1]} for t in Prepa.TypePrepa.choices]

        data = {
            "type_prepa": types,
            "centres": centres,
        }
        return self.success_response(data=data, message="Métadonnées Prépa récupérées avec succès.")

    # ---------------------------------------------------
    # 🔹 Helpers centres / permissions
    # ---------------------------------------------------
    def _admin_like(self, user) -> bool:
        # Indique si l'utilisateur dispose de droits complets de type
        # administrateur sur le module Prépa.
        return is_admin_like(user)

    def _prepa_staff_centre_ids(self, user):
        # Retourne les identifiants de centres accessibles dans Prépa
        # pour les profils staff/prepa_staff, ou None pour les profils
        # admin-like.
        if self._admin_like(user):
            return None

        if is_prepa_staff(user):
            centres = getattr(user, "centres", None)
            if not centres or not centres.exists():
                return []
            return list(centres.values_list("id", flat=True))

        if is_staff_or_staffread(user):
            centres = getattr(user, "centres", None)
            if not centres or not centres.exists():
                return []
            return list(centres.values_list("id", flat=True))

        return []

    def _scope_qs_to_user_centres(self, qs):
        # Restreint un queryset Prepa ou Centre au périmètre des centres
        # accessibles pour l'utilisateur connecté, ou renvoie qs.none()
        # pour les profils non autorisés.
        user = self.request.user

        if not user.is_authenticated or is_candidate(user):
            return qs.none()

        centre_ids = self._prepa_staff_centre_ids(user)

        if centre_ids is None:
            return qs

        if not centre_ids:
            return qs.none()

        model = getattr(qs, "model", None)
        if model and hasattr(model, "_meta"):
            field_names = [f.name for f in model._meta.get_fields()]

            if "centre_id" in field_names or "centre" in field_names:
                return qs.filter(centre_id__in=centre_ids).distinct()

            if model._meta.model_name == "centre" or "id" in field_names:
                return qs.filter(id__in=centre_ids).distinct()

        return qs.none()

    def _assert_user_can_use_centre(self, centre):
        # Empêche la création ou la modification d'une instance Prépa
        # liée à un centre en dehors du périmètre de l'utilisateur.
        if not centre:
            return
        user = self.request.user

        if self._admin_like(user):
            return

        allowed_ids = set(self._prepa_staff_centre_ids(user) or [])

        if getattr(centre, "id", None) not in allowed_ids:
            raise PermissionDenied("Centre hors de votre périmètre d'accès.")

    # ---------------------------------------------------
    # 🔹 CREATE / UPDATE contrôlé
    # ---------------------------------------------------
    def perform_create(self, serializer):
        """
        Crée une séance Prépa puis vérifie que le centre associé est
        autorisé pour l'utilisateur avant la sauvegarde finale.
        """
        self._assert_user_can_use_centre(serializer.validated_data.get("centre"))
        serializer.save()

    def perform_update(self, serializer):
        """
        Met à jour une séance Prépa après contrôle du centre associé
        et sauvegarde avec ou sans paramètre user selon le modèle.
        """
        current = serializer.instance
        new_centre = serializer.validated_data.get("centre", getattr(current, "centre", None))
        self._assert_user_can_use_centre(new_centre)

        serializer.save()

    # ---------------------------------------------------
    # 🔹 Actions statistiques
    # ---------------------------------------------------
    @action(detail=False, methods=["get"], url_path="stats-centres")
    def stats_centres(self, request):
        """
        Retourne les statistiques de séances Prépa par centre pour une
        année donnée en s'appuyant sur Prepa.accueillis_par_centre.
        """
        annee = int(request.query_params.get("annee", localdate().year))
        centres_qs = self._scope_qs_to_user_centres(Centre.objects.all()).order_by("nom")
        sessions_qs = self._scope_qs_to_user_centres(
            Prepa.objects.filter(date_prepa__year=annee, type_prepa=Prepa.TypePrepa.ATELIER1)
        )
        totals_by_centre = {
            row["centre_id"]: row["total"] or 0
            for row in sessions_qs.values("centre_id").annotate(total=Sum("nb_presents_prepa"))
        }
        data = {centre.nom: totals_by_centre.get(centre.id, 0) for centre in centres_qs}
        return self.success_response(data=data, message="Statistiques Prépa par centre récupérées avec succès.")

    @action(detail=False, methods=["get"], url_path="stats-departements")
    def stats_departements(self, request):
        """
        Retourne les statistiques de séances Prépa par département pour
        une année donnée en s'appuyant sur Prepa.accueillis_par_departement.
        """
        annee = int(request.query_params.get("annee", localdate().year))
        sessions_qs = self._scope_qs_to_user_centres(
            Prepa.objects.filter(date_prepa__year=annee, type_prepa=Prepa.TypePrepa.ATELIER1).select_related("centre")
        )
        data = {}
        for session in sessions_qs:
            centre = getattr(session, "centre", None)
            dep = getattr(centre, "departement", None) or ((getattr(centre, "code_postal", "") or "")[:2] or None)
            if dep:
                data[dep] = data.get(dep, 0) + (session.nb_presents_prepa or 0)
        data = dict(sorted(data.items()))
        return self.success_response(data=data, message="Statistiques Prépa par département récupérées avec succès.")

    @action(detail=False, methods=["get"], url_path="reste-a-faire-total")
    def reste_a_faire_total(self, request):
        """
        Retourne le reste à faire agrégé pour tous les centres Prépa
        sur une année donnée via Prepa.reste_a_faire_total.
        """
        annee = int(request.query_params.get("annee", localdate().year))
        objectifs_qs = self._scope_qs_to_user_centres(ObjectifPrepa.objects.filter(annee=annee))
        realise_qs = self._scope_qs_to_user_centres(
            Prepa.objects.filter(date_prepa__year=annee, type_prepa=Prepa.TypePrepa.ATELIER1)
        )
        objectif_total = objectifs_qs.aggregate(total=Sum("valeur_objectif"))["total"] or 0
        realise_total = realise_qs.aggregate(total=Sum("nb_presents_prepa"))["total"] or 0
        total = max(objectif_total - realise_total, 0)
        return self.success_response(
            data={"annee": annee, "reste_total": total},
            message="Reste à faire Prépa récupéré avec succès.",
        )

    # ---------------------------------------------------
    # 🔹 Export Excel complet PREPA
    # ---------------------------------------------------
    @extend_schema(
        summary="Export Excel complet Prépa",
        description="Exporte toutes les séances Prépa enrichies avec objectifs et indicateurs.",
        responses={200: {"type": "binary", "format": "xlsx"}},
        tags=["Prépa"],
    )
    @action(detail=False, methods=["get"], url_path="export-xlsx")
    def export_xlsx(self, request):
        """
        Exporte en fichier Excel les séances Prépa filtrées et
        accessibles à l'utilisateur, en incluant indicateurs et
        objectifs liés.
        """
        qs = self.filter_queryset(self.get_queryset())
        annee = int(request.query_params.get("annee", localdate().year))

        wb = Workbook()
        ws = wb.active
        ws.title = "Prépa – Données complètes"

        # === Logo + titre ===
        try:
            logo_path = Path(settings.BASE_DIR) / "rap_app/static/images/logo.png"
            if logo_path.exists():
                img = XLImage(str(logo_path))
                img.height = 60
                img.width = 120
                ws.add_image(img, "A1")
        except Exception:
            pass

        ws.merge_cells("B1:Z1")
        ws["B1"] = "Export complet Prépa — Rap_App"
        ws["B1"].font = Font(name="Calibri", bold=True, size=15, color="004C99")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("B2:Z2")
        ws["B2"] = f"Export réalisé le {dj_timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        ws["B2"].font = Font(name="Calibri", italic=True, size=10, color="666666")
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.append([])
        ws.append([])

        # === En-têtes ===
        headers = [
            "ID",
            "Type activité",
            "Date",
            "Centre",
            "Places ouvertes (IC)",
            "Prescriptions (IC)",
            "Présents (IC)",
            "Absents (IC)",
            "Adhésions (IC)",
            "Inscrits (Atelier)",
            "Présents (Atelier)",
            "Absents (Atelier)",
            "Taux prescription (%)",
            "Taux présence IC (%)",
            "Taux adhésion (%)",
            "Taux présence atelier (%)",
            "Année objectif",
            "Objectif annuel (centre)",
            "Réalisé (IC cumulés)",
            "Taux atteinte annuel (%)",
            "Reste à faire",
            "Département centre",
            "Taux prescription (objectif)",
            "Taux présence (objectif)",
            "Taux adhésion (objectif)",
            "Taux atteinte (objectif)",
            "Reste à faire (objectif centre)",
            "Commentaire séance",
            "Commentaire objectif",
        ]
        ws.append(headers)
        header_fill = PatternFill("solid", fgColor="DCE6F1")
        border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        for cell in ws[ws.max_row]:
            cell.font = Font(name="Calibri", bold=True, color="002060")
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

        # === Données ===
        even_fill = PatternFill("solid", fgColor="F8FBFF")
        odd_fill = PatternFill("solid", fgColor="FFFFFF")
        objectifs_cache = {(o.centre_id, o.annee): o for o in ObjectifPrepa.objects.select_related("centre").all()}

        for i, s in enumerate(qs, start=1):
            key = (s.centre_id, s.date_prepa.year if s.date_prepa else annee)
            obj = objectifs_cache.get(key)
            obj_data = obj.synthese_globale() if obj else {}
            dep = obj.departement if obj else getattr(s.centre, "departement", "")

            ws.append(
                [
                    s.id,
                    s.get_type_prepa_display(),
                    s.date_prepa.strftime("%d/%m/%Y") if s.date_prepa else "",
                    getattr(s.centre, "nom", ""),
                    s.nombre_places_ouvertes,
                    s.nombre_prescriptions,
                    s.nb_presents_info,
                    s.nb_absents_info,
                    s.nb_adhesions,
                    s.nb_inscrits_prepa,
                    s.nb_presents_prepa,
                    s.nb_absents_prepa,
                    s.taux_prescription,
                    s.taux_presence_info,
                    s.taux_adhesion,
                    s.taux_presence_prepa,
                    key[1],
                    s.objectif_annuel,
                    s.nb_presents_info,
                    s.taux_atteinte_annuel,
                    s.reste_a_faire,
                    dep,
                    obj_data.get("taux_prescription", ""),
                    obj_data.get("taux_presence", ""),
                    obj_data.get("taux_adhesion", ""),
                    obj_data.get("taux_atteinte", ""),
                    obj_data.get("reste_a_faire", ""),
                    (s.commentaire or "").replace("\n", " "),
                    (obj.commentaire if obj else "") or "",
                ]
            )

            fill = even_fill if i % 2 == 0 else odd_fill
            for cell in ws[ws.max_row]:
                cell.fill = fill
                cell.border = border
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(vertical="center", wrapText=True)
            ws.row_dimensions[ws.max_row].height = 22

        # === Filtres & largeur ===
        end_row = ws.max_row
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}{end_row}"
        ws.freeze_panes = "A2"

        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 45)

        # === Réponse ===
        buffer = BytesIO()
        wb.save(buffer)
        binary_content = buffer.getvalue()

        filename = f'prepa_unifie_{dj_timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            binary_content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = len(binary_content)
        return response
