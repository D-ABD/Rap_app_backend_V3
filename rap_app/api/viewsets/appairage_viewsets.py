from io import BytesIO
from pathlib import Path

import django_filters
from django.conf import settings
from django.core.exceptions import ValidationError as DjangoValidationError
from django.db.models import OuterRef, Prefetch, Q, Subquery
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone as dj_timezone
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import OpenApiResponse, OpenApiTypes, extend_schema
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from rest_framework import filters, status
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.response import Response

from ...models.appairage import Appairage, AppairageActivite, AppairageStatut
from ...models.commentaires_appairage import CommentaireAppairage
from ...services.placement_services import AppairagePlacementService, defer_appairage_snapshot_sync
from ...utils.filters import AppairageFilterSet
from ..paginations import RapAppPagination
from ..permissions import IsStaffOrAbove, is_staff_or_staffread
from ..roles import is_admin_like
from ..serializers.appairage_serializers import (
    AppairageCreateUpdateSerializer,
    AppairageListSerializer,
    AppairageMetaSerializer,
    AppairageSerializer,
    CommentaireAppairageSerializer,
)
from ..serializers.commentaires_appairage_serializers import CommentaireAppairageWriteSerializer
from .scoped_viewset import ScopedModelViewSet


class AppairageViewSet(ScopedModelViewSet):
    """
    ViewSet principal des appairages.

    Source de vérité actuelle :
    - accès protégé par `IsStaffOrAbove`
    - scoping via `ScopedModelViewSet` avec visibilité par centre sur la
      formation directe ou la formation du candidat
    - serializers par action :
      - `list` => `AppairageListSerializer`
      - `create` / `update` / `partial_update` => `AppairageCreateUpdateSerializer`
      - autres actions => `AppairageSerializer`
    - `perform_create()` et `perform_update()` orchestrent désormais :
      - validations métier
      - sauvegarde de l'appairage
      - appel explicite à `AppairagePlacementService.sync_after_save()`

    Les signaux résiduels sur ce domaine sont conservés en observation,
    mais la synchronisation métier passe désormais par le service.
    """

    base_queryset = (
        Appairage.objects.all()
        .select_related(
            "candidat",
            "candidat__formation",
            "candidat__formation__centre",
            "candidat__formation__type_offre",
            "candidat__formation__statut",
            "partenaire",
            "formation",
            "formation__centre",
            "formation__type_offre",
            "formation__statut",
            "created_by",
            "updated_by",
        )
        .prefetch_related(
            "historiques",
            Prefetch(
                "commentaires",
                queryset=CommentaireAppairage.objects.select_related("created_by").order_by("-created_at", "-pk"),
            ),
        )
    )

    permission_classes = [IsStaffOrAbove]
    pagination_class = RapAppPagination
    serializer_class = AppairageSerializer
    scope_mode = "centre"
    centre_lookup_paths = ("formation__centre_id", "candidat__formation__centre_id")

    filterset_class = AppairageFilterSet
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["statut", "formation", "candidat", "partenaire", "created_by"]

    search_fields = [
        "candidat__nom",
        "candidat__prenom",
        "partenaire__nom",
        "partenaire_contact_nom",
        "formation__nom",
        "candidat__formation__nom",
        "formation__centre__nom",
        "candidat__formation__centre__nom",
        "commentaires__body",
        "retour_partenaire",
        "created_by__first_name",
        "created_by__last_name",
        "created_by__username",
        "created_by__email",
    ]
    ordering_fields = [
        "date_appairage",
        "statut",
        "formation__nom",
        "formation__centre__nom",
        "updated_at",
        "created_at",
    ]

    def _assert_staff_can_use_formation(self, formation):
        # Refuse l’utilisation d’une formation d’un centre non autorisé lors d'une écriture
        if not formation:
            return
        user = self.request.user
        if is_admin_like(user):
            return
        if is_staff_or_staffread(user):
            allowed = set(user.centres.values_list("id", flat=True))
            if getattr(formation, "centre_id", None) not in allowed:
                raise PermissionDenied("Formation hors de votre périmètre (centre).")

    def _user_display(self, user):
        if not user:
            return ""
        return f"{user.get_full_name()} ({user.username})" if hasattr(user, "username") else str(user)

    def _partenaire_email(self, a):
        return getattr(a.partenaire, "email", "") or getattr(a, "partenaire_email", "") or ""

    def _partenaire_telephone(self, a):
        return getattr(a.partenaire, "telephone", "") or getattr(a, "partenaire_telephone", "") or ""

    def _formation_id(self, a):
        return getattr(a.formation, "id", "") or ""

    def _formation_label(self, a):
        return getattr(a.formation, "nom", "") or ""

    def _formation_type_offre(self, a):
        return getattr(getattr(a.formation, "type_offre", None), "libelle", "") or getattr(
            a, "formation_type_offre", ""
        )

    def _formation_num_offre(self, a):
        return getattr(a.formation, "num_offre", "") or getattr(a, "formation_numero_offre", "")

    def get_serializer_class(self):
        """
        Sélectionne le serializer adapté au type de réponse attendu par
        l'action courante.
        """
        if self.action == "list":
            return AppairageListSerializer
        elif self.action in ["create", "update", "partial_update"]:
            return AppairageCreateUpdateSerializer
        return AppairageSerializer

    def perform_create(self, serializer):
        """
        Orchestration de création d'un appairage.

        Le flux réel est :
        - refus candidats/stagiaires
        - contrôle du périmètre centre sur la formation
        - contrôle d'unicité métier `(candidat, partenaire, formation)`
        - sauvegarde avec différé du vieux sync implicite
        - synchronisation explicite du snapshot candidat via service
        """
        user = self.request.user
        if hasattr(user, "is_candidat_or_stagiaire") and user.is_candidat_or_stagiaire():
            raise PermissionDenied("Les candidats/stagiaires ne peuvent pas créer d’appairage.")

        formation_payload = serializer.validated_data.get("formation")
        candidat_payload = serializer.validated_data.get("candidat")
        formation = (
            formation_payload
            or (getattr(getattr(candidat_payload, "formation", None), "pk", None) and candidat_payload.formation)
            or None
        )

        if formation:
            self._assert_staff_can_use_formation(formation)

        partenaire_payload = serializer.validated_data.get("partenaire")
        if Appairage.objects.filter(
            candidat=candidat_payload, partenaire=partenaire_payload, formation=formation
        ).exists():
            raise ValidationError(
                {"detail": "Un appairage existe déjà pour ce candidat, ce partenaire et cette formation."}
            )

        with defer_appairage_snapshot_sync():
            instance = serializer.save(
                created_by=user,
                updated_by=user,
                formation=formation or serializer.validated_data.get("formation"),
            )

        AppairagePlacementService.sync_after_save(instance, actor=user)

    def perform_update(self, serializer):
        """
        Orchestration de mise à jour d'un appairage avec le même principe
        que `perform_create()`, puis resynchronisation explicite du placement.
        """
        user = self.request.user
        instance = serializer.instance
        previous_candidat = instance.candidat
        if hasattr(user, "is_candidat_or_stagiaire") and user.is_candidat_or_stagiaire():
            raise PermissionDenied("Les candidats/stagiaires ne peuvent pas modifier un appairage.")

        data_formation = serializer.validated_data.get("formation", instance.formation)
        if data_formation:
            self._assert_staff_can_use_formation(data_formation)

        with defer_appairage_snapshot_sync():
            instance = serializer.save(formation=data_formation, updated_by=user)

        AppairagePlacementService.sync_after_save(instance, actor=user, previous_candidat=previous_candidat)

    @action(detail=False, methods=["get"], url_path="meta")
    def meta(self, request):
        """
        Retourne les métadonnées utiles à l'initialisation des formulaires
        d'appairage pour l'utilisateur courant.
        """
        serializer = AppairageMetaSerializer(instance={}, context={"request": request})
        return self.success_response(data=serializer.data, message="Métadonnées appairage récupérées avec succès.")

    @action(detail=True, methods=["get", "post"], url_path="commentaires")
    def commentaires(self, request, pk=None):
        """Lit ou crée les commentaires attachés à un appairage donné."""
        appairage = self.get_object()
        if request.method == "GET":
            qs = appairage.commentaires.select_related("created_by").order_by("-created_at")
            serializer = CommentaireAppairageSerializer(qs, many=True)
            return self.success_response(data=serializer.data, message="Commentaires d'appairage récupérés avec succès.")

        if request.method == "POST":
            write_serializer = CommentaireAppairageWriteSerializer(data=request.data)
            if write_serializer.is_valid():
                try:
                    commentaire = write_serializer.save(created_by=request.user, appairage=appairage)
                except DjangoValidationError as exc:
                    return self.error_response(
                        message="Impossible de créer le commentaire d'appairage.",
                        errors=getattr(exc, "message_dict", None) or {"non_field_errors": list(exc.messages)},
                        status_code=status.HTTP_400_BAD_REQUEST,
                    )
                serializer = CommentaireAppairageSerializer(commentaire)
                return self.success_response(
                    data=serializer.data,
                    message="Commentaire d'appairage créé avec succès.",
                    status_code=status.HTTP_201_CREATED,
                )
            return self.error_response(
                message="Impossible de créer le commentaire d'appairage.",
                errors=write_serializer.errors,
                status_code=status.HTTP_400_BAD_REQUEST,
            )

    def get_queryset(self):
        """
        Retourne les appairages visibles après application du scope centre,
        des filtres métier de période/activité et de l'annotation du dernier
        commentaire visible pour les usages de liste ou d'export.
        """
        qs = self.base_queryset

        last_comment_qs = (
            CommentaireAppairage.objects.filter(appairage=OuterRef("pk")).order_by("-created_at").values("body")[:1]
        )
        qs = qs.annotate(last_commentaire=Subquery(last_comment_qs))

        annee = self.request.query_params.get("annee")
        if annee:
            try:
                annee = int(annee)
                qs = qs.filter(date_appairage__year=annee)
            except ValueError:
                pass

        date_min = self.request.query_params.get("date_min")
        date_max = self.request.query_params.get("date_max")

        if date_min:
            qs = qs.filter(date_appairage__date__gte=date_min)

        if date_max:
            qs = qs.filter(date_appairage__date__lte=date_max)

        activite = self.request.query_params.get("activite")
        if activite in [AppairageActivite.ACTIF, AppairageActivite.ARCHIVE]:
            qs = qs.filter(activite=activite)
        else:
            avec_archivees = self.request.query_params.get("avec_archivees")
            if not (avec_archivees and str(avec_archivees).lower() in ["1", "true", "yes", "on"]):
                qs = qs.exclude(activite=AppairageActivite.ARCHIVE)

        return self.scope_queryset(qs)

    def retrieve(self, request, *args, **kwargs):
        """
        [GET] /appairages/<id>/
        Donne accès à un appairage, y compris archivé, mais uniquement dans le périmètre centre autorisé.
        """
        obj = self._get_object_including_archived_scoped(kwargs.get("pk"))
        serializer = self.get_serializer(obj)
        return self.success_response(data=serializer.data, message="Appairage récupéré avec succès.")

    @action(detail=True, methods=["post"], url_path="archiver")
    def archiver(self, request, pk=None):
        """
        [POST] /appairages/<id>/archiver/
        Archive un appairage visible dans le périmètre utilisateur.
        """
        appairage = self._get_object_including_archived_scoped(pk)

        if appairage.activite == AppairageActivite.ARCHIVE:
            return self.error_response(message="Déjà archivé.", status_code=status.HTTP_400_BAD_REQUEST)

        if hasattr(appairage, "archiver"):
            appairage.archiver(user=request.user)
        else:
            appairage.activite = AppairageActivite.ARCHIVE
            appairage.save(user=request.user, update_fields=["activite"])

        AppairagePlacementService.sync_candidate_snapshot(appairage.candidat, actor=request.user)

        return self.success_response(data={"status": "archived"}, message="Appairage archivé avec succès.")

    @action(detail=True, methods=["post"], url_path="desarchiver")
    def desarchiver(self, request, pk=None):
        """
        [POST] /appairages/<id>/desarchiver/
        Désarchive un appairage visible dans le périmètre utilisateur.
        """
        appairage = self._get_object_including_archived_scoped(pk)

        if appairage.activite != AppairageActivite.ARCHIVE:
            return self.error_response(
                message="Cet appairage n’est pas archivé.",
                status_code=status.HTTP_400_BAD_REQUEST,
            )

        if hasattr(appairage, "desarchiver"):
            appairage.desarchiver(user=request.user)
        else:
            appairage.activite = AppairageActivite.ACTIF
            appairage.save(user=request.user, update_fields=["activite"])

        AppairagePlacementService.sync_candidate_snapshot(appairage.candidat, actor=request.user)

        return self.success_response(data={"status": "unarchived"}, message="Appairage désarchivé avec succès.")

    def destroy(self, request, *args, **kwargs):
        """
        Supprime un appairage puis resynchronise le snapshot de placement du
        candidat concerné à partir du dernier appairage actif restant.
        """
        appairage = self._get_object_including_archived_scoped(kwargs.get("pk"))
        candidat = appairage.candidat

        with defer_appairage_snapshot_sync():
            appairage.delete(user=request.user)

        AppairagePlacementService.sync_candidate_snapshot(candidat, actor=request.user)
        return self.success_response(data=None, message="Appairage supprimé avec succès.")

    def _get_object_including_archived_scoped(self, pk):
        """
        Retourne un appairage visible par l'utilisateur courant, y compris s'il est archivé.
        Applique le scope centre, mais n'applique pas le filtre d'activité de get_queryset().
        """
        qs = self.base_queryset

        last_comment_qs = (
            CommentaireAppairage.objects.filter(appairage=OuterRef("pk")).order_by("-created_at").values("body")[:1]
        )
        qs = qs.annotate(last_commentaire=Subquery(last_comment_qs))

        qs = self.scope_queryset(qs)
        return get_object_or_404(qs, pk=pk)

    # ----------------------------------------------------------------------
    # 🧮 Export : génération fichier Excel des appairages visibles
    # ----------------------------------------------------------------------
    def _get_export_queryset(self, request):
        """
        Renvoie le queryset pour l'export Excel, incluant tous les filtres list, recherche ;
        additions :
            - ids=... limite à certains appairages (GET/POST)
            - avec_archivees=... désarchivage (POST/GET)
        - Les droits d’accès de l’utilisateur sont appliqués en scoping préalable.
        - Optimisations N+1 :
            * select_related sur candidat/formation/partenaire/créateurs pour limiter les requêtes par ligne.
            * prefetch_related("commentaires", "commentaires__created_by") pour sérialiser les commentaires
            et leurs auteurs sans requêtes supplémentaires par commentaire lors de la boucle d’export.
        """
        qs = self.filter_queryset(self.get_queryset())

        avec_archivees = request.data.get("avec_archivees") or request.query_params.get("avec_archivees")
        if not (avec_archivees and str(avec_archivees).lower() in ["1", "true", "yes", "on"]):
            qs = qs.exclude(activite=AppairageActivite.ARCHIVE)

        ids = request.data.get("ids") or request.query_params.get("ids")
        if ids:
            if isinstance(ids, str):
                ids = [int(x) for x in ids.split(",") if x.isdigit()]
            elif isinstance(ids, list):
                ids = [int(x) for x in ids if str(x).isdigit()]
            qs = qs.filter(id__in=ids)

        return qs.select_related(
            "candidat",
            "candidat__formation",
            "candidat__formation__centre",
            "formation",
            "formation__centre",
            "partenaire",
            "created_by",
            "updated_by",
        )

    @extend_schema(
        summary="Exporter les appairages (Excel)",
        description="Exporte les appairages filtrés au format Excel (.xlsx).",
        responses={
            200: OpenApiResponse(
                description="Fichier Excel généré avec succès.",
                response=OpenApiTypes.BINARY,
                examples=None,
            )
        },
    )
    @action(detail=False, methods=["get", "post"], url_path="export-xlsx")
    def export_xlsx(self, request):
        """
        [GET|POST] /appairages/export-xlsx/
        - Export des appairages filtrés selon les critères de list, scoping user, etc.
        - Filtres additionnels possibles : ids=... (GET ou POST) + avec_archivees
        - Permissions : staff/admin uniquement
        - Réponse : fichier Excel (binaire)
        """
        import mimetypes

        # Fix openpyxl / mimetypes pour certains environnements où .webp n'est pas enregistré
        mimetypes.add_type("image/webp", ".webp")

        qs = self._get_export_queryset(request)
        wb = Workbook()
        ws = wb.active
        ws.title = "Appairages"

        # ==========================================================
        # 🖼️ Ajout logo, titres, styles, etc. (format Excel)
        # ==========================================================
        try:
            logo_path = Path(settings.BASE_DIR) / "rap_app/static/images/logo.png"
            if logo_path.exists():
                with PILImage.open(logo_path) as pil_img:
                    if pil_img.mode not in ("RGB", "RGBA"):
                        pil_img = pil_img.convert("RGBA")

                    png_buffer = BytesIO()
                    pil_img.save(png_buffer, format="PNG")
                    png_buffer.seek(0)

                    img = XLImage(png_buffer)
                    img.height = 60
                    img.width = 60
                    ws.add_image(img, "A1")
        except Exception:
            pass

        ws.merge_cells("B1:Z1")
        ws["B1"] = "Export des appairages — Rap_App"
        ws["B1"].font = Font(bold=True, size=14, color="004C99")
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("B2:Z2")
        ws["B2"] = f"Export réalisé le {dj_timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        ws["B2"].font = Font(italic=True, size=10, color="666666")
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.append([])

        headers = [
            "Activité (code)",
            "Date appairage",
            "Statut (code)",
            "Candidat",
            "Partenaire",
            "Contact",
            "Email",
            "Téléphone",
            "Formation",
            "Centre",
            "Type d’offre",
            "N° Offre",
            "Statut formation",
            "Places totales",
            "Places disponibles",
            "Date début",
            "Date fin",
            "Retour partenaire",
            "Date retour",
            "Créé par (nom)",
            "Créé le",
            "Maj par (nom)",
            "Maj le",
            "Dernier commentaire",
            "Commentaires",
        ]
        ws.append(headers)

        header_row = ws.max_row
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        for cell in ws[header_row]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            cell.fill = header_fill
            cell.border = border
        ws.row_dimensions[header_row].height = 28

        even_fill = PatternFill("solid", fgColor="F7FBFF")
        odd_fill = PatternFill("solid", fgColor="FFFFFF")

        def _safe(obj, path, default=""):
            try:
                for part in path.split("."):
                    obj = getattr(obj, part)
                    if obj is None:
                        return default
                return obj
            except Exception:
                return default

        def _to_text(v):
            if v is None:
                return ""
            if callable(v):
                try:
                    v = v()
                except Exception:
                    v = str(v)
            if isinstance(v, (int, float, str, bool)):
                return str(v)
            if isinstance(v, (list, tuple, set)):
                return "\n".join(map(_to_text, v))
            if isinstance(v, dict):
                return "; ".join(f"{k}: {v}" for k, v in v.items())
            if hasattr(v, "strftime"):
                try:
                    return v.strftime("%d/%m/%Y %H:%M")
                except Exception:
                    return str(v)
            if hasattr(v, "nom_complet"):
                return v.nom_complet
            if hasattr(v, "nom"):
                return v.nom
            return str(v)

        def _compute_places_disponibles(a):
            f = getattr(a, "formation", None)
            if not f:
                return ""
            inscrits_total = (getattr(f, "inscrits_crif", 0) or 0) + (getattr(f, "inscrits_mp", 0) or 0)
            prevus_total = (getattr(f, "prevus_crif", 0) or 0) + (getattr(f, "prevus_mp", 0) or 0)
            cap = getattr(f, "cap", None)
            if cap is not None:
                return max(int(cap) - int(inscrits_total), 0)
            if prevus_total:
                return max(int(prevus_total) - int(inscrits_total), 0)
            return ""

        for i, a in enumerate(qs, start=1):
            commentaires_text = ""
            try:
                prefetched = getattr(a, "_prefetched_objects_cache", {}).get("commentaires")
                if prefetched is not None:
                    coms = sorted(
                        prefetched,
                        key=lambda c: (
                            getattr(c, "created_at", None),
                            getattr(c, "pk", None),
                        ),
                        reverse=True,
                    )
                elif hasattr(a, "commentaires"):
                    coms = a.commentaires.select_related("created_by").order_by("-created_at", "-pk")
                else:
                    coms = []
                if coms:
                    commentaires_text = "\n".join(
                        f"- {getattr(c.created_by, 'get_full_name', lambda: str(c.created_by))()}: {getattr(c, 'body', '')}"
                        for c in coms
                    )
            except Exception:
                commentaires_text = ""

            row = [
                _to_text(_safe(a, "activite")),
                a.date_appairage.strftime("%d/%m/%Y") if _safe(a, "date_appairage") else "",
                _to_text(_safe(a, "statut")),
                _to_text(_safe(a, "candidat.nom_complet")) or _to_text(_safe(a, "candidat")),
                _to_text(_safe(a, "partenaire.nom")),
                _to_text(_safe(a, "partenaire.contact_nom")),
                _to_text(_safe(a, "partenaire.contact_email")),
                _to_text(_safe(a, "partenaire.contact_telephone")),
                _to_text(_safe(a, "formation.nom")),
                _to_text(_safe(a, "formation.centre.nom")),
                _to_text(_safe(a, "formation.type_offre.nom")),
                _to_text(_safe(a, "formation.num_offre")),
                _to_text(_safe(a, "formation.statut")),
                _to_text(_safe(a, "formation.cap")),
                _to_text(_compute_places_disponibles(a)),
                _to_text(_safe(a, "formation.start_date")),
                _to_text(_safe(a, "formation.end_date")),
                _to_text(getattr(a, "retour_partenaire", "")),
                a.date_retour.strftime("%d/%m/%Y") if _safe(a, "date_retour") else "",
                _to_text(_safe(a, "created_by.get_full_name")),
                a.created_at.strftime("%d/%m/%Y %H:%M") if _safe(a, "created_at") else "",
                _to_text(_safe(a, "updated_by.get_full_name")),
                a.updated_at.strftime("%d/%m/%Y %H:%M") if _safe(a, "updated_at") else "",
                _to_text(getattr(a, "last_commentaire", "")),
                commentaires_text,
            ]
            ws.append(row)

            fill = even_fill if i % 2 == 0 else odd_fill
            for j, cell in enumerate(ws[ws.max_row], start=1):
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrapText=True)
                val = str(cell.value).strip().lower() if cell.value else ""

                if j == 3:
                    if "actif" in val:
                        cell.font = Font(color="008000", bold=True)
                    elif "inactif" in val or "refus" in val:
                        cell.font = Font(color="C00000", bold=True)
                    elif "archive" in val:
                        cell.font = Font(color="7F7F7F", italic=True)
                    else:
                        cell.font = Font(color="1F4E78")
                elif j == 14:
                    cell.font = Font(color="1F4E78", bold=True)
                elif j == 15:
                    try:
                        num = int(float(val.replace(",", ".").strip()))
                    except Exception:
                        num = None
                    if num is None:
                        cell.font = Font(color="000000")
                    elif num == 0:
                        cell.font = Font(color="006100", bold=True)
                    elif num <= 4:
                        cell.font = Font(color="E46C0A", bold=True)
                    elif num <= 9:
                        cell.font = Font(color="1F4E78", bold=True)
                    else:
                        cell.font = Font(color="C00000", bold=True)
                elif j in [18, 19]:
                    cell.font = Font(color="548235")
                elif j in [20, 22]:
                    cell.font = Font(color="7030A0")
                else:
                    cell.font = Font(color="000000")
            ws.row_dimensions[ws.max_row].height = 26

        end_row = ws.max_row
        last_col_letter = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{end_row}"
        ws.freeze_panes = f"A{header_row + 1}"

        for col_cells in ws.columns:
            length = max(len(str(c.value)) if c.value else 0 for c in col_cells)
            column_letter = get_column_letter(col_cells[0].column)
            adjusted_width = min(length + 3, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        ws.oddFooter.center.text = f"© Rap_App — export du {dj_timezone.now().strftime('%d/%m/%Y %H:%M')}"

        buffer = BytesIO()
        wb.save(buffer)
        binary = buffer.getvalue()
        buffer.close()

        filename = f'appairages_{dj_timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            binary,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = len(binary)
        return response
